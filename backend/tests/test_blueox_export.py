"""Contract tests for the Blue Ox curve-drop workbook.

Pure-builder tests (no DB): construct ``BlueOxExportData`` directly and
assert the workbook shape + the acceptance-gate rules from the contract
(engineering_db docs/blue_ox_curve_drop_contract.md) — complete
triplets, ascending percentile monotonicity, Block B reconciliation
computed from the delivered sheets, naming/lateral bounds, history
tie-out. One assembly-level test pins the SPE->ascending flip end to
end (our p90 fit must feed the file's _p10 columns).
"""

from __future__ import annotations

import io
import uuid
from datetime import UTC, date, datetime
from typing import Any

import pytest
from openpyxl import load_workbook

from app.exports.blueox import (
    LEVEL_TO_SPE_KEY,
    NGL_BASIS,
    BlueOxContractError,
    BlueOxExportData,
    DsuMetaRow,
    InventoryRow,
    NoviComparisonZone,
    ZoneData,
    blueox_filename,
    build_blueox_workbook,
    monthly_volumes_from_rates,
)

_MONTHS = 24
_ANALOG_HEADERS = (
    "api10",
    "Wellname",
    "Operator",
    "Formation",
    "Lateral Length",
)

# Ascending Blue Ox convention: p10 < base < p90.
_LEVEL_SCALE = {"P10": 0.5, "P25": 0.75, "P50": 1.0, "P75": 1.25, "P90": 1.5}


def _volumes(levels: tuple[str, ...], with_water: bool = True) -> dict[str, dict[str, list[float]]]:
    base_oil = [1000.0 - 30.0 * i for i in range(_MONTHS)]
    base_gas = [5000.0 - 100.0 * i for i in range(_MONTHS)]
    out: dict[str, dict[str, list[float]]] = {}
    for lv in ("P50", *levels):
        scale = _LEVEL_SCALE[lv]
        out[lv] = {
            "oil": [v * scale for v in base_oil],
            "gas": [v * scale for v in base_gas],
        }
    if with_water:
        out["P50"]["water"] = [2000.0 - 50.0 * i for i in range(_MONTHS)]
    return out


def _params(levels: tuple[str, ...]) -> list[dict[str, Any]]:
    rows = []
    for stream in ("oil", "gas"):
        for lv in ("P50", *levels):
            rows.append(
                {
                    "stream": stream,
                    "level": lv,
                    "qi": 250.0 * _LEVEL_SCALE[lv],
                    "qi_units": "bbl/d per 1,000 ft" if stream == "oil" else "Mcf/d per 1,000 ft",
                    "b_factor": 1.1,
                    "di": 2.5,
                    "dmin": 0.08,
                    "notes": "peak_ramp alignment; qi at peak month 4",
                }
            )
    return rows


def _zone(
    name: str,
    apis: tuple[str, ...],
    levels: tuple[str, ...] = ("P10", "P90"),
    n_inventory: int = 3,
) -> ZoneData:
    return ZoneData(
        zone_name=name,
        reserve_category="PUD",
        normalization_basis="per_1000_lateral_ft",
        volumes=_volumes(levels),
        curve_params=_params(levels),
        analog_headers=_ANALOG_HEADERS,
        analog_rows=[
            (api, f"WELL {i} 123H", "OPERATOR X", "WOLFCAMP A", 9800.0 + i)
            for i, api in enumerate(apis)
        ],
        inventory=[
            InventoryRow(
                producing_lateral_ft=10_000.0,
                drilled_lateral_ft=12_500.0 + 100.0 * i,
                well_name=f"PLANNED {i + 1}",
            )
            for i in range(n_inventory)
        ],
    )


def _production_rows(apis: tuple[str, ...]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for api in sorted(apis):
        for m in (1, 2, 3):
            rows.append([api, f"2025-0{m}", 30_000.0 / m, 90_000.0 / m, 40_000.0 / m, 30.0])
    return rows


def _data(
    zones: list[ZoneData] | None = None,
    levels: tuple[str, ...] = ("P10", "P90"),
    codename: str = "holdtheline",
    production_rows: list[list[Any]] | None = None,
    history_exceptions: tuple[str, ...] = (),
) -> BlueOxExportData:
    if zones is None:
        zones = [
            _zone("WOLFCAMP A", ("4200000001", "4200000002"), levels),
            _zone("THIRD BONE SPRING", ("4200000003",), levels),
        ]
    apis = tuple(
        str(r[0]) for z in zones for r in z.analog_rows if str(r[0]) not in history_exceptions
    )
    return BlueOxExportData(
        codename=codename,
        export_date=date(2026, 7, 20),
        curve_months=_MONTHS,
        levels=levels,
        prepared_by="M. Mast",
        source_system="anduin type-curve DB (synced from oilgas warehouse)",
        governing_export="deal holdTheLine — holdtheline_curves_2026-07-20.xlsx",
        curve_params_source="wolfcamp_a_v3 saved 2026-07-18",
        zones=zones,
        production_headers=("api10", "date", "oil_bbl", "gas_mcf", "water_bbl", "days_on"),
        production_rows=production_rows if production_rows is not None else _production_rows(apis),
        production_history_through="2025-03",
        history_exceptions=history_exceptions,
    )


# ============================ shape ============================


def test_sheet_set_and_order() -> None:
    wb = load_workbook(io.BytesIO(build_blueox_workbook(_data())))
    assert wb.sheetnames == [
        "WOLFCAMP A",
        "THIRD BONE SPRING",
        "meta",
        "inventory",
        "WOLFCAMP A meta",
        "THIRD BONE SPRING meta",
        "analog_production",
        "curve_params",
        "manifest",
    ]


def test_zone_sheet_shape_and_ngl_zero() -> None:
    wb = load_workbook(io.BytesIO(build_blueox_workbook(_data())))
    ws = wb["WOLFCAMP A"]
    header = [c.value for c in ws[1]]
    # No month column; exact lowercase headers; water after the base
    # triplet; percentile triplets complete (ngl rides along all-zero).
    assert header == [
        "oil_bbl",
        "gas_mcf",
        "ngl_bbl",
        "water_bbl",
        "oil_bbl_p10",
        "gas_mcf_p10",
        "ngl_bbl_p10",
        "oil_bbl_p90",
        "gas_mcf_p90",
        "ngl_bbl_p90",
    ]
    assert ws.max_row == _MONTHS + 1
    ngl_cols = [i + 1 for i, h in enumerate(header) if "ngl" in str(h)]
    for col in ngl_cols:
        vals = [ws.cell(row=r, column=col).value for r in range(2, ws.max_row + 1)]
        assert all(v == 0.0 for v in vals)
    # Numbers as numbers (openpyxl reads whole floats back as int —
    # both satisfy the contract's numbers-not-text rule).
    assert isinstance(ws.cell(row=2, column=1).value, (int, float))


def test_analog_sheet_marker_and_single_api_column() -> None:
    wb = load_workbook(io.BytesIO(build_blueox_workbook(_data())))
    ws = wb["WOLFCAMP A meta"]
    assert ws.cell(row=1, column=1).value == "per_well_summary"
    headers = [c.value for c in ws[2]]
    api_cols = [h for h in headers if h and "api" in str(h).lower()]
    assert api_cols == ["api10"]


def test_meta_sheet_rows() -> None:
    wb = load_workbook(io.BytesIO(build_blueox_workbook(_data())))
    rows = list(wb["meta"].iter_rows(values_only=True))
    assert rows[0] == ("area", "normalization_basis", "reserve_category")
    assert rows[1] == ("WOLFCAMP A", "per_1000_lateral_ft", "PUD")


# ============================ manifest ============================


def _manifest_kv(wb: Any) -> dict[str, Any]:
    rows = list(wb["manifest"].iter_rows(values_only=True))
    return {r[0]: r[1] for r in rows if r and r[0] is not None and r[1] is not None}


def test_manifest_block_a_declarations() -> None:
    wb = load_workbook(io.BytesIO(build_blueox_workbook(_data())))
    kv = _manifest_kv(wb)
    assert kv["percentile_orientation"] == "ascending"
    assert kv["gas_basis"] == "wellhead_unshrunk"
    assert kv["risking"] == "unrisked"
    # 2026-07-20 amendment: Blue Ox applies their own yield.
    assert kv["ngl_basis"] == NGL_BASIS == "derived_by_blue_ox_via_yield"
    assert kv["curve_months"] == _MONTHS
    assert kv["production_history_through"] == "2025-03"
    assert kv["deal_codename"] == "holdtheline"


def test_manifest_block_b_ties_to_delivered_sheets() -> None:
    wb = load_workbook(io.BytesIO(build_blueox_workbook(_data())))
    rows = list(wb["manifest"].iter_rows(values_only=True))
    header_idx = next(i for i, r in enumerate(rows) if r and r[0] == "area")
    header = rows[header_idx]
    by_zone = {
        r[0]: dict(zip(header, r, strict=False)) for r in rows[header_idx + 1 :] if r and r[0]
    }

    ws = wb["WOLFCAMP A"]
    zone_header = [c.value for c in ws[1]]
    oil_col = zone_header.index("oil_bbl") + 1
    oil_sum = sum(float(ws.cell(row=r, column=oil_col).value) for r in range(2, ws.max_row + 1))
    block_b = by_zone["WOLFCAMP A"]
    # ±0.1% is the acceptance gate; delivered-sheet arithmetic should
    # tie essentially exactly.
    assert abs(float(block_b["eur_oil_bbl"]) - oil_sum) < 1e-6 * max(oil_sum, 1.0)
    assert block_b["eur_ngl_bbl"] == 0.0
    assert block_b["gross_locations"] == 3
    inv_ws = wb["inventory"]
    drilled = [float(r[3].value) for r in inv_ws.iter_rows(min_row=2) if r[0].value == "WOLFCAMP A"]
    assert len(drilled) == 3
    assert abs(float(block_b["avg_drilled_lateral_ft"]) - sum(drilled) / 3) < 1e-9


# ============================ gates ============================


def test_percentile_monotonicity_enforced() -> None:
    zones = [_zone("WOLFCAMP A", ("4200000001",))]
    vols = dict(zones[0].volumes)
    # Sabotage: hand the P10 (low case) the HIGH volumes — the classic
    # un-flipped SPE export.
    vols["P10"], vols["P90"] = vols["P90"], vols["P10"]
    bad = ZoneData(**{**zones[0].__dict__, "volumes": vols})
    with pytest.raises(BlueOxContractError, match="not strictly ascending"):
        build_blueox_workbook(_data(zones=[bad]))


def test_partial_triplet_is_hard_failure() -> None:
    zones = [_zone("WOLFCAMP A", ("4200000001",))]
    vols = {lv: dict(by) for lv, by in zones[0].volumes.items()}
    del vols["P10"]["gas"]
    bad = ZoneData(**{**zones[0].__dict__, "volumes": vols})
    with pytest.raises(BlueOxContractError, match="partial triplet"):
        build_blueox_workbook(_data(zones=[bad]))


def test_zone_name_rules() -> None:
    for name, fragment in (
        ("A ZONE NAME THAT RUNS MUCH TOO LONG", "exceeds 26"),
        ("manifest", "reserved"),
        ("WOLF/CAMP", "forbidden character"),
        (" WOLFCAMP A", "leading/trailing"),
    ):
        with pytest.raises(BlueOxContractError, match=fragment):
            build_blueox_workbook(_data(zones=[_zone(name, ("4200000001",))]))


def test_lateral_bounds() -> None:
    z = _zone("WOLFCAMP A", ("4200000001",))
    bad = ZoneData(
        **{
            **z.__dict__,
            "inventory": [InventoryRow(producing_lateral_ft=2_500.0, drilled_lateral_ft=12_000.0)],
        }
    )
    with pytest.raises(BlueOxContractError, match="outside 3000-25000 ft"):
        build_blueox_workbook(_data(zones=[bad]))


def test_history_tieout_both_ways() -> None:
    # A listed analog with no history rows must be declared or fail.
    with pytest.raises(BlueOxContractError, match="no production history"):
        build_blueox_workbook(
            _data(
                history_exceptions=(),
                production_rows=_production_rows(("4200000001", "4200000002")),
            )
        )
    # Declared exception passes.
    build_blueox_workbook(
        _data(
            history_exceptions=("4200000003",),
            production_rows=_production_rows(("4200000001", "4200000002")),
        )
    )
    # And the exception is surfaced in the manifest, never silent.
    wb = load_workbook(
        io.BytesIO(
            build_blueox_workbook(
                _data(
                    history_exceptions=("4200000003",),
                    production_rows=_production_rows(("4200000001", "4200000002")),
                )
            )
        )
    )
    kv = _manifest_kv(wb)
    assert "4200000003" in str(kv["analog_history_exceptions"])


def test_filename_and_reserved_stems() -> None:
    assert blueox_filename("holdtheline", date(2026, 7, 20)) == "holdtheline_curves_2026-07-20.xlsx"
    with pytest.raises(BlueOxContractError, match="reserved stem"):
        build_blueox_workbook(_data(codename="pinned_ranch"))
    # The mandated "_curves_" stem itself must not trip "type_curves"
    # when a codename ends in "type".
    build_blueox_workbook(_data(codename="prototype"))


# ============================ conventions ============================


def test_level_to_spe_flip() -> None:
    # Blue Ox ascending -> anduin SPE (migration 0021): their low case
    # is our P90 fit. If this mapping ever changes, a whole deal is
    # silently inverted — pin it.
    assert LEVEL_TO_SPE_KEY == {
        "P10": "p90",
        "P25": "p75",
        "P50": "p50",
        "P75": "p25",
        "P90": "p10",
    }


def test_monthly_volumes_trapezoid_and_zero_fill() -> None:
    d = 365.0 / 12.0
    vols = monthly_volumes_from_rates([100.0, 80.0, 60.0], 5, d)
    assert vols[0] == pytest.approx((100.0 + 80.0) / 2.0 * d)
    assert vols[1] == pytest.approx((80.0 + 60.0) / 2.0 * d)
    # Last available month flat-extrapolates; the tail zero-fills.
    assert vols[2] == pytest.approx(60.0 * d)
    assert vols[3] == 0.0 and vols[4] == 0.0


def test_inventory_category_column_pdp_display_only() -> None:
    """PDP rows appear on the inventory sheet (gunbarrel handoff) but
    are exempt from lateral bounds and invisible to Block B counts and
    lateral means; the category basis is declared in the manifest."""
    z = _zone("WOLFCAMP A", ("4200000001",))
    inv = [
        *z.inventory,
        # Existing producer: short lateral would fail planned-well
        # bounds — PDP rows are exempt.
        InventoryRow(
            producing_lateral_ft=2_000.0,
            drilled_lateral_ft=2_000.0,
            well_name="4249533594",
            category="PDP",
        ),
        InventoryRow(
            producing_lateral_ft=10_000.0,
            drilled_lateral_ft=11_000.0,
            well_name="UPSIDE 1",
            category="UPSIDE",
        ),
    ]
    zz = ZoneData(**{**z.__dict__, "inventory": inv})
    wb = load_workbook(io.BytesIO(build_blueox_workbook(_data(zones=[zz]))))

    inv_ws = wb["inventory"]
    header = [c.value for c in inv_ws[1]]
    assert header == [
        "area",
        "category",
        "producing_lateral_ft",
        "drilled_lateral_ft",
        "well_name",
    ]
    cats = [r[1] for r in inv_ws.iter_rows(min_row=2, values_only=True)]
    assert cats == ["PUD", "PUD", "PUD", "PDP", "UPSIDE"]

    rows = list(wb["manifest"].iter_rows(values_only=True))
    hi = next(i for i, r in enumerate(rows) if r and r[0] == "area")
    block_b = dict(zip(rows[hi], rows[hi + 1], strict=False))
    # 3 PUD + 1 UPSIDE counted; the producer is display-only.
    assert block_b["gross_locations"] == 4
    counted = [12_500.0, 12_600.0, 12_700.0, 11_000.0]
    assert abs(float(block_b["avg_drilled_lateral_ft"]) - sum(counted) / 4) < 1e-9
    kv = _manifest_kv(wb)
    assert "pdp_count_3mi >= 3" in str(kv["inventory_category_basis"])


def test_handoff_category_rule() -> None:
    from app.api.deals import _handoff_category
    from app.warehouse_client.narvi import NarviInventoryWell

    def w(**kw: Any) -> NarviInventoryWell:
        base: dict[str, Any] = {
            "deal_id": "d",
            "scenario_id": "s",
            "well_name": "w",
            "formation": "F",
            "completed_lateral_ft": 1.0,
            "drilled_lateral_ft": 1.0,
            "well_type": "single",
        }
        return NarviInventoryWell(**{**base, **kw})

    assert _handoff_category(w(category="pdp")) == "PDP"
    assert _handoff_category(w(pdp_count_3mi=3)) == "PUD"
    assert _handoff_category(w(pdp_count_3mi=7)) == "PUD"
    assert _handoff_category(w(pdp_count_3mi=2)) == "UPSIDE"
    assert _handoff_category(w(pdp_count_3mi=None)) == "UPSIDE"
    # narvi's user override wins over the derivation, either direction.
    assert _handoff_category(w(pdp_count_3mi=0, handoff_category="pud")) == "PUD"
    assert _handoff_category(w(pdp_count_3mi=9, handoff_category="UPSIDE")) == "UPSIDE"
    # Unrecognized override falls back to the rule.
    assert _handoff_category(w(pdp_count_3mi=9, handoff_category="bogus")) == "PUD"


def test_zone_reserve_category_pud_or_upside() -> None:
    # RES was renamed UPSIDE (aligned with the per-well handoff
    # vocabulary); the pure builder refuses anything else.
    zone = _zone("WOLFCAMP A", ("4200000001",))
    bad = ZoneData(**{**zone.__dict__, "reserve_category": "RES"})
    with pytest.raises(BlueOxContractError, match="must be PUD or UPSIDE"):
        build_blueox_workbook(_data(zones=[bad]))
    ok = ZoneData(**{**zone.__dict__, "reserve_category": "UPSIDE"})
    build_blueox_workbook(_data(zones=[ok]))


def test_zone_spec_coerces_legacy_res() -> None:
    # Configs saved under contract v1 carry "RES"; the request model
    # normalizes them so saved configs keep loading and exporting.
    from app.api.deals import BlueOxZoneSpec

    spec = BlueOxZoneSpec.model_validate(
        {
            "type_curve_id": str(uuid.uuid4()),
            "zone_name": "WCA",
            "reserve_category": "RES",
            "benches": [],
        }
    )
    assert spec.reserve_category == "UPSIDE"


def test_manifest_declares_inventory_exclusions() -> None:
    data = _data()
    data = BlueOxExportData(
        **{
            **data.__dict__,
            "inventory_exclusions": ("WDFD (4 wells)", "WCB_1 (1 wells)"),
        }
    )
    wb = load_workbook(io.BytesIO(build_blueox_workbook(data)))
    kv = _manifest_kv(wb)
    assert "WDFD (4 wells)" in str(kv["inventory_benches_excluded"])


def test_narvi_distribution_exclusions_and_unmapped() -> None:
    """Bench->zone routing: wells land in their zone, excluded benches
    are tallied for the manifest, unmapped benches hard-error, and an
    empty selection match hard-errors (typo protection)."""
    from unittest.mock import patch

    from app.api.deals import (
        BlueOxExportRequest,
        BlueOxZoneSpec,
        NarviSelection,
        _fetch_narvi_by_zone,
    )
    from app.warehouse_client.narvi import NarviInventoryWell

    def _w(
        scenario: str,
        name: str,
        bench: str,
        comp: float,
        drill: float,
        category: str = "generated",
    ) -> NarviInventoryWell:
        return NarviInventoryWell(
            deal_id="thecan_south_bs",
            scenario_id=scenario,
            well_name=name,
            formation=bench,
            completed_lateral_ft=comp,
            drilled_lateral_ft=drill,
            well_type="single",
            category=category,
        )

    wells = [
        _w("plan_a", "WCA 1H", "WCA_1", 10_000.0, 10_000.0),
        _w("plan_a", "WCA 2H", "WCA_2", 9_800.0, 9_800.0),
        _w("plan_a", "WDFD 1H", "WDFD", 10_400.0, 10_400.0),
        _w("plan_a", "MYSTERY 1H", "BS9", 10_000.0, 10_000.0),
        # Existing producer riding along as gunbarrel context — must be
        # invisible to the inventory (not mapped, not excluded-counted).
        _w("plan_a", "4249533594", "WCA_1", 4_326.0, 4_326.0, category="pdp"),
    ]
    req = BlueOxExportRequest(
        codename="x",
        prepared_by="m",
        zones=[
            BlueOxZoneSpec(
                type_curve_id=uuid.uuid4(),
                zone_name="WCA",
                reserve_category="PUD",
                benches=["WCA_1", "WCA_2"],
            )
        ],
        narvi_selections=[
            NarviSelection(deal_id="thecan_south_bs", scenario_id="plan_a"),
            NarviSelection(deal_id="thecan_south_bs", scenario_id="plan_missing"),
        ],
        exclude_benches=["WDFD"],
    )
    errors: list[str] = []
    with (
        patch("app.api.deals.get_warehouse_session", return_value=iter([None])),
        patch("app.api.deals.fetch_narvi_inventory", return_value=wells),
    ):
        by_zone, exclusions, unzoned_pdp = _fetch_narvi_by_zone(req, errors)

    # Planned wells route first; the existing producer rides along as a
    # display row for the gunbarrel automation (category PDP at the
    # InventoryRow stage), never as an error.
    assert [w.well_name for w in by_zone["WCA"]] == ["WCA 1H", "WCA 2H", "4249533594"]
    assert exclusions == ("WDFD (1 wells)",)
    assert unzoned_pdp == []  # this producer's bench IS mapped
    assert any("BS9 (1 wells)" in e for e in errors)  # unmapped planned -> error
    assert any("plan_missing matched no inventory wells" in e for e in errors)
    assert all("4249533594" not in e for e in errors)


def test_narvi_pdp_dedupe_and_duplicate_guard() -> None:
    """The same existing producer appearing in two merged scenarios
    dedupes to ONE display row; the same planned well appearing twice
    is an error (it would double the valued location count)."""
    from unittest.mock import patch

    from app.api.deals import (
        BlueOxExportRequest,
        BlueOxZoneSpec,
        NarviSelection,
        _fetch_narvi_by_zone,
    )
    from app.warehouse_client.narvi import NarviInventoryWell

    def _w(scenario: str, name: str, category: str) -> NarviInventoryWell:
        return NarviInventoryWell(
            deal_id="d",
            scenario_id=scenario,
            well_name=name,
            formation="WCA_1",
            completed_lateral_ft=10_000.0,
            drilled_lateral_ft=10_000.0,
            well_type="single",
            category=category,
        )

    wells = [
        _w("plan_all_pdp", "4249533594", "pdp"),
        _w("plan_b", "4249533594", "pdp"),  # pdp dupe across scenarios: fine
        _w("plan_b", "WCA_1-01", "generated"),
        _w("plan_c", "WCA_1-01", "generated"),  # planned dupe: error
    ]
    req = BlueOxExportRequest(
        codename="x",
        prepared_by="m",
        zones=[
            BlueOxZoneSpec(
                type_curve_id=uuid.uuid4(),
                zone_name="WCA",
                reserve_category="PUD",
                benches=["WCA_1"],
            )
        ],
        narvi_selections=[
            NarviSelection(deal_id="d", scenario_id="plan_all_pdp"),
            NarviSelection(deal_id="d", scenario_id="plan_b"),
            NarviSelection(deal_id="d", scenario_id="plan_c"),
        ],
    )
    errors: list[str] = []
    with (
        patch("app.api.deals.get_warehouse_session", return_value=iter([None])),
        patch("app.api.deals.fetch_narvi_inventory", return_value=wells),
    ):
        by_zone, _, _ = _fetch_narvi_by_zone(req, errors)

    assert any("'WCA_1-01'" in e and "plan_b" in e and "plan_c" in e for e in errors)
    assert all("4249533594" not in e for e in errors)
    # 2 generated rows (dupe flagged as an error) + the producer ONCE.
    names = [w.well_name for w in by_zone["WCA"]]
    assert names.count("4249533594") == 1
    assert names.count("WCA_1-01") == 2


def test_narvi_unzoned_pdp_never_dropped() -> None:
    """A producer whose bench maps to no zone (and one in an excluded
    bench) comes back as an unzoned PDP well — destined for the
    inventory sheet with its bench code as area — never dropped."""
    from unittest.mock import patch

    from app.api.deals import (
        BlueOxExportRequest,
        BlueOxZoneSpec,
        NarviSelection,
        _fetch_narvi_by_zone,
    )
    from app.warehouse_client.narvi import NarviInventoryWell

    def _w(name: str, bench: str, category: str) -> NarviInventoryWell:
        return NarviInventoryWell(
            deal_id="d",
            scenario_id="plan_a",
            well_name=name,
            formation=bench,
            completed_lateral_ft=4_231.0,
            drilled_lateral_ft=4_231.0,
            well_type="single",
            category=category,
        )

    wells = [
        _w("WCA_1-01", "WCA_1", "generated"),
        _w("4249533990", "WCB_1", "pdp"),  # bench has no zone -> unzoned
        _w("4249533991", "WDFD", "pdp"),  # excluded bench -> unzoned too
        _w("WDFD-01", "WDFD", "generated"),
    ]
    req = BlueOxExportRequest(
        codename="x",
        prepared_by="m",
        zones=[
            BlueOxZoneSpec(
                type_curve_id=uuid.uuid4(),
                zone_name="WCA",
                reserve_category="PUD",
                benches=["WCA_1"],
            )
        ],
        narvi_selections=[NarviSelection(deal_id="d", scenario_id="plan_a")],
        exclude_benches=["WDFD"],
    )
    errors: list[str] = []
    with (
        patch("app.api.deals.get_warehouse_session", return_value=iter([None])),
        patch("app.api.deals.fetch_narvi_inventory", return_value=wells),
    ):
        by_zone, exclusions, unzoned_pdp = _fetch_narvi_by_zone(req, errors)

    assert errors == []
    assert [w.well_name for w in by_zone["WCA"]] == ["WCA_1-01"]
    assert exclusions == ("WDFD (1 wells)",)  # the PLANNED WDFD well
    assert sorted(w.well_name for w in unzoned_pdp) == ["4249533990", "4249533991"]


def test_builder_writes_unzoned_pdp_rows_with_bench_area() -> None:
    """Unzoned PDP rows land on the inventory sheet with the bench code
    as area, invisible to Block B, exempt from lateral bounds."""
    data = _data()
    data = BlueOxExportData(
        **{
            **data.__dict__,
            "pdp_context_rows": (
                (
                    "WCB_1",
                    InventoryRow(
                        producing_lateral_ft=4_231.0,
                        drilled_lateral_ft=4_231.0,
                        well_name="4249533990",
                        category="PDP",
                    ),
                ),
            ),
        }
    )
    wb = load_workbook(io.BytesIO(build_blueox_workbook(data)))
    inv = list(wb["inventory"].iter_rows(min_row=2, values_only=True))
    pdp_row = next(r for r in inv if r[1] == "PDP")
    assert pdp_row[0] == "WCB_1"  # bench code, not a zone name
    assert pdp_row[4] == "4249533990"
    # Block B untouched: gross_locations still count only zone rows.
    rows = list(wb["manifest"].iter_rows(values_only=True))
    hi = next(i for i, r in enumerate(rows) if r and r[0] == "area")
    bb = {r[0]: r[4] for r in rows[hi + 1 :] if r and r[0]}
    assert bb["WOLFCAMP A"] == 3
    # non-PDP category in the context rows is a hard error
    bad = BlueOxExportData(
        **{
            **data.__dict__,
            "pdp_context_rows": (
                (
                    "WCB_1",
                    InventoryRow(
                        producing_lateral_ft=10_000.0,
                        drilled_lateral_ft=10_000.0,
                        well_name="X",
                        category="PUD",
                    ),
                ),
            ),
        }
    )
    with pytest.raises(BlueOxContractError, match="must be category PDP"):
        build_blueox_workbook(bad)


def test_narvi_wells_become_inventory_rows_uturn_laterals() -> None:
    """A U-turn well's completed (producing) and drilled laterals map to
    the contract's two columns without cohort-mean defaulting."""
    from unittest.mock import patch

    from app.api.deals import BlueOxZoneSpec, _collect_blueox_zone
    from app.db.models import AlignmentMethod, NormalizationBasis, TypeCurve
    from app.warehouse_client.narvi import NarviInventoryWell

    def _fit(qi: float) -> dict[str, Any]:
        return {"qi": qi, "Di": 0.0, "b": 0.0, "Df": 0.0, "qo": qi, "peak_index": 0}

    tc = TypeCurve()
    tc.id = uuid.uuid4()
    tc.name = "WCA"
    tc.included_api10s = ["4200000001"]
    tc.normalization_basis = NormalizationBasis.PER_LATERAL_FT
    tc.alignment_method = AlignmentMethod.PEAK_RAMP
    tc.created_at = datetime(2026, 7, 18, tzinfo=UTC)
    tc.series = {
        "streams": {
            "oil": {"fitted_per_percentile": {"p50": _fit(100.0)}},
            "gas": {"fitted_per_percentile": {"p50": _fit(600.0)}},
        },
    }
    spec = BlueOxZoneSpec(
        type_curve_id=tc.id,
        zone_name="WCA",
        reserve_category="PUD",
    )
    narvi_wells = [
        NarviInventoryWell(
            deal_id="thecan_44_44",
            scenario_id="plan_thecan_44",
            well_name="THECAN 44 U1",
            formation="WCA_1",
            completed_lateral_ft=8_926.0,
            drilled_lateral_ft=10_811.0,
            well_type="uturn",
        ),
    ]
    errors: list[str] = []
    with (
        patch(
            "app.api.deals.per_well_rows",
            return_value=[("4200000001", "W 1H", "OP", "WOLFCAMP A", 10_000.0)],
        ),
        patch("app.api.deals.well_geo_rows", return_value={}),
    ):
        zone = _collect_blueox_zone(
            None,
            tc,
            spec,
            ["P50"],
            12,
            errors,  # type: ignore[arg-type]
            narvi_wells=narvi_wells,
        )
    assert errors == []
    assert zone is not None
    assert zone.inventory[0].producing_lateral_ft == 8_926.0
    assert zone.inventory[0].drilled_lateral_ft == 10_811.0
    assert zone.inventory[0].well_name == "THECAN 44 U1"


def test_analog_sheet_appends_well_geo_columns() -> None:
    """The zone's analog sheet carries surface/heel/toe lat-lon appended
    after the shared 12 columns (2026-07-29 amendment); a well with no
    geometry gets blank cells, and PER_WELL_HEADERS itself is untouched
    (the deal xlsx / pptx exports stay at 12 columns)."""
    from unittest.mock import patch

    from app.api.deals import (
        BlueOxInventoryRowIn,
        BlueOxZoneSpec,
        _collect_blueox_zone,
    )
    from app.db.models import AlignmentMethod, NormalizationBasis, TypeCurve
    from app.exports.well_rows import PER_WELL_HEADERS, WELL_GEO_HEADERS

    def _fit(qi: float) -> dict[str, Any]:
        return {"qi": qi, "Di": 0.0, "b": 0.0, "Df": 0.0, "qo": qi, "peak_index": 0}

    tc = TypeCurve()
    tc.id = uuid.uuid4()
    tc.name = "WCA"
    tc.included_api10s = ["4200000001", "4200000002"]
    tc.normalization_basis = NormalizationBasis.PER_LATERAL_FT
    tc.alignment_method = AlignmentMethod.PEAK_RAMP
    tc.created_at = datetime(2026, 7, 18, tzinfo=UTC)
    tc.series = {
        "streams": {
            "oil": {"fitted_per_percentile": {"p50": _fit(100.0)}},
            "gas": {"fitted_per_percentile": {"p50": _fit(600.0)}},
        },
    }
    spec = BlueOxZoneSpec(
        type_curve_id=tc.id,
        zone_name="WCA",
        reserve_category="PUD",
        inventory=[BlueOxInventoryRowIn(drilled_lateral_ft=10_000.0)],
    )
    geo = {
        "4200000001": (
            -103.5501,
            31.9012,
            -103.5502,
            31.9101,
            -103.5503,
            31.9375,
            "LINESTRING(-103.5501 31.9012,-103.5502 31.9101,-103.55025 31.9238,-103.5503 31.9375)",
        ),
    }
    errors: list[str] = []
    with (
        patch(
            "app.api.deals.per_well_rows",
            return_value=[
                ("4200000001", "W 1H", "OP", "WOLFCAMP A", 10_000.0),
                ("4200000002", "W 2H", "OP", "WOLFCAMP A", 10_000.0),
            ],
        ),
        patch("app.api.deals.well_geo_rows", return_value=geo),
    ):
        zone = _collect_blueox_zone(
            None,
            tc,
            spec,
            ["P50"],
            12,
            errors,  # type: ignore[arg-type]
        )
    assert errors == []
    assert zone is not None
    assert zone.analog_headers == PER_WELL_HEADERS + WELL_GEO_HEADERS
    rows = {r[0]: r for r in zone.analog_rows}
    assert rows["4200000001"][-7:] == geo["4200000001"]
    assert rows["4200000002"][-7:] == (None,) * 7
    # The shared header tuple is untouched — geo rides on the drop only.
    assert "surface_lon" not in PER_WELL_HEADERS
    # lon-first pairs — the drop-wide standard (inventory heel_a_lon
    # precedent + WKT's inherent lon-lat order).
    assert zone.analog_headers[-7:-1] == (
        "surface_lon",
        "surface_lat",
        "heel_lon",
        "heel_lat",
        "toe_lon",
        "toe_lat",
    )


def test_assembly_flip_feeds_p10_columns_from_spe_p90_fit() -> None:
    """End-to-end pin of the orientation flip: a stub curve whose SPE
    p10 (high) fit is qi=200 and SPE p90 (low) fit is qi=50 must land
    qi=50-derived volumes in the file's _p10 columns."""
    from unittest.mock import patch

    from app.api.deals import BlueOxInventoryRowIn, BlueOxZoneSpec, _collect_blueox_zone
    from app.db.models import AlignmentMethod, NormalizationBasis, TypeCurve

    def _fit(qi: float) -> dict[str, Any]:
        # Constant-rate Arps (Di=0) — volumes reduce to qi * days.
        return {"qi": qi, "Di": 0.0, "b": 0.0, "Df": 0.0, "qo": qi, "peak_index": 0}

    def _stream(q_low: float, q_mid: float, q_high: float) -> dict[str, Any]:
        return {
            "fitted_per_percentile": {
                # SPE orientation as persisted: p10 = HIGH case.
                "p10": _fit(q_high),
                "p25": _fit(q_high * 0.9),
                "p50": _fit(q_mid),
                "p75": _fit(q_low * 1.1),
                "p90": _fit(q_low),
                "mean": _fit(q_mid),
            },
        }

    tc = TypeCurve()
    tc.id = uuid.uuid4()
    tc.name = "WOLFCAMP A"
    tc.included_api10s = ["4200000001"]
    tc.normalization_basis = NormalizationBasis.PER_LATERAL_FT
    tc.alignment_method = AlignmentMethod.PEAK_RAMP
    tc.created_at = datetime(2026, 7, 18, tzinfo=UTC)
    tc.series = {
        "streams": {
            "oil": _stream(50.0, 100.0, 200.0),
            "gas": _stream(300.0, 600.0, 1200.0),
        },
    }

    spec = BlueOxZoneSpec(
        type_curve_id=tc.id,
        reserve_category="PUD",
        inventory=[BlueOxInventoryRowIn(drilled_lateral_ft=10_000.0)],
    )
    errors: list[str] = []
    with (
        patch(
            "app.api.deals.per_well_rows",
            return_value=[("4200000001", "W 1H", "OP", "WOLFCAMP A", 10_000.0)],
        ),
        patch("app.api.deals.well_geo_rows", return_value={}),
    ):
        zone = _collect_blueox_zone(
            None,
            tc,
            spec,
            ["P10", "P50", "P90"],
            12,
            errors,  # type: ignore[arg-type]
        )
    assert errors == []
    assert zone is not None

    d = 365.0 / 12.0
    # Constant-rate: every month's volume = qi * days_per_month.
    assert zone.volumes["P10"]["oil"][0] == pytest.approx(50.0 * d)
    assert zone.volumes["P50"]["oil"][0] == pytest.approx(100.0 * d)
    assert zone.volumes["P90"]["oil"][0] == pytest.approx(200.0 * d)
    # curve_params quotes the SAME flipped fit (level P10 row = qi 50).
    p10_oil = next(r for r in zone.curve_params if r["stream"] == "oil" and r["level"] == "P10")
    assert p10_oil["qi"] == 50.0
    assert p10_oil["qi_units"] == "bbl/d per 1,000 ft"
    # producing_lateral_ft defaulted to the cohort mean.
    assert zone.inventory[0].producing_lateral_ft == 10_000.0


def test_blueox_config_roundtrip_and_staleness() -> None:
    """The saved config is the JSON dump of the export request model
    (guaranteeing config and ad-hoc payloads can't drift), and the
    staleness probe flags a narvi scenario re-saved after pinning."""
    from unittest.mock import patch

    from app.api.deals import (
        BlueOxExportRequest,
        BlueOxZoneSpec,
        NarviSelection,
        _scenario_status,
    )

    cfg = BlueOxExportRequest(
        codename="thecan",
        prepared_by="m",
        zones=[
            BlueOxZoneSpec(
                type_curve_id=uuid.uuid4(),
                zone_name="WCA",
                reserve_category="PUD",
                benches=["WCA_1", "WCA_2"],
            )
        ],
        narvi_selections=[
            NarviSelection(
                deal_id="thecan_south_bs",
                scenario_id="plan_a",
                pinned_updated_at="2026-07-22T15:00:00+00:00",
            )
        ],
    )
    assert BlueOxExportRequest.model_validate(cfg.model_dump(mode="json")) == cfg

    def probe(current: str | None) -> list[dict]:
        res = {("thecan_south_bs", "plan_a"): current} if current else {}
        with (
            patch("app.api.deals.get_warehouse_session", return_value=iter([None])),
            patch("app.api.deals.fetch_scenario_updated_at", return_value=res),
        ):
            return _scenario_status(cfg.narvi_selections)

    assert probe("2026-07-22T16:00:00+00:00")[0]["stale"] is True  # re-saved since pin
    assert probe("2026-07-22T15:00:00+00:00")[0]["stale"] is False  # unchanged
    assert probe(None)[0]["stale"] is False  # unreachable -> usable


# ==================== scenario-scoped zones ====================


def _scoped_req(zones: list[Any], selections: list[str]) -> Any:
    from app.api.deals import BlueOxExportRequest, NarviSelection

    return BlueOxExportRequest(
        codename="alchemist",
        prepared_by="m",
        zones=zones,
        narvi_selections=[NarviSelection(deal_id="alch", scenario_id=s) for s in selections],
    )


def _scoped_zone(name: str, benches: list[str], scope: list[str] | None) -> Any:
    from app.api.deals import BlueOxZoneSpec, ScenarioRef

    return BlueOxZoneSpec(
        type_curve_id=uuid.uuid4(),
        zone_name=name,
        reserve_category="PUD",
        benches=benches,
        scenario_scope=(
            None if scope is None else [ScenarioRef(deal_id="alch", scenario_id=s) for s in scope]
        ),
    )


def _alch_well(scenario: str, name: str, bench: str) -> Any:
    from app.warehouse_client.narvi import NarviInventoryWell

    return NarviInventoryWell(
        deal_id="alch",
        scenario_id=scenario,
        well_name=name,
        formation=bench,
        completed_lateral_ft=10_000.0,
        drilled_lateral_ft=10_000.0,
        well_type="single",
        category="generated",
    )


def test_scenario_scope_routes_same_bench_to_two_zones() -> None:
    """The alchemist case: the same bench (BS1_S) carries a WEST curve
    for the western DSU scenarios and an EAST curve for the SE DSU —
    disjoint scopes coexist, wells route by their scenario, and an
    unscoped second bench still routes everything."""
    from unittest.mock import patch

    from app.api.deals import _fetch_narvi_by_zone

    wells = [
        _alch_well("plan_west_a", "W A BS1_S-01", "BS1_S"),
        _alch_well("plan_west_b", "W B BS1_S-01", "BS1_S"),
        _alch_well("plan_east", "E BS1_S-01", "BS1_S"),
        _alch_well("plan_east", "E WCA_1-01", "WCA_1"),
    ]
    req = _scoped_req(
        zones=[
            _scoped_zone("bs1_s west", ["BS1_S"], ["plan_west_a", "plan_west_b"]),
            _scoped_zone("bs1_s east", ["BS1_S"], ["plan_east"]),
            _scoped_zone("wca", ["WCA_1"], None),  # unscoped rides along
        ],
        selections=["plan_west_a", "plan_west_b", "plan_east"],
    )
    errors: list[str] = []
    with (
        patch("app.api.deals.get_warehouse_session", return_value=iter([None])),
        patch("app.api.deals.fetch_narvi_inventory", return_value=wells),
    ):
        by_zone, _, _ = _fetch_narvi_by_zone(req, errors)

    assert errors == []
    assert [w.well_name for w in by_zone["bs1_s west"]] == [
        "W A BS1_S-01",
        "W B BS1_S-01",
    ]
    assert [w.well_name for w in by_zone["bs1_s east"]] == ["E BS1_S-01"]
    assert [w.well_name for w in by_zone["wca"]] == ["E WCA_1-01"]


def test_scenario_scope_overlap_and_coverage_errors() -> None:
    """Overlapping scopes on one bench error (incl. unscoped-vs-scoped);
    a well whose bench is claimed but whose scenario no zone covers is a
    hard error, never a silent drop; a scope naming an unselected
    scenario errors."""
    from unittest.mock import patch

    from app.api.deals import _fetch_narvi_by_zone

    wells = [
        _alch_well("plan_west_a", "W BS1_S-01", "BS1_S"),
        _alch_well("plan_east", "E BS1_S-01", "BS1_S"),
    ]

    # (a) unscoped + scoped on the same bench = overlap error
    req = _scoped_req(
        zones=[
            _scoped_zone("bs1_s west", ["BS1_S"], ["plan_west_a"]),
            _scoped_zone("bs1_s all", ["BS1_S"], None),
        ],
        selections=["plan_west_a", "plan_east"],
    )
    errors: list[str] = []
    with (
        patch("app.api.deals.get_warehouse_session", return_value=iter([None])),
        patch("app.api.deals.fetch_narvi_inventory", return_value=wells),
    ):
        _fetch_narvi_by_zone(req, errors)
    assert any("overlapping scenario scope" in e for e in errors)

    # (b) scope gap: east scenario's BS1_S wells covered by NO zone
    req = _scoped_req(
        zones=[_scoped_zone("bs1_s west", ["BS1_S"], ["plan_west_a"])],
        selections=["plan_west_a", "plan_east"],
    )
    errors = []
    with (
        patch("app.api.deals.get_warehouse_session", return_value=iter([None])),
        patch("app.api.deals.fetch_narvi_inventory", return_value=wells),
    ):
        by_zone, _, _ = _fetch_narvi_by_zone(req, errors)
    assert any(
        "scenario_scope covers" in e and "BS1_S in alch/plan_east (1 wells)" in e for e in errors
    )
    assert [w.well_name for w in by_zone["bs1_s west"]] == ["W BS1_S-01"]

    # (c) scope naming an unselected scenario
    req = _scoped_req(
        zones=[_scoped_zone("bs1_s west", ["BS1_S"], ["plan_typo"])],
        selections=["plan_west_a"],
    )
    errors = []
    with (
        patch("app.api.deals.get_warehouse_session", return_value=iter([None])),
        patch("app.api.deals.fetch_narvi_inventory", return_value=wells[:1]),
    ):
        _fetch_narvi_by_zone(req, errors)
    assert any("not in this drop's selections" in e and "alch/plan_typo" in e for e in errors)


def test_scenario_scope_legacy_config_and_roundtrip() -> None:
    """Configs saved before scenario_scope validate untouched (None =
    all scenarios), and a scoped config survives dump -> validate."""
    from app.api.deals import BlueOxExportRequest

    legacy = {
        "codename": "thecan",
        "prepared_by": "m",
        "zones": [
            {
                "type_curve_id": str(uuid.uuid4()),
                "zone_name": "WCA",
                "reserve_category": "PUD",
                "benches": ["WCA_1"],
            }
        ],
        "narvi_selections": [],
        "exclude_benches": [],
    }
    cfg = BlueOxExportRequest.model_validate(legacy)
    assert cfg.zones[0].scenario_scope is None
    assert cfg.zones[0].scope_pairs() is None

    scoped = _scoped_req(
        zones=[_scoped_zone("bs1_s west", ["BS1_S"], ["plan_west_a"])],
        selections=["plan_west_a"],
    )
    assert BlueOxExportRequest.model_validate(scoped.model_dump(mode="json")) == scoped


def test_manifest_declares_scoped_zones_only() -> None:
    """Scoped zones declare their DSU subset in Block A; unscoped zones
    add nothing (legacy drops byte-identical)."""
    zone = _zone("WOLFCAMP A", ("4200000001",))
    scoped_zone = type(zone)(
        **{
            **zone.__dict__,
            "zone_name": "BS1S WEST",
            "scenario_scope": ("alch/plan_west_a", "alch/plan_west_b"),
        }
    )
    data = _data(zones=[zone, scoped_zone])
    wb = load_workbook(io.BytesIO(build_blueox_workbook(data)))
    kv = {
        r[0]: r[1]
        for r in wb["manifest"].iter_rows(values_only=True)
        if r and r[0] is not None and r[1] is not None
    }
    assert kv["zone_scenario_scope[BS1S WEST]"] == "alch/plan_west_a; alch/plan_west_b"
    assert "zone_scenario_scope[WOLFCAMP A]" not in kv


# ================= inventory gunbarrel geometry (2026-07-27) =================


def _geo_inv(
    name: str,
    category: str = "PUD",
    dsu: str = "1_4_9/plan_a",
    xs: tuple[float, ...] = (-660.0,),
    legs: int = 1,
) -> InventoryRow:
    return InventoryRow(
        producing_lateral_ft=10_000.0,
        drilled_lateral_ft=12_000.0,
        well_name=name,
        category=category,
        dsu_id=dsu,
        bench="WCA_1",
        landing_tvd_ft=11_480.0,
        gunbarrel_offset_ft=xs[0],
        gunbarrel_offset_b_ft=xs[1] if len(xs) > 1 else None,
        lateral_azimuth_deg=105.3,
        heel_a_lon=-103.81,
        heel_a_lat=31.92,
        toe_a_lon=-103.78,
        toe_a_lat=31.92,
        heel_b_lon=-103.81 if legs > 1 else None,
        heel_b_lat=31.921 if legs > 1 else None,
        toe_b_lon=-103.78 if legs > 1 else None,
        toe_b_lat=31.921 if legs > 1 else None,
    )


def _dsu_frame(dsu: str = "1_4_9/plan_a") -> DsuMetaRow:
    return DsuMetaRow(
        dsu_id=dsu,
        azimuth_deg=105.3,
        origin_lon=-103.795,
        origin_lat=31.9205,
    )


def test_inventory_geometry_columns_and_dsu_meta() -> None:
    """Geometry columns append strictly AFTER the v1 columns; the U-turn
    second leg fills the _b cells; dsu_meta carries the projection frame."""
    z = _zone("WOLFCAMP A", ("4200000001",))
    inv = [
        _geo_inv("PLANNED 1"),
        _geo_inv("PLANNED 2 UTURN", xs=(-660.0, 660.0), legs=2),
        _geo_inv("4249533594", category="PDP"),
    ]
    zz = ZoneData(**{**z.__dict__, "inventory": inv})
    data = _data(zones=[zz])
    data = BlueOxExportData(**{**data.__dict__, "dsu_meta": (_dsu_frame(),)})
    wb = load_workbook(io.BytesIO(build_blueox_workbook(data)))

    ws = wb["inventory"]
    header = [c.value for c in ws[1]]
    assert header == [
        "area",
        "category",
        "producing_lateral_ft",
        "drilled_lateral_ft",
        "well_name",
        "dsu_id",
        "bench",
        "landing_tvd_ft",
        "gunbarrel_offset_ft",
        "gunbarrel_offset_b_ft",
        "lateral_azimuth_deg",
        "heel_a_lon",
        "heel_a_lat",
        "toe_a_lon",
        "toe_a_lat",
        "heel_b_lon",
        "heel_b_lat",
        "toe_b_lon",
        "toe_b_lat",
    ]
    rows = {r[4]: r for r in ws.iter_rows(min_row=2, values_only=True)}
    single = rows["PLANNED 1"]
    uturn = rows["PLANNED 2 UTURN"]
    col = {h: i for i, h in enumerate(header)}
    assert single[col["gunbarrel_offset_ft"]] == -660.0
    assert single[col["gunbarrel_offset_b_ft"]] is None
    assert single[col["heel_b_lon"]] is None
    assert uturn[col["gunbarrel_offset_b_ft"]] == 660.0
    assert uturn[col["toe_b_lat"]] == 31.921
    assert single[col["dsu_id"]] == "1_4_9/plan_a"
    assert single[col["landing_tvd_ft"]] == 11_480.0

    # dsu_meta sheet, right after inventory.
    names = wb.sheetnames
    assert names.index("dsu_meta") == names.index("inventory") + 1
    frames = list(wb["dsu_meta"].iter_rows(values_only=True))
    assert frames[0] == ("dsu_id", "azimuth_deg", "origin_lon", "origin_lat")
    assert frames[1] == ("1_4_9/plan_a", 105.3, -103.795, 31.9205)


def test_inventory_without_geometry_is_byte_stable_shape() -> None:
    """Legacy/manual inventory (no dsu_id anywhere) keeps the v1 header
    and writes no dsu_meta sheet."""
    wb = load_workbook(io.BytesIO(build_blueox_workbook(_data())))
    header = [c.value for c in wb["inventory"][1]]
    assert header == [
        "area",
        "category",
        "producing_lateral_ft",
        "drilled_lateral_ft",
        "well_name",
    ]
    assert "dsu_meta" not in wb.sheetnames


def test_dsu_meta_two_way_tie_enforced() -> None:
    z = _zone("WOLFCAMP A", ("4200000001",))
    zz = ZoneData(**{**z.__dict__, "inventory": [_geo_inv("PLANNED 1")]})

    # Referenced DSU with no frame -> hard error.
    data = _data(zones=[zz])
    with pytest.raises(BlueOxContractError, match="no dsu_meta frame"):
        build_blueox_workbook(data)

    # Frame with no referencing rows -> hard error too.
    data = _data()
    data = BlueOxExportData(**{**data.__dict__, "dsu_meta": (_dsu_frame(),)})
    with pytest.raises(BlueOxContractError, match="no inventory rows"):
        build_blueox_workbook(data)


# ================= TC-vs-Novi comparison sheets (2026-07-27) =================


def _novi_zone(name: str, n_sticks: int = 4, months: int = 12) -> NoviComparisonZone:
    series = tuple(100.0 - i for i in range(months)) if n_sticks else ()
    return NoviComparisonZone(
        zone_name=name,
        n_sticks=n_sticks,
        n_self=1 if n_sticks else 0,
        n_neighborhood=max(n_sticks - 1, 0),
        n_pud=n_sticks,
        n_res=0,
        n_wells_no_set=0 if n_sticks else 2,
        radius_m=1609.0,
        lateral_tol=0.25,
        intel_vintage="2025-09-30",
        low_n=n_sticks < 3,
        stale_vintage=False,
        tc_risked=False,
        oil_bbl=series,
        gas_mcf=tuple(v * 3.0 for v in series),
        water_bbl=tuple(v * 2.0 for v in series),
    )


def _with_novi(data: BlueOxExportData, comps: tuple[NoviComparisonZone, ...]) -> BlueOxExportData:
    return BlueOxExportData(
        **{
            **data.__dict__,
            "novi_comparison": comps,
            "novi_intel_vintage": "2025-09-30",
        }
    )


def test_novi_comparison_sheets_and_manifest_keys() -> None:
    data = _data()
    comps = (_novi_zone("WOLFCAMP A"), _novi_zone("THIRD BONE SPRING", n_sticks=0))
    wb = load_workbook(io.BytesIO(build_blueox_workbook(_with_novi(data, comps))))

    names = wb.sheetnames
    assert names.index("novi_comparison") == names.index("curve_params") + 1
    assert names.index("novi_comparison_meta") == names.index("novi_comparison") + 1
    assert names.index("manifest") == names.index("novi_comparison_meta") + 1

    ws = wb["novi_comparison"]
    assert [c.value for c in ws[1]] == ["area", "month", "oil_bbl", "gas_mcf", "water_bbl"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    # Only the zone with sticks contributes series rows, months 1..12.
    assert {r[0] for r in rows} == {"WOLFCAMP A"}
    assert [r[1] for r in rows] == list(range(1, 13))
    assert rows[0][2] == 100.0 and rows[0][3] == 300.0 and rows[0][4] == 200.0

    meta_rows = list(wb["novi_comparison_meta"].iter_rows(values_only=True))
    assert meta_rows[0] == (
        "area",
        "n_sticks",
        "n_self",
        "n_neighborhood",
        "n_pud",
        "n_res",
        "n_wells_no_set",
        "radius_m",
        "lateral_tol",
        "intel_vintage",
        "low_n_flag",
        "stale_vintage_flag",
        "tc_risked",
    )
    by_zone = {r[0]: r for r in meta_rows[1:]}
    assert by_zone["WOLFCAMP A"][1] == 4
    assert by_zone["THIRD BONE SPRING"][1] == 0
    assert by_zone["THIRD BONE SPRING"][10] is True  # low_n flagged, not hidden

    kv = _manifest_kv(wb)
    assert kv["novi_intel_vintage"] == "2025-09-30"
    assert kv["novi_selection_radius_m"] == 1609.0
    assert kv["novi_selection_lateral_tol"] == 0.25
    assert kv["novi_alignment"] == "novi_to_ip_tc_to_peak"
    assert kv["novi_rate_to_volume_days"] == 30


def test_novi_comparison_mixed_basin_tol_defers_to_meta() -> None:
    """Per-basin ll tolerance (2026-07-30): a single-basin deal declares
    the one value in the manifest; a mixed-basin deal must not declare
    one zone's tolerance for all — it defers to the per-zone meta column."""
    data = _data()
    midland = NoviComparisonZone(
        **{
            **_novi_zone("WOLFCAMP A").__dict__,
            "lateral_tol": 0.40,
        }
    )
    comps = (midland, _novi_zone("THIRD BONE SPRING", n_sticks=0))
    wb = load_workbook(io.BytesIO(build_blueox_workbook(_with_novi(data, comps))))
    kv = _manifest_kv(wb)
    assert kv["novi_selection_lateral_tol"] == "per_zone_see_novi_comparison_meta"
    by_zone = {r[0]: r for r in wb["novi_comparison_meta"].iter_rows(values_only=True)}
    assert by_zone["WOLFCAMP A"][8] == 0.40
    assert by_zone["THIRD BONE SPRING"][8] == 0.25

    # uniform wide tolerance declares the number, as before
    uniform = tuple(NoviComparisonZone(**{**z.__dict__, "lateral_tol": 0.40}) for z in comps)
    wb2 = load_workbook(io.BytesIO(build_blueox_workbook(_with_novi(data, uniform))))
    assert _manifest_kv(wb2)["novi_selection_lateral_tol"] == 0.40


def test_novi_comparison_absent_keeps_legacy_shape() -> None:
    wb = load_workbook(io.BytesIO(build_blueox_workbook(_data())))
    assert "novi_comparison" not in wb.sheetnames
    assert "novi_comparison_meta" not in wb.sheetnames
    kv = _manifest_kv(wb)
    assert "novi_intel_vintage" not in kv
    assert "novi_alignment" not in kv


def test_novi_comparison_must_cover_every_zone() -> None:
    data = _data()
    with pytest.raises(BlueOxContractError, match="cover every zone"):
        build_blueox_workbook(_with_novi(data, (_novi_zone("WOLFCAMP A"),)))


def test_novi_comparison_vector_rules() -> None:
    data = _data()
    ragged = NoviComparisonZone(
        **{
            **_novi_zone("WOLFCAMP A").__dict__,
            "gas_mcf": (1.0, 2.0),
        }
    )
    with pytest.raises(BlueOxContractError, match="equal-length"):
        build_blueox_workbook(
            _with_novi(data, (ragged, _novi_zone("THIRD BONE SPRING", n_sticks=0)))
        )

    sneaky = NoviComparisonZone(
        **{
            **_novi_zone("WOLFCAMP A", n_sticks=0).__dict__,
            "oil_bbl": (1.0,),
            "gas_mcf": (1.0,),
            "water_bbl": (1.0,),
        }
    )
    with pytest.raises(BlueOxContractError, match="must carry no series"):
        build_blueox_workbook(
            _with_novi(data, (sneaky, _novi_zone("THIRD BONE SPRING", n_sticks=0)))
        )
