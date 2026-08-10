"""Geologic risking — pure-function tests.

The invariants that make a scalar MUL safe to bake into delivered
volumes:

  * Arps homogeneity: scaling (qi, qo) scales the evaluated rate vector
    exactly — so param-level and rate-level scaling can never disagree.
  * apply_risking touches ONLY rate/volume-dimensioned values; shape
    params (b/Di/Df/peak_index), fit quality, well counts, and the
    observed-actuals overlay pass through untouched.
  * Identity at 1.0 (the unrisked default) returns the SAME object —
    the regression baselines can only move if a test opts in.
  * SPE orientation (P10 >= P50 >= P90) survives any positive scalar.
"""

from __future__ import annotations

import copy
import io
import uuid
from datetime import UTC, datetime
from typing import Any

import pytest
from openpyxl import load_workbook

from app.db.models import AlignmentMethod, NormalizationBasis, TypeCurve
from app.exports.blueox import (
    RISKING_APPLIED,
    RISKING_UNRISKED,
    build_blueox_workbook,
)
from app.exports.param_row import format_param_row
from app.type_curves.fit_p50 import evaluate_fit
from app.type_curves.risking import apply_risking, is_risked, normalize_multipliers

from .test_blueox_export import _data, _params, _zone


def _fit(qi: float = 300.0, qo: float = 120.0) -> dict[str, Any]:
    return {
        "qi": qi,
        "Di": 2.5,
        "b": 1.1,
        "Df": 0.08,
        "qo": qo,
        "peak_index": 3,
        "r2": 0.97,
        "ramp_eur": 20_000.0,
        "arps_eur": 180_000.0,
        "eur_per_unit": 200_000.0,
        "smoothed_rate": [qo, qi * 0.8, qi, qi * 0.9],
    }


def _series() -> dict[str, Any]:
    def stream(scale: float) -> dict[str, Any]:
        return {
            "p10": [400.0 * scale, 380.0 * scale, None],
            "p25": [350.0 * scale, 330.0 * scale, None],
            "p50": [300.0 * scale, 280.0 * scale, None],
            "p75": [250.0 * scale, 230.0 * scale, None],
            "p90": [200.0 * scale, 180.0 * scale, None],
            "mean": [305.0 * scale, 285.0 * scale, None],
            "well_count": [12, 12, 11],
            "implied_eur_per_1000ft": {"p50": 150_000.0 * scale, "mean": 155_000.0 * scale},
            "fitted": _fit(300.0 * scale, 120.0 * scale),
            "fitted_eur_per_unit": {
                "p10": 260_000.0 * scale,
                "p50": 200_000.0 * scale,
                "p90": None,
            },
            "fitted_per_percentile": {
                "p10": _fit(400.0 * scale, 160.0 * scale),
                "p50": _fit(300.0 * scale, 120.0 * scale),
                "p90": None,
            },
        }

    return {
        "n_months": 3,
        "n_wells": 12,
        "streams": {"oil": stream(1.0), "gas": stream(5.0), "water": stream(2.0)},
        "observed_streams": {"oil": stream(1.0)},
    }


# ======================= normalize / is_risked =======================


def test_normalize_defaults_and_validation() -> None:
    assert normalize_multipliers(None) == {"oil": 1.0, "gas": 1.0, "water": 1.0}
    assert normalize_multipliers({"oil": 0.85}) == {"oil": 0.85, "gas": 1.0, "water": 1.0}
    with pytest.raises(ValueError, match="unknown risking stream"):
        normalize_multipliers({"ngl": 0.9})
    for bad in (0.0, -1.0, float("nan"), float("inf")):
        with pytest.raises(ValueError, match="positive finite"):
            normalize_multipliers({"gas": bad})


def test_is_risked() -> None:
    assert not is_risked(None)
    assert not is_risked({})
    assert not is_risked({"oil": 1.0})
    assert is_risked({"oil": 0.85})


# ============================ apply_risking ============================


def test_identity_returns_same_object() -> None:
    s = _series()
    assert apply_risking(s, {}) is s
    assert apply_risking(s, {"oil": 1.0, "gas": 1.0, "water": 1.0}) is s


def test_scaling_touches_only_rate_dimensioned_values() -> None:
    s = _series()
    before = copy.deepcopy(s)
    out = apply_risking(s, {"oil": 0.85})

    assert s == before  # input never mutated
    oil = out["streams"]["oil"]
    ref = before["streams"]["oil"]
    # rate arrays scale; None holes survive
    assert oil["p50"] == [v * 0.85 if v is not None else None for v in ref["p50"]]
    assert oil["p50"][2] is None
    # fit: magnitude params + smoothed scale, shape params identical
    for key in ("qi", "qo", "ramp_eur", "arps_eur", "eur_per_unit"):
        assert oil["fitted"][key] == pytest.approx(ref["fitted"][key] * 0.85)
    for key in ("Di", "b", "Df", "peak_index", "r2"):
        assert oil["fitted"][key] == ref["fitted"][key]
    assert oil["fitted"]["smoothed_rate"] == [v * 0.85 for v in ref["fitted"]["smoothed_rate"]]
    # per-percentile store moves WITH the published fit (the
    # _apply_fit_overrides lesson); missing slots stay None
    assert oil["fitted_per_percentile"]["p10"]["qi"] == pytest.approx(400.0 * 0.85)
    assert oil["fitted_per_percentile"]["p90"] is None
    assert oil["fitted_eur_per_unit"]["p10"] == pytest.approx(260_000.0 * 0.85)
    assert oil["fitted_eur_per_unit"]["p90"] is None
    assert oil["implied_eur_per_1000ft"]["p50"] == pytest.approx(150_000.0 * 0.85)
    # counts + other streams + observed actuals untouched
    assert oil["well_count"] == ref["well_count"]
    assert out["streams"]["gas"] == before["streams"]["gas"]
    assert out["streams"]["water"] == before["streams"]["water"]
    assert out["observed_streams"] == before["observed_streams"]


def test_spe_order_survives_scaling() -> None:
    out = apply_risking(_series(), {"oil": 0.7})
    oil = out["streams"]["oil"]
    for i in range(2):
        assert oil["p10"][i] >= oil["p25"][i] >= oil["p50"][i]
        assert oil["p50"][i] >= oil["p75"][i] >= oil["p90"][i]


def test_arps_homogeneity_param_scaling_equals_rate_scaling() -> None:
    """The premise the whole design rests on: scaling (qi, qo) scales the
    evaluated rate vector exactly, so risking params at an export seam is
    identical to risking the delivered volumes."""
    mul = 0.85
    base = evaluate_fit(qi=300.0, Di=2.5, b=1.1, Df=0.08, qo=120.0, peak_index=3, n_months=120)
    scaled = evaluate_fit(
        qi=300.0 * mul,
        Di=2.5,
        b=1.1,
        Df=0.08,
        qo=120.0 * mul,
        peak_index=3,
        n_months=120,
    )
    for a, b_ in zip(base["smoothed_rate"], scaled["smoothed_rate"], strict=True):
        assert b_ == pytest.approx(a * mul, rel=1e-9)
    assert scaled["eur_per_unit"] == pytest.approx(base["eur_per_unit"] * mul, rel=1e-9)


# ============================ param table ============================


def _tc(risk: dict[str, Any] | None = None) -> TypeCurve:
    tc = TypeCurve()
    tc.id = uuid.uuid4()
    tc.name = "wca_west"
    tc.notes = None
    tc.filter_spec = {}
    tc.included_api10s = ["4200000001"]
    tc.normalization_basis = NormalizationBasis.PER_LATERAL_FT
    tc.alignment_method = AlignmentMethod.PEAK_RAMP
    tc.series = _series()
    tc.created_at = datetime(2026, 7, 24, tzinfo=UTC)
    tc.version_of = None
    tc.deal_id = None
    tc.is_stale = False
    tc.forecast_overrides = {}
    tc.risk_multipliers = risk or {}
    return tc


def test_param_row_unrisked_unchanged_and_risked_scales() -> None:
    clean = format_param_row(_tc())
    risked = format_param_row(_tc({"oil": 0.5}))
    assert clean[0] == "wca_west"
    # Factor in the suffix: per-stream list when multipliers differ …
    assert risked[0] == "wca_west [RISKED ×0.50 oil]"
    # … collapsed to one factor when all three streams share it.
    uniform = format_param_row(_tc({"oil": 0.8, "gas": 0.8, "water": 0.8}))
    assert uniform[0] == "wca_west [RISKED ×0.80]"
    # Oil Peak (qi, per-10k scaled): 300 -> 3,000 clean, 1,500 risked
    assert clean[3] == "3,000"
    assert risked[3] == "1,500"
    # Oil B / Oil Di are shape params — identical either way
    assert risked[4] == clean[4]
    assert risked[5] == clean[5]
    # Gas untouched by an oil-only MUL
    assert risked[7:13] == clean[7:13]


# ============================ blue ox workbook ============================


def test_manifest_risking_token_and_risk_mult_column() -> None:
    # Unrisked default: token + all-1.0 risk_mult column
    wb = load_workbook(io.BytesIO(build_blueox_workbook(_data())))
    kv = {r[0]: r[1] for r in wb["manifest"].iter_rows(values_only=True) if r and r[0] and r[1]}
    assert kv["risking"] == RISKING_UNRISKED
    ws = wb["curve_params"]
    header = [c.value for c in ws[1]]
    assert header.index("risk_mult") == header.index("dmin") + 1
    basis_col = header.index("qi_basis") + 1
    mult_col = header.index("risk_mult") + 1
    for r in range(2, ws.max_row + 1):
        assert ws.cell(row=r, column=basis_col).value == "fitted_qi"
        assert ws.cell(row=r, column=mult_col).value == 1.0

    # Risked: token flips; per-row basis + MUL carry through verbatim
    params = [
        {**p, "qi_basis": "fitted_qi_risked", "risk_mult": 0.85, "qi": p["qi"] * 0.85}
        if p["stream"] == "oil"
        else p
        for p in _params(("P10", "P90"))
    ]
    zone = _zone("WOLFCAMP A", ("4200000001",))
    risked_zone = type(zone)(**{**zone.__dict__, "curve_params": params})
    data = _data(zones=[risked_zone])
    data = type(data)(**{**data.__dict__, "risking": RISKING_APPLIED})
    wb = load_workbook(io.BytesIO(build_blueox_workbook(data)))
    kv = {r[0]: r[1] for r in wb["manifest"].iter_rows(values_only=True) if r and r[0] and r[1]}
    assert kv["risking"] == RISKING_APPLIED
    ws = wb["curve_params"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    for row in rows[1:]:
        cells = dict(zip(header, row, strict=False))
        if cells["stream"] == "oil":
            assert cells["qi_basis"] == "fitted_qi_risked"
            assert cells["risk_mult"] == 0.85
        else:
            assert cells["qi_basis"] == "fitted_qi"
            assert cells["risk_mult"] == 1.0
