"""Workbook-shape tests for the deal export.

Builds the workbook bytes directly via ``_build_workbook`` (no DB
required) and asserts the sheet layout the engineering team handed us:
two sheets per assigned type curve — one metadata, one forecast — with
the forecast tab holding the 50-year fitted projection (600 monthly
rows + 1 header).
"""

from __future__ import annotations

import io
import uuid
from datetime import datetime, timezone
from typing import Any

from openpyxl import load_workbook

from app.api.deals import _build_workbook, _sheet_slug
from app.db.models import AlignmentMethod, Deal, NormalizationBasis, TypeCurve


def _stub_stream(rate: float = 100.0, n: int = 24) -> dict[str, Any]:
    """Minimal per-stream payload: flat-rate per-percentile arrays + a
    ``fitted_per_percentile`` entry so ``_evaluate_fitted_rates`` can
    skip the refit path and just evaluate the persisted params."""
    arr = [rate] * n
    fit = {
        # Constant-rate Arps: Di=0 reduces to qi forever. Good enough
        # to verify the workbook plumbing without exercising scipy.
        "qi": rate,
        "Di": 0.0,
        "b": 0.0,
        "Df": 0.0,
        "qo": rate,
        "peak_index": 0,
        "eur_per_unit": rate * 365.0 * 50.0,
    }
    return {
        "p10": arr, "p25": arr, "p50": arr, "p75": arr, "p90": arr,
        "mean": arr, "well_count": [1] * n,
        "implied_eur_per_1000ft": {
            "p10": 100.0, "p25": 100.0, "p50": 100.0,
            "p75": 100.0, "p90": 100.0, "mean": 100.0,
        },
        "fitted_per_percentile": {
            "p10": fit, "p25": fit, "p50": fit,
            "p75": fit, "p90": fit, "mean": fit,
        },
    }


def _curve(name: str) -> TypeCurve:
    tc = TypeCurve()
    tc.id = uuid.uuid4()
    tc.name = name
    tc.notes = "stub for export test"
    tc.filter_spec = {"formation": "Wolfcamp A", "county": "Loving"}
    tc.included_api10s = ["42100000000001", "42100000000002"]
    tc.normalization_basis = NormalizationBasis.PER_LATERAL_FT
    tc.alignment_method = AlignmentMethod.FIRST_PROD_MONTH
    tc.series = {
        "n_months": 24,
        "n_wells": 2,
        "streams": {
            "oil": _stub_stream(rate=100.0),
            "gas": _stub_stream(rate=500.0),
            "water": _stub_stream(rate=200.0),
        },
    }
    tc.created_at = datetime(2026, 5, 1, tzinfo=timezone.utc)
    tc.version_of = None
    tc.deal_id = None
    return tc


def _deal() -> Deal:
    d = Deal()
    d.id = uuid.uuid4()
    d.name = "holdTheLine"
    d.notes = None
    d.created_at = datetime(2026, 5, 1, tzinfo=timezone.utc)
    return d


def test_export_has_two_sheets_per_curve() -> None:
    curves = [_curve("holdTheLine_bs1s_v2"), _curve("holdTheLine_bs2s_v2")]
    content = _build_workbook(_deal(), curves)
    wb = load_workbook(io.BytesIO(content), read_only=True)
    names = wb.sheetnames
    assert len(names) == 4
    # Sheet pairs are emitted in input order: meta then forecast per curve.
    assert "meta" in names[0] and "bs1s_v2" in names[0]
    assert "forecast" in names[1] and "bs1s_v2" in names[1]
    assert "meta" in names[2] and "bs2s_v2" in names[2]
    assert "forecast" in names[3] and "bs2s_v2" in names[3]


def test_forecast_sheet_has_full_50yr_horizon() -> None:
    curves = [_curve("holdTheLine_bs1s_v2")]
    content = _build_workbook(_deal(), curves)
    wb = load_workbook(io.BytesIO(content))
    forecast_ws = [s for s in wb.sheetnames if "forecast" in s][0]
    ws = wb[forecast_ws]
    # 600 monthly rows + 1 header.
    assert ws.max_row == 601
    # 1 month column + 3 streams × 6 pcts × 2 (rate/cum).
    assert ws.max_column == 1 + 3 * 6 * 2
    # Header sanity: first column reflects alignment method, last column
    # is the water-mean cum.
    assert ws.cell(row=1, column=1).value == "month_since_first_prod"
    assert ws.cell(row=1, column=ws.max_column).value == "water_mean_cum"
    # Row 1 month index = 1; row 600's last data row = 600.
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=601, column=1).value == 600


def test_metadata_sheet_contains_curve_fields() -> None:
    tc = _curve("holdTheLine_bs1s_v2")
    content = _build_workbook(_deal(), [tc])
    wb = load_workbook(io.BytesIO(content))
    meta_ws = [s for s in wb.sheetnames if "meta" in s][0]
    ws = wb[meta_ws]
    # Convert to a flat field->value dict for the two-column section so
    # the test stays robust to non-essential row ordering tweaks.
    rows = list(ws.iter_rows(values_only=True))
    kv = {r[0]: r[1] for r in rows if r and r[0] is not None}
    assert kv["name"] == "holdTheLine_bs1s_v2"
    # Humanized at the export boundary — the DB enum is "per_lateral_ft"
    # but the aggregation math is actually per 1,000 ft, matching the
    # chart axes and the EUR-table header.
    assert kv["normalization_basis"] == "per_1000_lateral_ft"
    assert kv["alignment_method"] == "first_prod_month"
    assert kv["n_wells"] == 2
    # filter_spec entries land further down in the same column-A layout.
    assert kv["formation"] == "Wolfcamp A"


def test_metadata_sheet_includes_fitted_50yr_eur() -> None:
    """The deal-handoff workbook reports the 50-yr Arps projection per
    percentile (technical EUR), NOT the data-window cumsum (which got
    mistaken for an EUR). The stub fit is constant-rate (Di=0=Df=b),
    so its 50-yr integral collapses to qi * (365/12) * 600 months."""
    tc = _curve("holdTheLine_bs1s_v2")
    content = _build_workbook(_deal(), [tc])
    wb = load_workbook(io.BytesIO(content))
    meta_ws = [s for s in wb.sheetnames if "meta" in s][0]
    ws = wb[meta_ws]
    rows = list(ws.iter_rows(values_only=True))

    # Find the fitted-EUR header row and confirm the legacy implied-EUR
    # header is no longer emitted.
    header_labels = [r[0] for r in rows if r and r[0]]
    assert any(
        "fitted_eur_per_1000ft" in str(label) for label in header_labels
    )
    assert not any(
        "implied_eur_per_1000ft" in str(label) for label in header_labels
    )

    # The row immediately following ["stream", "p10", ..., "mean"] is oil.
    header_idx = next(
        i for i, r in enumerate(rows) if r and r[0] == "stream"
    )
    oil_row = rows[header_idx + 1]
    assert oil_row[0] == "oil"
    # Constant-rate stub: 100 BOPD/1000 ft × (365/12) days × 600 months
    # — the shared display-EUR convention (ramp_arps.trapezoid_eur).
    expected = 100.0 * (365.0 / 12.0) * 600
    p50 = oil_row[3]
    assert p50 is not None
    # ±1% tolerance — the integration uses 600 monthly samples; numeric
    # noise is negligible but we don't want to pin to an exact float.
    assert abs(p50 - expected) / expected < 0.01


def test_deepcopy_isolation_in_patch_path() -> None:
    """The patch handler must deepcopy row.series before passing it to
    _apply_fit_overrides; otherwise the nested mutation also mutates the
    original (shared streams ref), SQLAlchemy's snapshot-vs-new compare
    returns equal, and the UPDATE is silently skipped. This test pins
    the contract by asserting that running the helper on a deep-copied
    payload leaves the source untouched."""
    import copy
    from app.api.type_curves import FitOverride, _apply_fit_overrides

    source: dict[str, Any] = {
        "n_months": 24,
        "streams": {"oil": _stub_stream(rate=100.0)},
    }
    # The Di in the auto-fit baseline we're about to mutate.
    pre_di = source["streams"]["oil"]["fitted_per_percentile"]["p50"]["Di"]
    assert pre_di == 0.0

    # SHALLOW copy (the buggy old behavior) — confirm it leaks into the
    # source via the shared streams ref. If this assertion ever stops
    # holding it means dict() started deep-copying, and the prod fix is
    # safe to relax (currently no plans for that).
    shallow = dict(source)
    override = FitOverride(qi=250.0, Di=1.7, b=0.0, Df=0.0, qo=250.0, peak_index=0)
    _apply_fit_overrides(shallow, {"oil": override})
    assert source["streams"]["oil"]["fitted_per_percentile"]["p50"]["Di"] == 1.7, (
        "shallow copy unexpectedly isolated the mutation — re-evaluate the "
        "deepcopy fix in patch_type_curve"
    )

    # Reset and try with the deepcopy the prod handler uses. The source
    # must remain untouched so SQLAlchemy's snapshot stays valid.
    source = {
        "n_months": 24,
        "streams": {"oil": _stub_stream(rate=100.0)},
    }
    deep = copy.deepcopy(source)
    _apply_fit_overrides(deep, {"oil": override})
    assert deep["streams"]["oil"]["fitted_per_percentile"]["p50"]["Di"] == 1.7
    # And the snapshot still reads as the pre-mutation value:
    assert source["streams"]["oil"]["fitted_per_percentile"]["p50"]["Di"] == 0.0


def test_metadata_sheet_includes_p50_params_block() -> None:
    """Downstream econ tools sometimes want the raw Arps params rather
    than the pre-integrated monthly streams (so they can re-evaluate at
    quarterly cash horizons, apply their own econ cutoff, etc). The
    metadata sheet carries a `fitted_p50_params (per stream)` block —
    one row per stream with model_type, qi, Di, b, Df, qo, peak_index.
    Sourced from fitted_per_percentile.p50 (which the Update Fit path
    now keeps in lock-step with the published `fitted`)."""
    tc = _curve("holdTheLine_bs1s_v2")
    content = _build_workbook(_deal(), [tc])
    wb = load_workbook(io.BytesIO(content))
    meta_ws = [s for s in wb.sheetnames if "meta" in s][0]
    ws = wb[meta_ws]
    rows = list(ws.iter_rows(values_only=True))

    # Locate the params block by its header label.
    header_idx = next(
        (
            i
            for i, r in enumerate(rows)
            if r and r[0] and "fitted_p50_params" in str(r[0])
        ),
        None,
    )
    assert header_idx is not None, "fitted_p50_params header missing"

    # Column header row immediately follows the section title.
    col_header = rows[header_idx + 1]
    assert col_header[0] == "stream"
    assert "qi" in col_header
    assert "Di" in col_header
    assert "peak_index" in col_header

    # Pull oil row — stub fit is constant-rate (qi=100, Di=0, b=0).
    oil_row_idx = header_idx + 2
    oil_row = rows[oil_row_idx]
    assert oil_row[0] == "oil"
    # Map column header → value for stable assertions across reorderings.
    by_col = dict(zip(col_header, oil_row, strict=False))
    assert by_col["qi"] == 100.0
    assert by_col["Di"] == 0.0
    assert by_col["b"] == 0.0
    assert by_col["peak_index"] == 0


def test_p50_override_flows_to_fitted_per_percentile() -> None:
    """Update Fit must mirror the override into fitted_per_percentile.p50
    (and fitted_eur_per_unit.p50). Otherwise the export's forecast tab
    + metadata-sheet EUR row keep showing the stale auto-fit, which is
    the "two exports of the same deal look identical" bug. This test
    pins the contract at the unit-of-work level so it can't drift
    back."""
    from app.api.type_curves import FitOverride, _apply_fit_overrides

    # Start from an auto-fit payload where p50 says qi=100. Override to
    # qi=250 and assert the per-percentile p50 picks up the new value.
    payload: dict[str, Any] = {
        "n_months": 24,
        "streams": {"oil": _stub_stream(rate=100.0)},
    }
    pre = payload["streams"]["oil"]["fitted_per_percentile"]["p50"]
    assert pre["qi"] == 100.0  # the auto-fit baseline

    override = FitOverride(qi=250.0, Di=0.0, b=0.0, Df=0.0, qo=250.0, peak_index=0)
    _apply_fit_overrides(payload, {"oil": override})

    fitted = payload["streams"]["oil"]["fitted"]
    per_pct = payload["streams"]["oil"]["fitted_per_percentile"]["p50"]
    eur_map = payload["streams"]["oil"]["fitted_eur_per_unit"]

    assert fitted["qi"] == 250.0
    # The key invariant: per-percentile P50 sees the new qi (was 100,
    # now 250 to match the published fit).
    assert per_pct["qi"] == 250.0
    assert per_pct["Di"] == 0.0
    # smoothed_rate is dropped from the per-percentile slot — params
    # only, matching the auto-fit's emission shape.
    assert "smoothed_rate" not in per_pct
    # The scalar EUR companion was also refreshed.
    assert eur_map["p50"] == fitted["eur_per_unit"]
    # Other percentiles stay at the auto-fit (only p50 was tweaked).
    for key in ("p10", "p25", "p75", "p90", "mean"):
        assert payload["streams"]["oil"]["fitted_per_percentile"][key]["qi"] == 100.0


def test_sheet_slug_collision_disambiguates() -> None:
    used: set[str] = set()
    # Long names with identical 22-char prefix should produce distinct slugs.
    a = _sheet_slug("holdTheLine_bs1s_v2_alpha", used)
    b = _sheet_slug("holdTheLine_bs1s_v2_beta", used)
    assert a != b
    assert a in used and b in used


