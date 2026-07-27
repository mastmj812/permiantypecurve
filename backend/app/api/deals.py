"""Deals API + deal-level Excel export.

A deal groups type curves for a single acquisition / divestiture so an
engineer can hand off all of them as one workbook (one metadata sheet +
one fitted-forecast sheet per curve). Cardinality is 1:N — a curve has
at most one deal via ``type_curves.deal_id``.

  POST   /api/deals                        create
  GET    /api/deals                        list with curve counts
  GET    /api/deals/{id}                   detail incl. assigned curve summaries
  PATCH  /api/deals/{id}                   rename / notes
  DELETE /api/deals/{id}                   delete (un-assigns curves via FK SET NULL)
  GET    /api/deals/{id}/export.xlsx       Excel workbook bundling the deal
"""

from __future__ import annotations

import io
import json
import math
import re
import uuid
from contextlib import contextmanager
from datetime import UTC, date, datetime
from typing import Any, Literal

from fastapi import APIRouter, Depends, HTTPException, Request, Response
from starlette.datastructures import UploadFile as StarletteUploadFile
from pydantic import BaseModel, Field, field_validator
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.api.type_curves import (
    _DAYS_PER_MONTH,
    _FITTED_PARAM_KEYS,
    _FORECAST_N_MONTHS,
    _PERCENTILE_KEYS_WITH_MEAN,
    TypeCurveSummary,
    _evaluate_fitted_rates,
    _fitted_eur_per_1000ft,
    _fitted_p50_params,
)
from app.core.logging import get_logger
from app.db.models import Deal, NormalizationBasis, TypeCurve
from app.db.models.production_monthly import ProductionMonthly
from app.db.session import get_session
from app.exports.blueox import (
    LEVEL_TO_SPE_KEY,
    RISKING_APPLIED,
    RISKING_UNRISKED,
    BlueOxContractError,
    BlueOxExportData,
    DsuMetaRow,
    InventoryRow,
    NoviComparisonZone,
    ZoneData,
    blueox_filename,
    build_blueox_workbook,
    monthly_volumes_from_rates,
)
from app.exports.well_rows import (
    PER_WELL_COL_FORMATS,
    PER_WELL_HEADERS,
    per_well_rows,
)
from app.type_curves.aggregate import PERCENTILE_KEYS
from app.type_curves.risking import apply_risking, is_risked, normalize_multipliers
from app.exports.dossier import (
    CurveSlideInput,
    ScenarioSlideInput,
    build_deal_dossier_pptx,
)
from app.warehouse_client.intel_forecast import (
    REP_LATERAL_TOL,
    REP_LOW_N,
    REP_RADIUS_M,
    fetch_intel_median_series,
    fetch_intel_vintage,
    resolve_rep_set,
)
from app.warehouse_client.narvi import (
    NarviInventoryWell,
    NarviScenario,
    derive_handoff_category,
    fetch_narvi_dsu_frames,
    fetch_narvi_inventory,
    fetch_narvi_scenario_detail,
    fetch_narvi_scenarios,
    fetch_scenario_updated_at,
)
from app.warehouse_client.session import get_warehouse_session

router = APIRouter(prefix="/deals", tags=["deals"])
log = get_logger("api.deals")

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


# ============================ schemas ============================


class DealCreate(BaseModel):
    name: str = Field(min_length=1, max_length=255)
    notes: str | None = None


class DealPatch(BaseModel):
    name: str | None = Field(default=None, min_length=1, max_length=255)
    notes: str | None = None


class DealSummary(BaseModel):
    id: uuid.UUID
    name: str
    notes: str | None
    created_at: datetime
    n_curves: int


class DealRow(BaseModel):
    id: uuid.UUID
    name: str
    notes: str | None
    created_at: datetime
    curves: list[TypeCurveSummary]


# ============================ helpers ============================


def _curve_summary(tc: TypeCurve) -> TypeCurveSummary:
    return TypeCurveSummary(
        id=tc.id,
        name=tc.name,
        notes=tc.notes,
        normalization_basis=tc.normalization_basis,
        alignment_method=tc.alignment_method,
        n_wells=len(tc.included_api10s or []),
        created_at=tc.created_at,
        version_of=tc.version_of,
        deal_id=tc.deal_id,
        is_stale=bool(tc.is_stale),
    )


def _load_or_404(session: Session, deal_id: uuid.UUID) -> Deal:
    deal = session.get(Deal, deal_id)
    if deal is None:
        raise HTTPException(status_code=404, detail="deal not found")
    return deal


# ============================ CRUD ============================


@router.post("", response_model=DealRow, status_code=201)
def create_deal(req: DealCreate, session: Session = Depends(get_session)) -> DealRow:
    deal = Deal(id=uuid.uuid4(), name=req.name, notes=req.notes)
    session.add(deal)
    try:
        session.commit()
    except IntegrityError:
        # Unique constraint on name — surface as 409 so the frontend can
        # show a friendly "deal already exists" message.
        session.rollback()
        raise HTTPException(status_code=409, detail="deal name already exists") from None
    session.refresh(deal)
    log.info("deal_created", id=str(deal.id), name=deal.name)
    return DealRow(
        id=deal.id,
        name=deal.name,
        notes=deal.notes,
        created_at=deal.created_at,
        curves=[],
    )


@router.get("", response_model=list[DealSummary])
def list_deals(session: Session = Depends(get_session)) -> list[DealSummary]:
    # Single query: join deals to a per-deal curve-count subquery so the
    # list endpoint stays O(1) regardless of how many curves exist.
    count_subq = (
        select(TypeCurve.deal_id, func.count().label("n"))
        .where(TypeCurve.deal_id.is_not(None))
        .group_by(TypeCurve.deal_id)
        .subquery()
    )
    rows = session.execute(
        select(Deal, func.coalesce(count_subq.c.n, 0))
        .outerjoin(count_subq, count_subq.c.deal_id == Deal.id)
        .order_by(Deal.created_at.desc())
    ).all()
    return [
        DealSummary(
            id=d.id,
            name=d.name,
            notes=d.notes,
            created_at=d.created_at,
            n_curves=int(n),
        )
        for d, n in rows
    ]


@router.get("/{deal_id}", response_model=DealRow)
def get_deal(deal_id: uuid.UUID, session: Session = Depends(get_session)) -> DealRow:
    deal = _load_or_404(session, deal_id)
    curves = session.execute(
        select(TypeCurve)
        .where(TypeCurve.deal_id == deal_id)
        .order_by(TypeCurve.name)
    ).scalars().all()
    return DealRow(
        id=deal.id,
        name=deal.name,
        notes=deal.notes,
        created_at=deal.created_at,
        curves=[_curve_summary(c) for c in curves],
    )


@router.patch("/{deal_id}", response_model=DealRow)
def patch_deal(
    deal_id: uuid.UUID,
    req: DealPatch,
    session: Session = Depends(get_session),
) -> DealRow:
    deal = _load_or_404(session, deal_id)
    if req.name is not None:
        deal.name = req.name
    if "notes" in req.model_fields_set:
        # Allow clearing notes via explicit null — symmetric with the
        # deal_id un-assign behavior on type-curves PATCH.
        deal.notes = req.notes
    try:
        session.commit()
    except IntegrityError:
        session.rollback()
        raise HTTPException(status_code=409, detail="deal name already exists") from None
    session.refresh(deal)
    return get_deal(deal_id, session)


@router.delete("/{deal_id}", status_code=204)
def delete_deal(deal_id: uuid.UUID, session: Session = Depends(get_session)) -> Response:
    deal = _load_or_404(session, deal_id)
    session.delete(deal)
    session.commit()
    # FK is ON DELETE SET NULL, so any assigned curves now have deal_id=null
    # rather than being cascade-deleted.
    log.info("deal_deleted", id=str(deal_id))
    return Response(status_code=204)


# ============================ Excel export ============================


# Excel caps sheet names at 31 chars and forbids \ / ? * [ ] :.
# We reserve room for the " — meta" / " — forecast" suffix (longest is
# 11 chars after sanitization, since the en-dash is one char). 22 keeps
# us inside the cap with a margin.
_SHEET_SLUG_MAX = 22
_SHEET_FORBIDDEN_RE = re.compile(r"[\\/?*\[\]:]+")


def _sheet_slug(name: str, used: set[str]) -> str:
    """Sanitize a curve name into a unique sheet-name prefix.

    Strips Excel-forbidden chars, truncates to ``_SHEET_SLUG_MAX``, and
    appends ``_2``, ``_3`` … on collision so two curves with names that
    truncate to the same prefix get distinct sheets.
    """
    base = _SHEET_FORBIDDEN_RE.sub("_", name).strip()
    base = base[:_SHEET_SLUG_MAX] or "curve"
    candidate = base
    n = 2
    while candidate in used:
        suffix = f"_{n}"
        candidate = f"{base[: _SHEET_SLUG_MAX - len(suffix)]}{suffix}"
        n += 1
    used.add(candidate)
    return candidate


# Humanize the normalization-basis enum for export labels. The DB enum
# value "per_lateral_ft" is historical — the actual aggregation math
# (aggregate.py::_normalizer) divides by lateral_ft / 1000, so rates and
# EUR are expressed PER 1,000 LATERAL FT. The chart axes and the EUR
# table header already say "1,000 ft"; this brings the export metadata
# sheet into line so the downstream economics tool isn't misled.
_NORMALIZATION_LABEL: dict[str, str] = {
    "per_lateral_ft": "per_1000_lateral_ft",
    "per_proppant_lb": "per_million_proppant_lb",
    "per_well": "per_well",
}


def _write_metadata_sheet(
    ws: Any, tc: TypeCurve, session: Session | None, series: dict[str, Any]
) -> None:
    """Two-column field/value layout mirroring ``_metadata_csv`` for the
    top section, then a wide per-well summary block at the bottom that
    expands the bare api10 list into completion-design + EUR metrics
    the engineer wants alongside the type-curve params. The CSV-zip
    export still ships just the api10 list — when an engineer needs
    the wider context they're already in xlsx territory.

    ``session`` may be None — the per-well summary block is then
    skipped (the rest of the metadata sheet still renders). Used by
    pure-function workbook-shape tests that don't stand up a DB.

    ``series`` is the (possibly risked) view the sibling forecast sheet
    was written from — passed in so EUR/param blocks can never disagree
    with the rate columns in the same workbook."""
    from openpyxl.styles import Font

    bold = Font(bold=True)
    ws.append(["field", "value"])
    ws["A1"].font = bold
    ws["B1"].font = bold

    ws.append(["id", str(tc.id)])
    ws.append(["name", tc.name])
    ws.append(["notes", tc.notes or ""])
    ws.append([
        "normalization_basis",
        _NORMALIZATION_LABEL.get(
            tc.normalization_basis.value, tc.normalization_basis.value
        ),
    ])
    ws.append(["alignment_method", tc.alignment_method.value])
    # SPE / PRMS reserves orientation: the p10 columns are the HIGH
    # case (exceeded by 10% of wells), p90 the LOW case. Stamped here
    # so downstream econ tooling can't misread the percentile columns.
    ws.append(["percentile_convention", "SPE (P10 = high case, P90 = low case)"])
    # Geologic risking disclosure — this sheet's EUR/param blocks and
    # the sibling forecast sheet already carry the multiplier.
    if is_risked(tc.risk_multipliers or {}):
        ws.append(["risking", "geologic_multipliers_applied"])
        for s_name, mul in sorted(
            normalize_multipliers(tc.risk_multipliers or {}).items()
        ):
            ws.append([f"risk_mult_{s_name}", round(mul, 4)])
    else:
        ws.append(["risking", "unrisked"])
    ws.append(["created_at", tc.created_at.isoformat()])
    ws.append(["version_of", str(tc.version_of) if tc.version_of else ""])
    ws.append(["n_wells", len(tc.included_api10s or [])])
    ws.append([])

    ws.append(["filter_spec_key", "filter_spec_value"])
    last_row = ws.max_row
    ws.cell(row=last_row, column=1).font = bold
    ws.cell(row=last_row, column=2).font = bold
    for k, v in (tc.filter_spec or {}).items():
        # JSON-serializable scalars; lists/dicts get stringified so the
        # sheet stays in two-column shape regardless of value type.
        ws.append([k, v if isinstance(v, (str, int, float, bool)) else str(v)])
    ws.append([])

    # 50-yr Arps projection per percentile — the EUR the downstream
    # economics tool actually wants. Computed at export time from the
    # persisted fitted_per_percentile params (no econ cutoff; this tool
    # is technical-only). The previous block here reported the
    # data-window cumsum, which is a look-back QC artifact and was
    # easily mistaken for an EUR — dropped on purpose.
    ws.append(["fitted_eur_per_1000ft (50-yr Arps projection)"])
    ws.cell(row=ws.max_row, column=1).font = bold
    header = ["stream", *PERCENTILE_KEYS, "mean"]
    ws.append(header)
    for col_idx in range(1, len(header) + 1):
        ws.cell(row=ws.max_row, column=col_idx).font = bold
    streams = series.get("streams", {})
    for s_name in ("oil", "gas", "water"):
        s = streams.get(s_name, {})
        eur = _fitted_eur_per_1000ft(s)
        ws.append(
            [s_name, *(eur.get(k) for k in PERCENTILE_KEYS), eur.get("mean")]
        )
    ws.append([])

    # Raw P50 Arps params per stream — same shape the
    # `app.type_curves.fit_p50.evaluate_fit` helper produces. Lets the
    # downstream econ tool re-evaluate the curve at its own time grid or
    # apply economic-limit cutoffs without re-fitting from rate columns.
    # Units: qi/qo in BOPD or MCFD per 1000 ft (matches the
    # normalization_basis row above); Di/Df in per-year nominal; b
    # dimensionless; peak_index in months.
    ws.append(["fitted_p50_params (per stream)"])
    ws.cell(row=ws.max_row, column=1).font = bold
    param_header = ["stream", *_FITTED_PARAM_KEYS]
    ws.append(param_header)
    for col_idx in range(1, len(param_header) + 1):
        ws.cell(row=ws.max_row, column=col_idx).font = bold
    for s_name in ("oil", "gas", "water"):
        s = streams.get(s_name, {})
        params = _fitted_p50_params(s)
        if params is None:
            ws.append([s_name, *[None for _ in _FITTED_PARAM_KEYS]])
            continue
        ws.append([s_name, *(params.get(k) for k in _FITTED_PARAM_KEYS)])
    ws.append([])

    # Per-well summary block — expands the bare api10 list into a wide
    # table with completion-design (BWPF / PPF) and per-well EUR
    # metrics. Engineers compare these against the cohort's published
    # fit params above to spot wells dragging the TC away from the
    # cohort average. EURs are recomputed via monthly trapezoid so the
    # workbook matches the Review tab + the slide-export probit.
    ws.append(["per_well_summary"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(list(PER_WELL_HEADERS))
    header_row = ws.max_row
    for col_idx in range(1, len(PER_WELL_HEADERS) + 1):
        ws.cell(row=header_row, column=col_idx).font = bold
    if session is not None:
        for row in per_well_rows(session, list(tc.included_api10s or [])):
            ws.append(list(row))
            data_row = ws.max_row
            for col, fmt in PER_WELL_COL_FORMATS.items():
                ws.cell(row=data_row, column=col).number_format = fmt

    # Top metadata block is 2 columns; the per-well block extends out
    # to column 12. Size the first two for the metadata field/value
    # readability and let the per-well columns auto-size by content.
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    # Wellname tends to be long ("SLINGSHOT A1 2760H"); give it room.
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 18
    for col_letter in ("E", "F", "G", "H", "I", "J", "K", "L"):
        ws.column_dimensions[col_letter].width = 14


def _write_forecast_sheet(ws: Any, tc: TypeCurve, series: dict[str, Any]) -> None:
    """Wide layout: month index then ``{stream}_{pct}_{rate|cum}`` columns.

    Three streams × six percentiles × two value columns = 36 data columns
    plus the month column = 37. 600 monthly rows = the same 50-year
    horizon the CSV export uses. Empty cells for any percentile whose
    fit is missing (same fail-open semantic as ``_forecast_csv``)."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    streams = series.get("streams", {})
    alignment_val = tc.alignment_method.value
    month_col = (
        "month_since_first_prod" if alignment_val == "first_prod_month" else "month_since_peak"
    )

    header: list[str] = [month_col]
    for s_name in ("oil", "gas", "water"):
        for key in _PERCENTILE_KEYS_WITH_MEAN:
            header.append(f"{s_name}_{key}_rate")
            header.append(f"{s_name}_{key}_cum")
    ws.append(header)
    bold = Font(bold=True)
    for col_idx in range(1, len(header) + 1):
        ws.cell(row=1, column=col_idx).font = bold

    # Evaluate once per stream, cache rate arrays + running cum counters.
    rates_by_stream: dict[str, dict[str, list[float] | None]] = {}
    cums_by_stream: dict[str, dict[str, float]] = {}
    for s_name in ("oil", "gas", "water"):
        s = streams.get(s_name)
        rates_by_stream[s_name] = _evaluate_fitted_rates(s or {})
        cums_by_stream[s_name] = {k: 0.0 for k in _PERCENTILE_KEYS_WITH_MEAN}

    # Trapezoid rule: month K's volume = (rate at start + rate at end)
    # / 2 * dt, last month flat-extrapolated. Mirrors the frontend's
    # cumulateNumericArray + the metadata sheet's _fitted_eur_per_1000ft
    # so the xlsx cum column tracks the closed-form scipy.quad EUR
    # (= PPTX value, = UI EUR-table value) to <0.1% on Permian fits.
    # The older right-endpoint rectangle rule had a ~0.2% bias.
    for i in range(_FORECAST_N_MONTHS):
        row: list[Any] = [i + 1]
        for s_name in ("oil", "gas", "water"):
            rates_by_pct = rates_by_stream[s_name]
            cum_by_pct = cums_by_stream[s_name]
            for key in _PERCENTILE_KEYS_WITH_MEAN:
                arr = rates_by_pct.get(key)
                if arr is None or i >= len(arr):
                    row.append(None)
                    row.append(None)
                    continue
                rate = float(arr[i])
                nxt = arr[i + 1] if i + 1 < len(arr) else None
                b = (
                    float(nxt)
                    if nxt is not None and math.isfinite(float(nxt))
                    else rate
                )
                cum_by_pct[key] += (rate + b) / 2.0 * _DAYS_PER_MONTH
                row.append(rate)
                row.append(cum_by_pct[key])
        ws.append(row)

    # Freeze the header row + month column so the column labels stay
    # visible while scrolling through 600 monthly rows.
    ws.freeze_panes = "B2"
    # Number format: rates to 4 decimals, cums to 1. month column is
    # an integer. Apply via column letter so openpyxl writes the format
    # into the column-level XF.
    for col_idx in range(2, len(header) + 1):
        # Even columns (B, D, F, ...) are *_rate; odd are *_cum.
        # 1-indexed: rate at 2, cum at 3, rate at 4, cum at 5, …
        is_rate = (col_idx % 2) == 0
        fmt = "0.0000" if is_rate else "0.0"
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].number_format = fmt
        ws.column_dimensions[letter].width = 16
    ws.column_dimensions["A"].width = 10


def _build_workbook(
    deal: Deal, curves: list[TypeCurve], session: Session | None = None
) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    # Workbook ships with a default empty sheet; remove it so the first
    # curve's metadata sheet is the active sheet on open.
    default = wb.active
    wb.remove(default)

    used_slugs: set[str] = set()
    for tc in curves:
        slug = _sheet_slug(tc.name, used_slugs)
        # Export boundary: geologic risking applies here (stored series
        # is the clean fit; identity pass-through when unrisked).
        series = apply_risking(tc.series or {}, tc.risk_multipliers)
        meta_ws = wb.create_sheet(f"{slug} — meta")
        _write_metadata_sheet(meta_ws, tc, session, series)
        forecast_ws = wb.create_sheet(f"{slug} — forecast")
        _write_forecast_sheet(forecast_ws, tc, series)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_FILENAME_SLUG_RE = re.compile(r"[^A-Za-z0-9]+")


@router.get("/{deal_id}/export.xlsx")
def export_deal(
    deal_id: uuid.UUID, session: Session = Depends(get_session)
) -> Response:
    deal = _load_or_404(session, deal_id)
    curves = session.execute(
        select(TypeCurve)
        .where(TypeCurve.deal_id == deal_id)
        .order_by(TypeCurve.name)
    ).scalars().all()
    if not curves:
        raise HTTPException(status_code=400, detail="deal has no curves assigned")
    content = _build_workbook(deal, list(curves), session)
    safe_name = _FILENAME_SLUG_RE.sub("_", deal.name)[:64] or "deal"
    log.info(
        "deal_exported",
        id=str(deal_id),
        n_curves=len(curves),
        bytes=len(content),
    )
    return Response(
        content=content,
        media_type=XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": f'attachment; filename="{safe_name}.xlsx"',
        },
    )


# ==================== Blue Ox curve-drop export ====================
#
# Deliverable A of the Blue Ox Engineering Curve-Drop Contract v1
# (engineering_db docs/blue_ox_curve_drop_contract.md). Assembly only —
# workbook shape, validation, and the SPE->ascending percentile flip
# live in app.exports.blueox (pure module, no DB/HTTP coupling).


class BlueOxInventoryRowIn(BaseModel):
    """One manual inventory-sheet row (contract §1.3 + category
    amendment)."""

    drilled_lateral_ft: float = Field(ge=3_000, le=25_000)
    # Analog lateral basis behind the curve; defaults to the zone's
    # analog-cohort mean lateral when omitted.
    producing_lateral_ft: float | None = Field(default=None, ge=3_000, le=25_000)
    well_name: str | None = None
    category: Literal["PDP", "PUD", "UPSIDE"] = "PUD"


class ScenarioRef(BaseModel):
    """A (deal_id, scenario_id) pair naming one narvi scenario."""

    deal_id: str = Field(min_length=1)
    scenario_id: str = Field(min_length=1)

    @property
    def pair(self) -> tuple[str, str]:
        return (self.deal_id, self.scenario_id)


class BlueOxZoneSpec(BaseModel):
    type_curve_id: uuid.UUID
    # Adopted verbatim by Blue Ox as the zone identifier (contract
    # Principle 2 — stable across re-drops). Defaults to the curve name.
    zone_name: str | None = Field(default=None, min_length=1, max_length=26)
    # PUD = proven undeveloped (scheduled and valued); UPSIDE = non-
    # proven, carried unscheduled. Same vocabulary as the per-well
    # handoff category so the zone label and its narvi wells align.
    reserve_category: Literal["PUD", "UPSIDE"]

    @field_validator("reserve_category", mode="before")
    @classmethod
    def _legacy_res(cls, v: Any) -> Any:
        # Contract v1 shipped this as PUD | RES; configs saved before
        # the rename still carry "RES".
        if isinstance(v, str) and v.strip().upper() == "RES":
            return "UPSIDE"
        return v
    # narvi formation_blueox bench codes whose planned wells belong to
    # this zone (a zone commonly spans benches: WCA_1 + WCA_2 -> WCA).
    # Benches must be disjoint across zones WITHIN OVERLAPPING SCENARIO
    # SCOPE — two zones may claim the same bench when their scopes are
    # disjoint (the west/east same-bench split on wide deals).
    benches: list[str] = Field(default_factory=list)
    # Scenario scope: None or [] = ALL selected scenarios (the legacy
    # behavior — configs saved before this field validate untouched).
    # Otherwise the subset of narvi_selections whose wells this zone's
    # benches capture — each scenario is one DSU, so this is the
    # geographic grain (e.g. a western BS1_S curve scoped to the
    # western DSU scenarios, an eastern one to the SE DSU).
    scenario_scope: list[ScenarioRef] | None = None
    # Manual rows — the fallback for deals without narvi coverage, or
    # extras on top of the narvi pull.
    inventory: list[BlueOxInventoryRowIn] = Field(default_factory=list)

    def scope_pairs(self) -> frozenset[tuple[str, str]] | None:
        """Scope as a pair set; None = unscoped (covers everything)."""
        if not self.scenario_scope:
            return None
        return frozenset(ref.pair for ref in self.scenario_scope)


class NarviSelection(BaseModel):
    """One narvi scenario feeding this drop's inventory. A drop may
    merge several (one per DSU, plus geometry-forced splits — e.g.
    U-turn wells in half a DSU saved as their own scenario)."""

    deal_id: str = Field(min_length=1)
    scenario_id: str = Field(min_length=1)
    # Stamped server-side when the blueox_config is saved: the narvi
    # scenario's updated_at at pin time. An export from saved config
    # compares it against the live value and refuses (409) when the
    # scenario changed since — different inventory must not ship under
    # an unchanged drop structure without the user seeing it.
    pinned_updated_at: str | None = None


class BlueOxExportRequest(BaseModel):
    codename: str = Field(min_length=1, max_length=64)
    curve_months: int = Field(default=360, ge=12, le=600)
    # Blue Ox ASCENDING labels (their P10 = low case); the flip to our
    # SPE fits happens in assembly via LEVEL_TO_SPE_KEY.
    levels: list[Literal["P10", "P25", "P75", "P90"]] = Field(default_factory=list)
    prepared_by: str = Field(min_length=1, max_length=128)
    zones: list[BlueOxZoneSpec] = Field(min_length=1)
    # narvi scenarios whose inventory_well rows populate the inventory
    # sheet (via each zone's `benches` mapping).
    narvi_selections: list[NarviSelection] = Field(default_factory=list)
    # Benches present in the selected scenarios but deliberately NOT in
    # this workbook (e.g. carried in a sibling deal's drop). Declared in
    # the manifest; an unmapped bench that is not excluded is a hard
    # error — silently dropping wells would understate the location
    # count Blue Ox values.
    exclude_benches: list[str] = Field(default_factory=list)


_BLUEOX_BASIS: dict[NormalizationBasis, str] = {
    NormalizationBasis.PER_LATERAL_FT: "per_1000_lateral_ft",
    NormalizationBasis.PER_WELL: "per_well",
}
_BLUEOX_QI_UNITS: dict[str, str] = {"oil": "bbl/d", "gas": "Mcf/d"}
_BLUEOX_SOURCE_SYSTEM = "anduin type-curve DB (synced from oilgas warehouse)"


def _handoff_category(w: NarviInventoryWell) -> str:
    """PDP / PUD / UPSIDE for one narvi well.

    narvi's user-facing category (auto-derived there and overridable in
    its UI, like the TVD override) wins when persisted. Fallback is the
    agreed rule: existing producers -> PDP; pdp_count_3mi >= 3 -> PUD;
    <= 2 or unscored -> UPSIDE.
    """
    return derive_handoff_category(
        w.handoff_category,
        w.category,
        w.pdp_count_3mi is not None and w.pdp_count_3mi >= 3,
    )


def _inventory_row_from_narvi(nw: NarviInventoryWell) -> InventoryRow:
    """InventoryRow with the gunbarrel geometry columns (2026-07-27
    amendment) filled from what narvi persisted. Legs order is narvi's:
    leg A first, leg B = the U-turn second leg (absent on singles)."""
    xs = nw.gunbarrel_xs
    quads = nw.legs_lonlat
    leg_a = quads[0] if len(quads) >= 1 else (None, None, None, None)
    leg_b = quads[1] if len(quads) >= 2 else (None, None, None, None)
    return InventoryRow(
        producing_lateral_ft=nw.completed_lateral_ft or 0.0,
        drilled_lateral_ft=nw.drilled_lateral_ft or 0.0,
        well_name=nw.well_name,
        category=_handoff_category(nw),
        dsu_id=f"{nw.deal_id}/{nw.scenario_id}",
        bench=nw.formation,
        landing_tvd_ft=nw.target_tvd_ft,
        gunbarrel_offset_ft=xs[0] if len(xs) >= 1 else None,
        gunbarrel_offset_b_ft=xs[1] if len(xs) >= 2 else None,
        lateral_azimuth_deg=nw.lateral_azimuth_deg,
        heel_a_lon=leg_a[0], heel_a_lat=leg_a[1],
        toe_a_lon=leg_a[2], toe_a_lat=leg_a[3],
        heel_b_lon=leg_b[0], heel_b_lat=leg_b[1],
        toe_b_lon=leg_b[2], toe_b_lat=leg_b[3],
    )


def _collect_novi_comparison(
    wh: Session,
    zone_name: str,
    wells: list[NarviInventoryWell],
    tc_risked: bool,
    vintage: str | None,
) -> NoviComparisonZone:
    """One zone's TC-vs-Novi benchmark (2026-07-27 amendment).

    Per captured PLANNED well the representative set resolves via
    ``resolve_rep_set`` (persisted narvi ``novi_rep`` first, warehouse
    fallback for legacy saves); the zone medians over the DEDUPED UNION
    of all sets — a zone mixing generated and curated wells blends
    neighborhood and self sticks by design (union-median, unweighted).
    PDP wells contribute nothing (no intel ML forecast exists). A thin
    result is flagged (low_n), never widened.
    """
    self_ids: set[int] = set()
    hood_ids: set[int] = set()
    low_n = False
    stale = False
    n_no_set = 0
    for w in wells:
        if (w.category or "").lower() == "pdp":
            continue
        rep = resolve_rep_set(wh, w)
        if rep is None or not rep.stick_ids:
            n_no_set += 1
            if rep is not None and rep.mode == "neighborhood":
                low_n = True  # an empty neighborhood is the thinnest case
            continue
        if rep.mode == "self":
            self_ids.update(rep.stick_ids)
        else:
            hood_ids.update(rep.stick_ids)
            low_n = low_n or rep.low_n
        if rep.intel_vintage and vintage and rep.intel_vintage != vintage:
            stale = True

    union = tuple(sorted(self_ids | hood_ids))
    series = fetch_intel_median_series(wh, union) if union else None
    if series is not None and series.dropped_sticks:
        log.info(
            "blueox_novi_comparison_dropped_sticks",
            zone=zone_name,
            dropped=list(series.dropped_sticks),
        )
    n_sticks = series.n_sticks if series is not None else 0
    return NoviComparisonZone(
        zone_name=zone_name,
        n_sticks=n_sticks,
        n_self=len(self_ids),
        n_neighborhood=len(hood_ids - self_ids),
        n_pud=series.n_pud if series is not None else 0,
        n_res=series.n_res if series is not None else 0,
        n_wells_no_set=n_no_set,
        radius_m=REP_RADIUS_M,
        lateral_tol=REP_LATERAL_TOL,
        intel_vintage=vintage,
        low_n=low_n or n_sticks < REP_LOW_N,
        stale_vintage=stale,
        tc_risked=tc_risked,
        oil_bbl=series.oil_bbl if series is not None else (),
        gas_mcf=series.gas_mcf if series is not None else (),
        water_bbl=series.water_bbl if series is not None else (),
    )


def _one_yr_effective(di_nominal: float, b: float) -> float | None:
    """1-yr secant effective decline from nominal Di (per-yr) and Arps b.

    Quoted in curve_params notes so the finance side reads both
    conventions — a nominal 2.5/yr Permian Di is ~75% effective and the
    bare nominal number invites misreading.
    """
    try:
        if b > 0:
            return 1.0 - float((1.0 + b * di_nominal) ** (-1.0 / b))
        return 1.0 - math.exp(-di_nominal)
    except (ValueError, OverflowError, ZeroDivisionError):
        return None


def _collect_blueox_zone(
    session: Session,
    tc: TypeCurve,
    spec: BlueOxZoneSpec,
    delivered: list[str],
    curve_months: int,
    errors: list[str],
    narvi_wells: list[NarviInventoryWell] | None = None,
) -> ZoneData | None:
    """Assemble one zone's volumes / params / analogs / inventory.

    Appends human-readable problems to ``errors`` and returns None when
    the zone can't be assembled (missing fits, unsupported basis, no
    usable producing-lateral default).
    """
    from app.api.type_curves import _fitted_params_by_pct
    from app.type_curves.fit_p50 import evaluate_fit

    zone_name = spec.zone_name or tc.name
    basis = _BLUEOX_BASIS.get(tc.normalization_basis)
    if basis is None:
        errors.append(
            f"{zone_name}: normalization basis "
            f"{tc.normalization_basis.value!r} is not exportable under the contract"
        )
        return None

    # Export boundary: geologic risking applies here — the risked view
    # feeds the fits, so zone-sheet volumes, curve_params qi, and the
    # manifest's Block B sums stay self-consistent by construction.
    muls = normalize_multipliers(tc.risk_multipliers or {})
    series = apply_risking(tc.series or {}, tc.risk_multipliers)
    streams = series.get("streams", {})
    fits_by_stream = {
        s: _fitted_params_by_pct(streams.get(s) or {})
        for s in ("oil", "gas", "water")
    }

    def _rates(fit: dict[str, Any]) -> list[float]:
        evaluated = evaluate_fit(
            qi=fit["qi"], Di=fit["Di"], b=fit["b"], Df=fit["Df"],
            qo=fit.get("qo", fit["qi"]),
            peak_index=fit.get("peak_index", 0),
            n_months=_FORECAST_N_MONTHS,
        )
        return [float(r) for r in evaluated["smoothed_rate"]]

    volumes: dict[str, dict[str, list[float]]] = {}
    params_rows: list[dict[str, Any]] = []
    for lv in delivered:
        spe_key = LEVEL_TO_SPE_KEY[lv]
        by_stream: dict[str, list[float]] = {}
        for stream in ("oil", "gas"):
            fit = fits_by_stream[stream].get(spe_key)
            if fit is None:
                errors.append(
                    f"{zone_name}: no {stream} fit behind level {lv} (SPE "
                    f"{spe_key}) — a partial triplet is a contract failure; "
                    "drop the level or re-save the curve"
                )
                continue
            by_stream[stream] = monthly_volumes_from_rates(
                _rates(fit), curve_months, _DAYS_PER_MONTH
            )
            di = float(fit["Di"])
            b_factor = float(fit["b"])
            peak_month = int(fit.get("peak_index", 0)) + 1
            eff = _one_yr_effective(di, b_factor)
            eff_txt = f"; 1-yr secant effective {eff * 100.0:.1f}%" if eff is not None else ""
            qi_units = _BLUEOX_QI_UNITS[stream]
            if basis == "per_1000_lateral_ft":
                qi_units += " per 1,000 ft"
            params_rows.append({
                "stream": stream,
                "level": lv,
                # Already risked (the fit came from the risked series);
                # risk_mult + qi_basis disclose it per the amendment.
                "qi": float(fit["qi"]),
                "qi_units": qi_units,
                "qi_basis": "fitted_qi_risked" if muls[stream] != 1.0 else "fitted_qi",
                "risk_mult": round(muls[stream], 4),
                "b_factor": b_factor,
                "di": di,
                "dmin": float(fit["Df"]),
                "notes": (
                    f"{tc.alignment_method.value} alignment; qi at peak "
                    f"month {peak_month}{eff_txt}"
                ),
            })
        # Water rides on the base level only (contract: no percentile
        # water columns); included when a water fit exists.
        if lv == "P50":
            wfit = fits_by_stream["water"].get("p50")
            if wfit is not None:
                by_stream["water"] = monthly_volumes_from_rates(
                    _rates(wfit), curve_months, _DAYS_PER_MONTH
                )
        volumes[lv] = by_stream

    analog_rows = per_well_rows(session, list(tc.included_api10s or []))
    lat_idx = PER_WELL_HEADERS.index("Lateral Length")
    laterals = [float(r[lat_idx]) for r in analog_rows if r[lat_idx]]
    cohort_mean_lat = sum(laterals) / len(laterals) if laterals else None

    inventory: list[InventoryRow] = []
    # narvi wells first: completed lateral (EUR driver) is the producing
    # lateral, drilled (legs + turn arc) drives per-ft capex. Category
    # comes from narvi (override-aware) via _handoff_category.
    for nw in narvi_wells or []:
        if nw.completed_lateral_ft is None or nw.drilled_lateral_ft is None:
            errors.append(
                f"{zone_name}: narvi well {nw.well_name!r} "
                f"({nw.deal_id}/{nw.scenario_id}) has no laterals persisted"
            )
            continue
        inventory.append(_inventory_row_from_narvi(nw))
    # Manual rows — fallback / extras on top of the narvi pull.
    for i, inv in enumerate(spec.inventory, start=1):
        producing = inv.producing_lateral_ft or cohort_mean_lat
        if producing is None:
            errors.append(
                f"{zone_name}: inventory row {i} has no producing_lateral_ft "
                "and the analog cohort has no laterals to default from"
            )
            continue
        inventory.append(InventoryRow(
            producing_lateral_ft=float(producing),
            drilled_lateral_ft=float(inv.drilled_lateral_ft),
            well_name=inv.well_name,
            category=inv.category,
        ))

    return ZoneData(
        zone_name=zone_name,
        reserve_category=spec.reserve_category,
        normalization_basis=basis,
        volumes=volumes,
        curve_params=params_rows,
        analog_headers=PER_WELL_HEADERS,
        analog_rows=analog_rows,
        inventory=inventory,
        # Declared in the manifest when non-empty (scoped zones only).
        scenario_scope=tuple(
            f"{ref.deal_id}/{ref.scenario_id}"
            for ref in (spec.scenario_scope or [])
        ),
    )


def _fetch_narvi_by_zone(
    req: BlueOxExportRequest, errors: list[str]
) -> tuple[dict[str, list[NarviInventoryWell]], tuple[str, ...], list[NarviInventoryWell]]:
    """Pull the selected narvi scenarios and distribute wells to zones
    by each zone's bench list.

    Returns ({zone_name: wells}, manifest exclusion strings, unzoned
    PDP wells). Every fetched PLANNED well must land in a zone or an
    explicitly excluded bench — an unmapped bench is an error, never a
    silent drop (the inventory row count IS the location count Blue Ox
    values). Existing producers are NEVER dropped: a PDP well whose
    bench has no zone here comes back in the third element and lands on
    the inventory sheet with its bench code as `area` (the downstream
    gunbarrel automation needs the whole DSU stack).

    Scenario scope: a zone may restrict itself to a subset of the
    selected scenarios (each scenario ≈ one DSU), which is how the same
    bench gets a WEST curve and an EAST curve on a wide deal. Two zones
    may claim one bench only when their scopes are disjoint (an
    unscoped zone overlaps everything). A planned well whose bench is
    claimed but whose scenario no claiming zone covers is a hard error
    — scope must never silently shrink the location count.
    """
    if not req.narvi_selections:
        return {}, (), []

    selected_pairs = {(sel.deal_id, sel.scenario_id) for sel in req.narvi_selections}
    # bench -> [(zone_name, scope-pair-set | None)]; None = all scenarios.
    bench_claims: dict[str, list[tuple[str, frozenset[tuple[str, str]] | None]]] = {}
    for spec in req.zones:
        zname = spec.zone_name or ""
        scope = spec.scope_pairs()
        if scope is not None:
            unknown = scope - selected_pairs
            if unknown:
                errors.append(
                    f"zone {zname!r} scenario_scope names scenarios not in "
                    "this drop's selections: "
                    + ", ".join(f"{d}/{s}" for d, s in sorted(unknown))
                )
        for bench in spec.benches:
            claims = bench_claims.setdefault(bench, [])
            for other_zone, other_scope in claims:
                if scope is None or other_scope is None or (scope & other_scope):
                    errors.append(
                        f"bench {bench!r} mapped to more than one zone with "
                        f"overlapping scenario scope ({other_zone!r} and {zname!r})"
                    )
            claims.append((zname, scope))

    def _zone_for(bench: str, pair: tuple[str, str]) -> str | None:
        """The unique claiming zone whose scope covers this well's
        scenario (uniqueness is guaranteed by the overlap guard)."""
        for zname, scope in bench_claims.get(bench, ()):
            if scope is None or pair in scope:
                return zname
        return None

    pairs = [(sel.deal_id, sel.scenario_id) for sel in req.narvi_selections]
    try:
        with contextmanager(get_warehouse_session)() as wh:
            wells = fetch_narvi_inventory(wh, pairs)
    except Exception as exc:  # surface as a 422, not a 500
        errors.append(f"narvi inventory fetch failed: {exc}")
        return {}, (), []

    seen_pairs = {(w.deal_id, w.scenario_id) for w in wells}
    for pair in pairs:
        if pair not in seen_pairs:
            errors.append(
                f"narvi selection {pair[0]}/{pair[1]} matched no inventory wells"
            )

    # EXISTING producers (narvi category 'pdp', well_name = api10) ride
    # along as display rows for the downstream gunbarrel automation.
    # They dedupe across merged scenarios (the same producer shows up
    # in several DSU plans) and are NEVER dropped — unmapped/excluded
    # benches route them to the unzoned list (bench code as area).
    # Planned wells (everything else) must map or be excluded, and must
    # NOT duplicate across merged scenarios.
    planned = [w for w in wells if (w.category or "").lower() != "pdp"]
    pdp_context: list[NarviInventoryWell] = []
    seen_pdp: set[tuple[str, str]] = set()
    for w in wells:
        if (w.category or "").lower() != "pdp":
            continue
        key = (w.well_name, w.formation or "")
        if key not in seen_pdp:
            seen_pdp.add(key)
            pdp_context.append(w)

    seen_names: dict[tuple[str, str], tuple[str, str]] = {}
    for w in planned:
        key = (w.well_name, w.formation or "")
        if key in seen_names and seen_names[key] != (w.deal_id, w.scenario_id):
            errors.append(
                f"planned well {w.well_name!r} ({w.formation}) appears in both "
                f"{seen_names[key][1]} and {w.scenario_id}"
            )
        seen_names[key] = (w.deal_id, w.scenario_id)

    excluded = dict.fromkeys(req.exclude_benches, 0)
    unmapped: dict[str, int] = {}
    scope_missed: dict[tuple[str, str, str], int] = {}
    by_zone: dict[str, list[NarviInventoryWell]] = {}
    for w in planned:
        bench = w.formation or "(null)"
        if bench in excluded:
            excluded[bench] += 1
            continue
        zone = _zone_for(bench, (w.deal_id, w.scenario_id))
        if zone is None:
            if bench in bench_claims:
                # Claimed bench, uncovered scenario — scoping must never
                # silently shrink the location count.
                key = (bench, w.deal_id, w.scenario_id)
                scope_missed[key] = scope_missed.get(key, 0) + 1
            else:
                unmapped[bench] = unmapped.get(bench, 0) + 1
            continue
        by_zone.setdefault(zone, []).append(w)
    unzoned_pdp: list[NarviInventoryWell] = []
    for w in pdp_context:
        zone = _zone_for(w.formation or "", (w.deal_id, w.scenario_id))
        if zone is None or (w.formation or "") in excluded:
            unzoned_pdp.append(w)
            continue
        by_zone.setdefault(zone, []).append(w)
    if unmapped:
        errors.append(
            "narvi benches with no zone mapping and no exclusion: "
            + ", ".join(f"{b} ({n} wells)" for b, n in sorted(unmapped.items()))
        )
    if scope_missed:
        errors.append(
            "planned wells whose bench is zoned but whose scenario no "
            "zone's scenario_scope covers: "
            + ", ".join(
                f"{b} in {d}/{s} ({n} wells)"
                for (b, d, s), n in sorted(scope_missed.items())
            )
            + " — widen a zone's scope or exclude the bench"
        )
    for bench, n in excluded.items():
        if n == 0:
            errors.append(f"exclude_benches entry {bench!r} matched no wells")
    exclusions = tuple(
        f"{b} ({n} wells)" for b, n in sorted(excluded.items()) if n > 0
    )
    return by_zone, exclusions, unzoned_pdp


def _collect_blueox_data(
    session: Session, deal: Deal, req: BlueOxExportRequest, export_date: date
) -> BlueOxExportData:
    """Assemble the full pure-builder input; 422 on any assembly problem."""
    errors: list[str] = []
    delivered = [
        lv for lv in ("P10", "P25", "P50", "P75", "P90")
        if lv == "P50" or lv in req.levels
    ]

    # Zone names must be resolved before the narvi distribution (the
    # bench map keys on them), so default them from the curve names
    # up front.
    for spec in req.zones:
        if spec.zone_name is None:
            tc = session.get(TypeCurve, spec.type_curve_id)
            if tc is not None:
                spec.zone_name = tc.name[:26].strip()

    narvi_by_zone, inventory_exclusions, unzoned_pdp = _fetch_narvi_by_zone(req, errors)

    # Existing producers whose bench has no zone in THIS workbook still
    # ride the inventory sheet (bench code as `area`) — the gunbarrel
    # automation needs the whole DSU stack, zone coverage or not.
    pdp_context_rows: list[tuple[str, InventoryRow]] = []
    for w in unzoned_pdp:
        if w.completed_lateral_ft is None or w.drilled_lateral_ft is None:
            log.info("blueox_pdp_context_missing_laterals", well=w.well_name)
            continue
        pdp_context_rows.append((
            w.formation or "(unknown)",
            _inventory_row_from_narvi(w),
        ))

    zones: list[ZoneData] = []
    curves: list[TypeCurve] = []
    risked_by_zone: dict[str, bool] = {}
    all_apis: set[str] = set()
    for spec in req.zones:
        tc = session.get(TypeCurve, spec.type_curve_id)
        if tc is None:
            errors.append(f"type curve {spec.type_curve_id} not found")
            continue
        if tc.deal_id != deal.id:
            # Allowed on purpose: one curve can legitimately serve zones
            # in two workbooks (e.g. a formation whose curve lives on a
            # sibling deal but whose planned wells sit in this deal's
            # DSUs). Logged so a mis-picked curve is visible.
            log.info(
                "blueox_cross_deal_curve",
                deal=deal.name,
                curve=tc.name,
                curve_deal_id=str(tc.deal_id),
            )
        curves.append(tc)
        zone = _collect_blueox_zone(
            session, tc, spec, delivered, req.curve_months, errors,
            narvi_wells=narvi_by_zone.get(spec.zone_name or tc.name, []),
        )
        if zone is not None:
            zones.append(zone)
            risked_by_zone[zone.zone_name] = is_risked(tc.risk_multipliers or {})
            all_apis.update(str(r[0]) for r in zone.analog_rows)

    prod_rows_db = session.execute(
        select(ProductionMonthly)
        .where(ProductionMonthly.api10.in_(sorted(all_apis)))
        .order_by(ProductionMonthly.api10, ProductionMonthly.prod_date)
    ).scalars().all() if all_apis else []

    if not prod_rows_db:
        errors.append("no monthly production history found for any analog well")
    if errors:
        raise HTTPException(status_code=422, detail="; ".join(errors))

    # Optional columns obey the no-blank-cells rule: water is included
    # (None coalesced to 0, the standard load-time convention) whenever
    # any month reports it; days_on only when EVERY month has it — a
    # zero producing_days means shut-in, not missing, so None can't be
    # coalesced there.
    has_water = any(r.water_bbl is not None for r in prod_rows_db)
    has_days = all(r.producing_days is not None for r in prod_rows_db)
    headers = ["api10", "date", "oil_bbl", "gas_mcf"]
    if has_water:
        headers.append("water_bbl")
    if has_days:
        headers.append("days_on")
    prod_rows: list[list[Any]] = []
    for r in prod_rows_db:
        row: list[Any] = [
            r.api10,
            r.prod_date.strftime("%Y-%m"),
            float(r.oil_bbl or 0.0),
            float(r.gas_mcf or 0.0),
        ]
        if has_water:
            row.append(float(r.water_bbl or 0.0))
        if has_days:
            row.append(float(r.producing_days or 0.0))
        prod_rows.append(row)

    history_through = max(r.prod_date for r in prod_rows_db).strftime("%Y-%m")
    history_exceptions = tuple(sorted(all_apis - {r.api10 for r in prod_rows_db}))

    # Gunbarrel projection frames (dsu_meta sheet) for exactly the DSUs
    # the written inventory rows reference — the builder validates the
    # two-way tie (no offset without its frame, no orphan frame).
    pair_by_dsu_id = {
        f"{sel.deal_id}/{sel.scenario_id}": (sel.deal_id, sel.scenario_id)
        for sel in req.narvi_selections
    }
    referenced_dsu_ids = sorted(
        {inv.dsu_id for z in zones for inv in z.inventory if inv.dsu_id}
        | {inv.dsu_id for _a, inv in pdp_context_rows if inv.dsu_id}
    )
    dsu_meta: list[DsuMetaRow] = []
    novi_comparison: list[NoviComparisonZone] = []
    novi_vintage: str | None = None
    if referenced_dsu_ids:
        try:
            with contextmanager(get_warehouse_session)() as wh:
                frames = fetch_narvi_dsu_frames(
                    wh, [pair_by_dsu_id[d] for d in referenced_dsu_ids
                         if d in pair_by_dsu_id],
                )
                # TC-vs-Novi benchmark (2026-07-27 amendment): one row
                # per zone, always — n=0 rows declare stickless zones.
                novi_vintage = fetch_intel_vintage(wh)
                for zone in zones:
                    novi_comparison.append(_collect_novi_comparison(
                        wh,
                        zone.zone_name,
                        narvi_by_zone.get(zone.zone_name, []),
                        risked_by_zone.get(zone.zone_name, False),
                        novi_vintage,
                    ))
        except Exception as exc:  # surface as a 422, not a 500
            raise HTTPException(
                status_code=422,
                detail=f"narvi dsu frame / novi comparison fetch failed: {exc}",
            ) from exc
        dsu_meta = [
            DsuMetaRow(
                dsu_id=f.dsu_id, azimuth_deg=f.azimuth_deg,
                origin_lon=f.origin_lon, origin_lat=f.origin_lat,
            )
            for f in frames
        ]

    filename = blueox_filename(req.codename, export_date)
    return BlueOxExportData(
        codename=req.codename,
        export_date=export_date,
        curve_months=req.curve_months,
        levels=tuple(lv for lv in delivered if lv != "P50"),
        prepared_by=req.prepared_by,
        source_system=_BLUEOX_SOURCE_SYSTEM,
        governing_export=f"deal {deal.name} ({deal.id}) — {filename}",
        # Workbook-level risking token (2026-07-24 amendment): risked
        # when ANY zone's curve carries a geologic MUL; per-stream
        # values disclose on curve_params.risk_mult.
        risking=(
            RISKING_APPLIED
            if any(is_risked(tc.risk_multipliers or {}) for tc in curves)
            else RISKING_UNRISKED
        ),
        curve_params_source="; ".join(
            f"{tc.name} ({tc.id}) saved {tc.created_at:%Y-%m-%d}" for tc in curves
        ),
        zones=zones,
        production_headers=headers,
        production_rows=prod_rows,
        production_history_through=history_through,
        history_exceptions=history_exceptions,
        inventory_exclusions=inventory_exclusions,
        pdp_context_rows=pdp_context_rows,
        dsu_meta=tuple(dsu_meta),
        novi_comparison=tuple(novi_comparison),
        novi_intel_vintage=novi_vintage,
    )


def _scenario_status(
    selections: list[NarviSelection],
) -> list[dict[str, Any]]:
    """Pinned vs live updated_at per selection. Warehouse failures
    degrade to current=None / stale=False (the GET must stay usable
    when the warehouse is briefly unreachable; export re-checks)."""
    pairs = [(s.deal_id, s.scenario_id) for s in selections]
    try:
        with contextmanager(get_warehouse_session)() as wh:
            current = fetch_scenario_updated_at(wh, pairs)
    except Exception:
        current = {}
    out = []
    for s in selections:
        cur = current.get((s.deal_id, s.scenario_id))
        out.append({
            "deal_id": s.deal_id,
            "scenario_id": s.scenario_id,
            "pinned_updated_at": s.pinned_updated_at,
            "current_updated_at": cur,
            "stale": (
                s.pinned_updated_at is not None
                and cur is not None
                and cur != s.pinned_updated_at
            ),
        })
    return out


class BlueOxConfigResponse(BaseModel):
    config: BlueOxExportRequest | None
    narvi_status: list[dict[str, Any]]


@router.get("/{deal_id}/blueox-config", response_model=BlueOxConfigResponse)
def get_blueox_config(
    deal_id: uuid.UUID, session: Session = Depends(get_session)
) -> BlueOxConfigResponse:
    deal = _load_or_404(session, deal_id)
    if not deal.blueox_config:
        return BlueOxConfigResponse(config=None, narvi_status=[])
    cfg = BlueOxExportRequest.model_validate(deal.blueox_config)
    return BlueOxConfigResponse(
        config=cfg, narvi_status=_scenario_status(cfg.narvi_selections)
    )


@router.put("/{deal_id}/blueox-config", response_model=BlueOxConfigResponse)
def put_blueox_config(
    deal_id: uuid.UUID,
    req: BlueOxExportRequest,
    session: Session = Depends(get_session),
) -> BlueOxConfigResponse:
    """Save the deal's drop configuration. Pins each narvi selection's
    live updated_at (server-side — whatever the client sent is
    overwritten) so a later export can detect a re-saved scenario."""
    deal = _load_or_404(session, deal_id)
    selected = {(s.deal_id, s.scenario_id) for s in req.narvi_selections}
    for spec in req.zones:
        if session.get(TypeCurve, spec.type_curve_id) is None:
            raise HTTPException(
                status_code=422,
                detail=f"type curve {spec.type_curve_id} not found",
            )
        # A scope naming an unselected scenario is a config bug — catch
        # it at save time, not at export.
        scope = spec.scope_pairs()
        if scope is not None and (unknown := scope - selected):
            raise HTTPException(
                status_code=422,
                detail=(
                    f"zone {spec.zone_name or spec.type_curve_id!s} "
                    "scenario_scope names scenarios not in this config's "
                    "selections: "
                    + ", ".join(f"{d}/{s}" for d, s in sorted(unknown))
                ),
            )
    pairs = [(s.deal_id, s.scenario_id) for s in req.narvi_selections]
    if pairs:
        try:
            with contextmanager(get_warehouse_session)() as wh:
                current = fetch_scenario_updated_at(wh, pairs)
        except Exception as exc:
            raise HTTPException(
                status_code=502, detail=f"narvi scenario lookup failed: {exc}"
            ) from exc
        missing = [p for p in pairs if p not in current]
        if missing:
            raise HTTPException(
                status_code=422,
                detail="narvi scenarios not found: "
                + ", ".join(f"{d}/{s}" for d, s in missing),
            )
        for sel in req.narvi_selections:
            sel.pinned_updated_at = current[(sel.deal_id, sel.scenario_id)]
    deal.blueox_config = req.model_dump(mode="json")
    session.commit()
    log.info("deal_blueox_config_saved", id=str(deal_id), codename=req.codename)
    return BlueOxConfigResponse(
        config=req, narvi_status=_scenario_status(req.narvi_selections)
    )


# Read-only narvi passthroughs (the config UI's scenario pick list).
# Separate router: /api/deals/{deal_id} would swallow a /api/deals/narvi
# path as a UUID parse error.
narvi_router = APIRouter(prefix="/narvi", tags=["narvi"])


class NarviBenchCountOut(BaseModel):
    formation: str | None  # formation_blueox bench code
    category: str  # PDP / PUD / UPSIDE
    n: int


class NarviScenarioOut(BaseModel):
    deal_id: str
    scenario_id: str
    name: str | None
    well_type: str
    total_wells: int | None
    total_completed_ft: float | None
    updated_at: str
    # Per-bench well counts by handoff class — the config UI's surface
    # for confirming zone categories and bench coverage pre-export.
    breakdown: list[NarviBenchCountOut]


@narvi_router.get("/scenarios", response_model=list[NarviScenarioOut])
def list_narvi_scenarios() -> list[NarviScenarioOut]:
    try:
        with contextmanager(get_warehouse_session)() as wh:
            scenarios: list[NarviScenario] = fetch_narvi_scenarios(wh)
    except Exception as exc:
        raise HTTPException(
            status_code=502, detail=f"narvi scenario listing failed: {exc}"
        ) from exc
    return [
        NarviScenarioOut(
            deal_id=s.deal_id,
            scenario_id=s.scenario_id,
            name=s.name,
            well_type=s.well_type,
            total_wells=s.total_wells,
            total_completed_ft=s.total_completed_ft,
            updated_at=s.updated_at,
            breakdown=[
                NarviBenchCountOut(formation=b.formation, category=b.category, n=b.n)
                for b in s.breakdown
            ],
        )
        for s in scenarios
    ]


class NarviWellGeoOut(BaseModel):
    well_name: str
    formation: str | None
    category: str  # PDP / PUD / UPSIDE
    provenance: str | None  # generated / pud / res / pdp
    well_type: str
    n_legs: int
    completed_lateral_ft: float | None
    target_tvd_ft: float | None
    legs_geojson: str | None
    turn_geojson: str | None
    gunbarrel_xs: list[float]


class NarviScenarioDetailOut(BaseModel):
    deal_id: str
    scenario_id: str
    name: str | None
    well_type: str
    aoi_geojson: str | None
    wells: list[NarviWellGeoOut]


@narvi_router.get("/scenario-detail", response_model=NarviScenarioDetailOut)
def get_narvi_scenario_detail(
    deal_id: str, scenario_id: str
) -> NarviScenarioDetailOut:
    """Geometry payload for one scenario — the dossier preview's map
    (AOI + legs/turns) and gunbarrel (offset vs TVD) inputs."""
    try:
        with contextmanager(get_warehouse_session)() as wh:
            detail = fetch_narvi_scenario_detail(wh, deal_id, scenario_id)
    except Exception as exc:
        raise HTTPException(
            status_code=502, detail=f"narvi scenario detail failed: {exc}"
        ) from exc
    if detail is None:
        raise HTTPException(
            status_code=404,
            detail=f"narvi scenario {deal_id}/{scenario_id} not found",
        )
    return NarviScenarioDetailOut(
        deal_id=detail.deal_id,
        scenario_id=detail.scenario_id,
        name=detail.name,
        well_type=detail.well_type,
        aoi_geojson=detail.aoi_geojson,
        wells=[
            NarviWellGeoOut(
                well_name=w.well_name,
                formation=w.formation,
                category=w.category,
                provenance=w.provenance,
                well_type=w.well_type,
                n_legs=w.n_legs,
                completed_lateral_ft=w.completed_lateral_ft,
                target_tvd_ft=w.target_tvd_ft,
                legs_geojson=w.legs_geojson,
                turn_geojson=w.turn_geojson,
                gunbarrel_xs=list(w.gunbarrel_xs),
            )
            for w in detail.wells
        ],
    )


PPTX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.presentationml.presentation"
)


@router.post("/{deal_id}/dossier.pptx")
async def export_deal_dossier(
    deal_id: uuid.UUID,
    request: Request,
    session: Session = Depends(get_session),
) -> Response:
    """Assemble the deal dossier deck from client-captured panels.

    Multipart form: a ``manifest`` JSON field plus PNG files named by
    the manifest's order —

      manifest: {"scenarios": [{"title", "subtitle"}, ...],
                 "curves": [{"type_curve_id", ...}, ...]}
      files:    s{i}_map, s{i}_gunbarrel per scenario;
                c{i}_rate_{stream}, c{i}_cum_{stream} (oil/gas/water)
                and c{i}_map per curve.

    Dynamic file counts rule out fixed UploadFile params — the form is
    parsed by hand and missing parts are 422s, never silently skipped.
    """
    deal = _load_or_404(session, deal_id)
    form = await request.form()
    raw_manifest = form.get("manifest")
    if not isinstance(raw_manifest, str):
        raise HTTPException(status_code=422, detail="missing manifest field")
    try:
        manifest = json.loads(raw_manifest)
    except ValueError as exc:
        raise HTTPException(
            status_code=422, detail=f"manifest is not valid JSON: {exc}"
        ) from exc

    pngs: dict[str, bytes] = {}
    for key, value in form.multi_items():
        if isinstance(value, StarletteUploadFile):
            pngs[key] = await value.read()

    def png(name: str) -> bytes:
        data = pngs.get(name)
        if not data:
            raise HTTPException(
                status_code=422, detail=f"missing panel image {name!r}"
            )
        return data

    scenarios: list[ScenarioSlideInput] = []
    for i, sc in enumerate(manifest.get("scenarios", [])):
        scenarios.append(
            ScenarioSlideInput(
                title=str(sc.get("title") or f"Scenario {i + 1}"),
                subtitle=str(sc.get("subtitle") or ""),
                map_png=png(f"s{i}_map"),
                gunbarrel_png=png(f"s{i}_gunbarrel"),
            )
        )
    curves: list[CurveSlideInput] = []
    for i, cv in enumerate(manifest.get("curves", [])):
        try:
            tc_id = uuid.UUID(str(cv.get("type_curve_id")))
        except ValueError as exc:
            raise HTTPException(
                status_code=422, detail=f"curve {i}: bad type_curve_id"
            ) from exc
        curves.append(
            CurveSlideInput(
                type_curve_id=tc_id,
                stream_pngs={
                    stream: (png(f"c{i}_rate_{stream}"), png(f"c{i}_cum_{stream}"))
                    for stream in ("oil", "gas", "water")
                },
                map_png=png(f"c{i}_map"),
            )
        )

    try:
        content = build_deal_dossier_pptx(session, scenarios, curves)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    safe_name = re.sub(r"[^a-z0-9]+", "_", deal.name.lower()).strip("_") or "deal"
    filename = f"{safe_name}_dossier_{datetime.now(UTC).date().isoformat()}.pptx"
    log.info(
        "deal_dossier_exported",
        id=str(deal_id),
        n_scenarios=len(scenarios),
        n_curves=len(curves),
        bytes=len(content),
    )
    return Response(
        content=content,
        media_type=PPTX_MEDIA_TYPE,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@router.post("/{deal_id}/blueox-export.xlsx")
def export_deal_blueox(
    deal_id: uuid.UUID,
    req: BlueOxExportRequest | None = None,
    allow_stale: bool = False,
    session: Session = Depends(get_session),
) -> Response:
    deal = _load_or_404(session, deal_id)
    if req is None:
        # Run from the saved config — the normal path. Staleness gate:
        # a narvi scenario re-saved after the config was pinned means
        # different inventory under an unchanged structure; refuse
        # unless the caller explicitly allows it (or re-saves the
        # config, which re-pins).
        if not deal.blueox_config:
            raise HTTPException(
                status_code=400,
                detail="deal has no saved blue ox config and no payload was supplied",
            )
        req = BlueOxExportRequest.model_validate(deal.blueox_config)
        stale = [s for s in _scenario_status(req.narvi_selections) if s["stale"]]
        if stale and not allow_stale:
            raise HTTPException(
                status_code=409,
                detail="narvi scenarios changed since the config was saved: "
                + ", ".join(f"{s['deal_id']}/{s['scenario_id']}" for s in stale)
                + " — re-save the config to re-pin, or pass allow_stale=true",
            )
    export_date = datetime.now(UTC).date()
    data = _collect_blueox_data(session, deal, req, export_date)
    try:
        content = build_blueox_workbook(data)
    except BlueOxContractError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    filename = blueox_filename(req.codename, export_date)
    log.info(
        "deal_blueox_exported",
        id=str(deal_id),
        codename=req.codename,
        n_zones=len(data.zones),
        curve_months=req.curve_months,
        levels=list(data.levels),
        bytes=len(content),
    )
    return Response(
        content=content,
        media_type=XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )
