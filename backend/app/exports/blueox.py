"""Blue Ox curve-drop workbook builder.

Implements Deliverable A of the Blue Ox Engineering Curve-Drop Contract
v1 (2026-07-20, checked into engineering_db as
``docs/blue_ox_curve_drop_contract.md``): one governing curve workbook
per deal — zone sheets of monthly volumes, ``meta`` / ``inventory`` /
``analog_production`` / ``curve_params`` / ``manifest`` sheets, and one
``<Zone> meta`` analog sheet per zone.

Pure function of :class:`BlueOxExportData` -> xlsx bytes; no DB, no
HTTP (callable from a future server action). DB assembly lives in
``app.api.deals.export_deal_blueox``.

Convention boundary — the two flips this module owns:

* **Percentiles.** Blue Ox files are ASCENDING (their p10 = the low
  case). anduin persists the SPE orientation (P10 = high, migration
  0021). The caller maps levels via :data:`LEVEL_TO_SPE_KEY` when
  assembling ``ZoneData.volumes`` / ``curve_params`` (our p90 fit
  feeds their ``_p10`` columns); this module validates the result is
  genuinely ascending (contract §1.1 self-check) so a half-flipped
  workbook can never leave the building.
* **NGL.** Per the 2026-07-20 amendment, Blue Ox derives NGL via their
  own yield. ``ngl_bbl`` columns are emitted all-zero at every level
  (keeping the contract's complete-triplet rule) and ``ngl_basis``
  reads ``derived_by_blue_ox_via_yield``.
"""

from __future__ import annotations

import io
import itertools
import math
import re
from collections.abc import Mapping, Sequence
from dataclasses import dataclass, field
from datetime import date
from functools import lru_cache
from typing import TYPE_CHECKING, Any, Literal

if TYPE_CHECKING:
    from pyproj import Geod, Transformer

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Contract §1.1: sums must be strictly ascending across levels. Order of
# delivered percentile levels (base P50 slots between P25 and P75).
BLUEOX_LEVEL_ORDER: tuple[str, ...] = ("P10", "P25", "P50", "P75", "P90")
OPTIONAL_LEVELS: tuple[str, ...] = ("P10", "P25", "P75", "P90")

# Blue Ox ascending level -> anduin SPE fit key. Their p10 is the LOW
# case = our P90 fit (SPE P10 = high). The ONLY place this mapping may
# live — callers import it rather than re-deriving it.
LEVEL_TO_SPE_KEY: dict[str, str] = {
    "P10": "p90",
    "P25": "p75",
    "P50": "p50",
    "P75": "p25",
    "P90": "p10",
}

# Contract Principle 2: reserved sheet names (case-insensitive) + zone
# sheet-name mechanics.
RESERVED_SHEET_NAMES: frozenset[str] = frozenset(
    {
        "meta",
        "inventory",
        "manifest",
        "analog_production",
        "curve_params",
        # 2026-07-27 amendments (pending Blue Ox ack): gunbarrel frame
        # sheet + the TC-vs-Novi comparison sheets.
        "dsu_meta",
        "novi_comparison",
        "novi_comparison_meta",
    }
)
ZONE_NAME_MAX = 26
_ZONE_FORBIDDEN_RE = re.compile(r"[:\\/?*\[\]]")

# Contract §1 filename rule: these stems are reserved on the Blue Ox
# side and must not appear anywhere in the workbook filename.
RESERVED_FILENAME_STEMS: tuple[str, ...] = ("type_curves", "areas", "pdp", "pinned")

# Contract §1.3 hard bounds on inventory laterals.
LATERAL_MIN_FT = 3_000.0
LATERAL_MAX_FT = 25_000.0

# Manifest Block A constants the contract mandates verbatim.
PERCENTILE_ORIENTATION = "ascending"
GAS_BASIS = "wellhead_unshrunk"
# Manifest `risking` tokens (2026-07-24 amendment, pending Blue Ox ack):
# `unrisked` when every curve ships at MUL 1.0; the risked token when
# any zone's volumes carry a declared geologic multiplier — the
# per-stream values are disclosed in curve_params.risk_mult.
RISKING_UNRISKED = "unrisked"
RISKING_APPLIED = "geologic_multipliers_applied"
# 2026-07-20 amendment: Blue Ox applies their own yield; the drop
# carries all-zero ngl_bbl columns.
NGL_BASIS = "derived_by_blue_ox_via_yield"
# 2026-07-27 amendment: the TC-vs-Novi comparison is a benchmark screen
# — the zone-sheet vectors stay the sole economic input. The Novi
# series is IP-aligned while the TC vectors keep their peak-fit head;
# Novi per-day rates convert to period volumes on the Novi-native
# 30-day grid.
NOVI_ALIGNMENT = "novi_to_ip_tc_to_peak"
NOVI_RATE_TO_VOLUME_DAYS = 30

_BOLD = Font(bold=True)


class BlueOxContractError(ValueError):
    """A contract violation that must block the export (never a warning)."""


# Handoff categories on the inventory sheet. PUD/UPSIDE are the valued
# location classes (they count toward Block B gross_locations and
# lateral means); PDP rows are EXISTING producers carried for the
# downstream gunbarrel automation — displayed, never counted, and
# exempt from the planned-well lateral bounds.
INVENTORY_CATEGORIES: tuple[str, ...] = ("PDP", "PUD", "UPSIDE")


@dataclass(frozen=True)
class InventoryRow:
    """One row on the handoff inventory sheet (contract §1.3 + the
    category amendment): a planned undeveloped well (PUD/UPSIDE) or an
    existing producer shown for gunbarrel context (PDP).

    The geometry fields (2026-07-27 amendment, pending Blue Ox ack) let
    the receiver rebuild the per-DSU gunbarrel: plot gunbarrel_offset_ft
    (signed cross-section X, one per producing leg — the ``_b`` values
    are U-turn second legs, blank for singles) against landing_tvd_ft,
    grouped by dsu_id; the projection frame per dsu_id lives on the
    ``dsu_meta`` sheet. heel/toe lon/lat are the true WGS84 leg
    endpoints for map QC. All optional — manual inventory rows and
    legacy narvi saves simply leave them blank."""

    producing_lateral_ft: float
    drilled_lateral_ft: float
    well_name: str | None = None
    category: str = "PUD"
    dsu_id: str | None = None  # "<narvi deal_id>/<scenario_id>"
    bench: str | None = None  # formation_blueox (finer than `area`)
    landing_tvd_ft: float | None = None
    gunbarrel_offset_ft: float | None = None  # producing leg A
    gunbarrel_offset_b_ft: float | None = None  # leg B (U-turn only)
    lateral_azimuth_deg: float | None = None
    heel_a_lon: float | None = None
    heel_a_lat: float | None = None
    toe_a_lon: float | None = None
    toe_a_lat: float | None = None
    heel_b_lon: float | None = None
    heel_b_lat: float | None = None
    toe_b_lon: float | None = None
    toe_b_lat: float | None = None


# The geometry column block appended to the inventory sheet when any
# row carries geometry (attribute name == column header).
_INVENTORY_GEO_COLS: tuple[str, ...] = (
    "dsu_id",
    "bench",
    "landing_tvd_ft",
    "gunbarrel_offset_ft",
    "gunbarrel_offset_b_ft",
    "lateral_azimuth_deg",
    "heel_a_lon",
    "heel_a_lat",
    "toe_a_lon",
    "toe_a_lat",
    "heel_b_lon",
    "heel_b_lat",
    "toe_b_lon",
    "toe_b_lat",
)


@dataclass(frozen=True)
class NoviComparisonZone:
    """One zone's TC-vs-Novi benchmark (2026-07-27 amendment): the
    MEDIAN Novi Intelligence ML forecast of the representative stick
    set behind the zone's captured wells — per-1,000-ft period volumes
    on the Novi-native 30-day grid, IP-aligned. ``n_sticks == 0`` means
    no eligible sticks (e.g. a PDP-dominant zone): the meta sheet still
    carries the row (flagged), the long sheet carries no series."""

    zone_name: str
    n_sticks: int
    n_self: int  # sticks contributed by curated pud/res wells (their own)
    n_neighborhood: int  # sticks contributed by generated wells' neighborhoods
    n_pud: int
    n_res: int
    n_wells_no_set: int  # planned wells with no resolvable set
    radius_m: float
    lateral_tol: float
    intel_vintage: str | None
    low_n: bool  # any contributing set (or the union) below the n floor
    stale_vintage: bool  # a persisted set predates the current vintage
    tc_risked: bool  # the TC side of the comparison carries geologic MULs
    oil_bbl: tuple[float, ...] = ()
    gas_mcf: tuple[float, ...] = ()
    water_bbl: tuple[float, ...] = ()


@dataclass(frozen=True)
class DsuMetaRow:
    """One row of the ``dsu_meta`` sheet: the gunbarrel projection frame
    of one DSU/scenario. offset = signed projection of the leg midpoint
    onto the axis 90° clockwise of azimuth_deg (folded to [0°, 180°))
    through the origin (parcel centroid), in feet."""

    dsu_id: str
    azimuth_deg: float | None
    origin_lon: float | None
    origin_lat: float | None


@dataclass(frozen=True)
class ZoneData:
    """Everything the workbook needs for one zone.

    ``volumes`` maps Blue Ox level label -> stream -> monthly gross
    wellhead volumes (bbl or Mcf per month), already percentile-flipped
    by the caller and already at the zone's normalization basis. ``P50``
    must carry ``oil`` and ``gas`` (``water`` optional); every other
    delivered level carries exactly ``oil`` and ``gas``.

    ``curve_params`` rows carry the engineer's decline params per
    stream x level with keys ``stream`` / ``level`` / ``qi`` /
    ``qi_units`` / ``qi_basis`` / ``risk_mult`` / ``b_factor`` /
    ``di`` / ``dmin`` / ``notes``. ``qi_basis`` defaults to
    ``fitted_qi`` and ``risk_mult`` to 1.0 when a row omits them.
    """

    zone_name: str
    reserve_category: str  # "PUD" | "UPSIDE"
    normalization_basis: str  # "per_1000_lateral_ft" | "per_well"
    volumes: Mapping[str, Mapping[str, Sequence[float]]]
    curve_params: Sequence[Mapping[str, Any]]
    analog_headers: Sequence[str]
    analog_rows: Sequence[Sequence[Any]]
    inventory: Sequence[InventoryRow] = field(default_factory=tuple)
    # narvi scenario scope this zone's inventory was drawn from
    # ("deal/scenario" strings). Empty = all selected scenarios (the
    # unscoped default) — nothing is declared in the manifest, keeping
    # legacy output byte-identical. Non-empty scopes are declared
    # (Principle 4: everything declared, nothing implied) so a
    # west/east same-bench split is auditable downstream.
    scenario_scope: tuple[str, ...] = ()


@dataclass(frozen=True)
class BlueOxExportData:
    """Pure input to :func:`build_blueox_workbook`."""

    codename: str
    export_date: date
    curve_months: int
    levels: tuple[str, ...]  # delivered levels beyond P50, e.g. ("P10", "P90")
    prepared_by: str
    source_system: str
    governing_export: str
    curve_params_source: str
    zones: Sequence[ZoneData]
    production_headers: Sequence[str]
    production_rows: Sequence[Sequence[Any]]
    production_history_through: str  # "YYYY-MM"
    # api10s listed on analog sheets with no available history, declared
    # in the manifest per contract §1.5 ("never silently absent").
    history_exceptions: tuple[str, ...] = ()
    # Inventory benches deliberately excluded from this workbook (e.g.
    # carried in a sibling deal's drop), declared in the manifest —
    # Principle 4: everything declared, nothing implied. Strings like
    # "WDFD (4 wells)".
    inventory_exclusions: tuple[str, ...] = ()
    # Existing producers whose bench maps to NO zone in this workbook —
    # still written to the inventory sheet (the downstream gunbarrel
    # automation needs the whole DSU stack), with the bench code as
    # `area`. Deliberately the one place an `area` value may not match
    # a zone sheet; loaders must tolerate it on category=PDP rows only.
    pdp_context_rows: Sequence[tuple[str, InventoryRow]] = ()
    # Manifest Block A `risking` value — RISKING_UNRISKED unless any
    # zone's curve carries a geologic multiplier (2026-07-24 amendment).
    # The caller computes this from the curves it assembled.
    risking: str = RISKING_UNRISKED
    # Gunbarrel projection frames, one per DSU/scenario referenced by
    # the inventory rows' dsu_id (2026-07-27 amendment). Empty = no
    # dsu_meta sheet (legacy output unchanged).
    dsu_meta: Sequence[DsuMetaRow] = ()
    # TC-vs-Novi benchmark (2026-07-27 amendment). Empty = no
    # novi_comparison sheets and no manifest keys (legacy output
    # unchanged); non-empty must cover EVERY zone (n=0 rows included).
    novi_comparison: Sequence[NoviComparisonZone] = ()
    # Warehouse intel vintage the comparison was computed against.
    novi_intel_vintage: str | None = None


def blueox_filename(codename: str, export_date: date) -> str:
    return f"{codename}_curves_{export_date.isoformat()}.xlsx"


def monthly_volumes_from_rates(
    rates: Sequence[float], curve_months: int, days_per_month: float
) -> list[float]:
    """Daily-rate array -> monthly volumes via the shared trapezoid rule.

    Month K's volume = (rate at start + rate at end) / 2 * dt, last
    available month flat-extrapolated — the SAME rule as the deal xlsx /
    CSV cum columns and ``ramp_arps.trapezoid_eur``, so the drop's Block
    B EURs stay on the one integration convention. Months past the end
    of ``rates`` zero-fill (contract §1.1 tail rule).
    """
    n = len(rates)
    out: list[float] = []
    for i in range(curve_months):
        if i >= n:
            out.append(0.0)
            continue
        a = float(rates[i])
        nxt = rates[i + 1] if i + 1 < n else None
        b = float(nxt) if nxt is not None and math.isfinite(float(nxt)) else a
        out.append((a + b) / 2.0 * days_per_month)
    return out


# ============================ validation ============================

# Ratio-mode refusal (hard, by design): the drop contract's curve_params
# sheet declares Arps params (qi/b/di/dmin) per stream — a ratio-mode
# stream (fitted ratio of cumulative oil; see app.forecasting.ratio) has
# none, and emitting placeholder Arps rows would misdeclare the curve.
# Do NOT relax this or touch the contract/ledger to accommodate it.
RATIO_REFUSAL_NOTE = (
    "ratio-mode streams are not yet in the drop contract — refit the "
    "stream as Arps or await a contract amendment"
)


def ratio_mode_streams(series: Mapping[str, Any] | None) -> list[str]:
    """Streams whose PUBLISHED fitted block is ratio-mode in a saved
    type-curve ``series`` JSONB. Pure + import-light so the zone
    collector, the dossier builder, and tests share one rule."""
    streams = (series or {}).get("streams") or {}
    return [
        s
        for s in ("oil", "gas", "water")
        if ((streams.get(s) or {}).get("fitted") or {}).get("mode") == "ratio"
    ]


def _validate_zone_name(name: str) -> str | None:
    """Return an error string if ``name`` breaks Principle 2, else None."""
    if not name:
        return "zone name is empty"
    if len(name) > ZONE_NAME_MAX:
        return f"zone name {name!r} exceeds {ZONE_NAME_MAX} characters"
    if _ZONE_FORBIDDEN_RE.search(name):
        return f"zone name {name!r} contains a forbidden character (: \\ / ? * [ ])"
    if name != name.strip() or name.startswith("'") or name.endswith("'"):
        return f"zone name {name!r} has leading/trailing spaces or apostrophes"
    if name.casefold() in RESERVED_SHEET_NAMES:
        return f"zone name {name!r} is a reserved sheet name"
    return None


def _delivered_levels(data: BlueOxExportData) -> list[str]:
    """All delivered levels in ascending contract order, P50 included."""
    return [lv for lv in BLUEOX_LEVEL_ORDER if lv == "P50" or lv in data.levels]


def _validate(data: BlueOxExportData) -> None:
    errors: list[str] = []

    filename = blueox_filename(data.codename, data.export_date).lower()
    for stem in RESERVED_FILENAME_STEMS:
        # "curves" is the mandated stem; only flag the reserved ones.
        if stem in filename.replace("_curves_", "_"):
            errors.append(f"filename {filename!r} contains reserved stem {stem!r}")
    if not data.codename or not data.codename.strip():
        errors.append("codename is empty")
    if data.curve_months < 1:
        errors.append(f"curve_months must be >= 1 (got {data.curve_months})")
    for lv in data.levels:
        if lv not in OPTIONAL_LEVELS:
            errors.append(f"unknown percentile level {lv!r}")
    if len(set(data.levels)) != len(data.levels):
        errors.append(f"duplicate percentile levels in {data.levels!r}")
    if not data.zones:
        errors.append("no zones supplied")
    if not re.fullmatch(r"\d{4}-\d{2}", data.production_history_through):
        errors.append(
            f"production_history_through must be YYYY-MM (got {data.production_history_through!r})"
        )

    seen_names: set[str] = set()
    levels = _delivered_levels(data)
    for z in data.zones:
        zerr = _validate_zone_name(z.zone_name)
        if zerr:
            errors.append(zerr)
        if z.zone_name.casefold() in seen_names:
            errors.append(f"duplicate zone name {z.zone_name!r}")
        seen_names.add(z.zone_name.casefold())

        if z.reserve_category not in ("PUD", "UPSIDE"):
            errors.append(
                f"{z.zone_name}: reserve_category must be PUD or UPSIDE "
                f"(got {z.reserve_category!r})"
            )
        if z.normalization_basis not in ("per_1000_lateral_ft", "per_well"):
            errors.append(
                f"{z.zone_name}: normalization_basis must be "
                f"per_1000_lateral_ft or per_well (got {z.normalization_basis!r})"
            )

        # Complete triplets: every delivered level needs oil + gas
        # vectors (ngl is emitted all-zero on our side). Water rides on
        # the base level only.
        for lv in levels:
            by_stream = z.volumes.get(lv)
            if by_stream is None:
                errors.append(f"{z.zone_name}: level {lv} missing entirely")
                continue
            required = ("oil", "gas")
            for stream in required:
                vec = by_stream.get(stream)
                if vec is None:
                    errors.append(
                        f"{z.zone_name}: level {lv} missing {stream} "
                        "(partial triplet is a hard failure)"
                    )
                    continue
                if len(vec) != data.curve_months:
                    errors.append(
                        f"{z.zone_name}: {stream} {lv} has {len(vec)} rows, "
                        f"expected curve_months={data.curve_months}"
                    )
                if any(v is None or not math.isfinite(float(v)) or float(v) < 0 for v in vec):
                    errors.append(
                        f"{z.zone_name}: {stream} {lv} contains negative or non-finite values"
                    )
            extra = set(by_stream) - {"oil", "gas", "water"}
            if extra:
                errors.append(f"{z.zone_name}: level {lv} has unsupported streams {sorted(extra)}")
            if lv != "P50" and "water" in by_stream:
                errors.append(
                    f"{z.zone_name}: percentile water columns are not allowed (water at level {lv})"
                )
        water = z.volumes.get("P50", {}).get("water")
        if water is not None and len(water) != data.curve_months:
            errors.append(
                f"{z.zone_name}: water has {len(water)} rows, "
                f"expected curve_months={data.curve_months}"
            )

        # Ascending monotonicity (contract §1.1 self-check): strictly
        # increasing column sums across delivered levels per stream.
        if len(levels) > 1:
            for stream in ("oil", "gas"):
                sums = [
                    sum(float(v) for v in z.volumes.get(lv, {}).get(stream, [])) for lv in levels
                ]
                if any(a >= b for a, b in itertools.pairwise(sums)):
                    errors.append(
                        f"{z.zone_name}: {stream} sums are not strictly ascending "
                        f"across {levels} (got {[round(s, 1) for s in sums]}) — "
                        "check the SPE->ascending percentile flip"
                    )

        # curve_params: oil + gas P50 rows minimum, every row labelled
        # with a delivered level.
        param_keys = {(str(r.get("stream")), str(r.get("level"))) for r in z.curve_params}
        for stream in ("oil", "gas"):
            if (stream, "P50") not in param_keys:
                errors.append(f"{z.zone_name}: curve_params missing {stream} P50 row")
        for stream, lv in param_keys:
            if lv not in levels:
                errors.append(
                    f"{z.zone_name}: curve_params row {stream} {lv} is not a delivered level"
                )

        # Analog sheet: exactly one header containing "api".
        api_cols = [h for h in z.analog_headers if "api" in str(h).lower()]
        if len(api_cols) != 1:
            errors.append(
                f"{z.zone_name}: analog sheet must have exactly one column "
                f"containing 'api' (got {api_cols!r})"
            )
        if not z.analog_rows:
            errors.append(f"{z.zone_name}: analog sheet has no wells")

        for inv in z.inventory:
            if inv.category not in INVENTORY_CATEGORIES:
                errors.append(
                    f"{z.zone_name}: inventory category {inv.category!r} not in "
                    f"{INVENTORY_CATEGORIES}"
                )
            if inv.category == "PDP":
                # Existing producers are display rows for the gunbarrel
                # automation — the planned-well bounds don't apply.
                continue
            for label, val in (
                ("producing_lateral_ft", inv.producing_lateral_ft),
                ("drilled_lateral_ft", inv.drilled_lateral_ft),
            ):
                if not (LATERAL_MIN_FT <= float(val) <= LATERAL_MAX_FT):
                    errors.append(
                        f"{z.zone_name}: {label} {val} outside "
                        f"{LATERAL_MIN_FT:.0f}-{LATERAL_MAX_FT:.0f} ft"
                    )

    # analog_production <-> analog sheets tie-out (both directions),
    # minus the declared exceptions.
    analog_apis = {
        str(row[_api_col_index(z.analog_headers)]) for z in data.zones for row in z.analog_rows
    }
    prod_headers_l = [str(h).lower() for h in data.production_headers]
    for col in ("api10", "date", "oil_bbl", "gas_mcf"):
        if col not in prod_headers_l:
            errors.append(f"analog_production missing required column {col!r}")
    if "api10" in prod_headers_l:
        api_idx = prod_headers_l.index("api10")
        prod_apis = {str(r[api_idx]) for r in data.production_rows}
        missing = analog_apis - prod_apis - set(data.history_exceptions)
        if missing:
            errors.append(
                "analog wells with no production history and no declared "
                f"exception: {sorted(missing)}"
            )
        orphans = prod_apis - analog_apis
        if orphans:
            errors.append(
                f"analog_production rows for wells not on any analog sheet: {sorted(orphans)}"
            )

    # Unzoned PDP context rows: display-only, must actually be PDP.
    for area, invr in data.pdp_context_rows:
        if invr.category != "PDP":
            errors.append(
                f"pdp_context_rows entry {invr.well_name!r} ({area}) must be "
                f"category PDP (got {invr.category!r})"
            )

    # Gunbarrel geometry (2026-07-27 amendment): every dsu_id referenced
    # by an inventory row must have its projection frame on dsu_meta —
    # an offset without its frame is not reproducible downstream.
    frame_ids = {f.dsu_id for f in data.dsu_meta}
    referenced = {inv.dsu_id for inv in _all_inventory_rows(data) if inv.dsu_id is not None}
    missing_frames = referenced - frame_ids
    if missing_frames:
        errors.append(
            "inventory rows reference DSUs with no dsu_meta frame: "
            + ", ".join(sorted(missing_frames))
        )
    orphan_frames = frame_ids - referenced
    if orphan_frames:
        errors.append("dsu_meta frames with no inventory rows: " + ", ".join(sorted(orphan_frames)))

    # TC-vs-Novi comparison (2026-07-27 amendment): when present it
    # must cover every zone exactly once (n=0 rows for stickless zones
    # — a missing row is indistinguishable from a forgotten zone), and
    # every zone with sticks carries three equal-length vectors.
    if data.novi_comparison:
        zone_names = {z.zone_name for z in data.zones}
        comp_names = [c.zone_name for c in data.novi_comparison]
        if sorted(comp_names) != sorted(zone_names):
            errors.append(
                "novi_comparison must cover every zone exactly once "
                f"(zones {sorted(zone_names)}, comparison {sorted(comp_names)})"
            )
        for c in data.novi_comparison:
            lengths = {len(c.oil_bbl), len(c.gas_mcf), len(c.water_bbl)}
            if c.n_sticks > 0:
                if len(lengths) != 1 or 0 in lengths:
                    errors.append(
                        f"novi_comparison {c.zone_name!r}: oil/gas/water "
                        "vectors must be non-empty and equal-length "
                        f"(got {sorted(len(v) for v in (c.oil_bbl, c.gas_mcf, c.water_bbl))})"
                    )
            elif lengths != {0}:
                errors.append(
                    f"novi_comparison {c.zone_name!r}: n_sticks=0 rows must carry no series"
                )
        if data.novi_intel_vintage is None:
            errors.append("novi_comparison present but novi_intel_vintage undeclared")

    if errors:
        raise BlueOxContractError("Blue Ox contract violations:\n- " + "\n- ".join(errors))


def _api_col_index(headers: Sequence[str]) -> int:
    for i, h in enumerate(headers):
        if "api" in str(h).lower():
            return i
    return 0


# ======================== pre-send value sweep ========================
#
# ``_validate`` above enforces the contract's STRUCTURE (sheets,
# triplets, monotonicity, tie-outs). The sweep below checks what the
# NUMBERS look like. Each check encodes the mechanical signature of a
# defect Blue Ox caught after a send — bro_time's mid-lateral heels
# (2026-07-29), toucan's zero-spread azimuths (2026-08-03), the §8
# lat-first coordinate swap — plus the §6 offset-reproducibility
# invariant and basic zone-vector sanity. Assert-only: the sweep never
# changes an emitted value, column name, or ordering. Any FAIL blocks
# the build (same :class:`BlueOxContractError` path); WARNs ride along
# on the :func:`presend_sweep` report for the human pre-send checklist.

# Cross-repo constants: must match narvi ``src/narvi/records.py``
# FT_PER_M and ``src/narvi/parcel.py`` WORK_EPSG — the persisted §6
# offsets are computed there (``placement.gunbarrel_offset_ft`` in the
# UTM 13N work CRS); reproducing them here requires the same numbers.
FT_PER_M = 3.280839895
WORK_EPSG = 32613  # UTM zone 13N — narvi's planar work CRS

# Permian coordinate envelope (generous). Every emitted pair is
# lon-first (§8); a lat/lon swap puts both values outside these ranges
# immediately, which is exactly the defect signature this guards.
PERMIAN_LON_RANGE = (-110.0, -100.0)
PERMIAN_LAT_RANGE = (30.0, 34.0)

# bro_time signature: a mid-lateral "heel" makes heel->toe ~half the
# stated lateral. 40% catches the halving with headroom for genuine
# chord-vs-along-hole differences on curved sticks.
HEEL_TOE_MISMATCH_FRAC = 0.40

# toucan signature: adopted rows (as-built bearings) in one unit all
# stamped with a single lateral_azimuth_deg. Real per-well as-built
# bearings never agree this tightly across a unit. Axial comparison —
# azimuths fold to [0°, 180°).
AZIMUTH_SPREAD_MIN_DEG = 0.1

# §6 invariant: every offset reproduces from dsu_meta azimuth + origin.
# Persisted offsets are rounded to 0.1 ft and lon/lats to ~1e-6 deg;
# 1 ft covers that rounding while catching any real frame drift.
OFFSET_REPRO_TOL_FT = 1.0

# Loose per-month magnitude ceilings for a per-1,000-ft-normalized zone
# sheet (warning-only — far above any physical Permian month, so a trip
# means the normalization basis itself is suspect, e.g. volumes shipped
# pre-multiplied by lateral). oil ceiling ≈ 2,000 bbl/d per 1,000 ft
# for a full month. per_well sheets scale by the 25,000 ft lateral cap.
_STREAM_MONTH_CEILING_PER_1000FT: dict[str, float] = {
    "oil": 60_000.0,  # bbl/month per 1,000 ft
    "gas": 300_000.0,  # Mcf/month per 1,000 ft
    "water": 100_000.0,  # bbl/month per 1,000 ft
}

SweepStatus = Literal["pass", "fail", "warn"]


@dataclass(frozen=True)
class SweepFinding:
    """One line of the pre-send sweep report."""

    check: str
    status: SweepStatus
    detail: str


@lru_cache(maxsize=1)
def _geod() -> Geod:
    # Lazy: pyproj stays out of the import graph until a sweep needs it
    # (same pattern as well_rows' lazy ramp_arps import).
    from pyproj import Geod

    return Geod(ellps="WGS84")


@lru_cache(maxsize=1)
def _to_work_crs() -> Transformer:
    from pyproj import Transformer

    return Transformer.from_crs(4326, WORK_EPSG, always_xy=True)


def _geodesic_ft(lon1: float, lat1: float, lon2: float, lat2: float) -> float:
    _, _, dist_m = _geod().inv(lon1, lat1, lon2, lat2)
    return float(dist_m) * FT_PER_M


def _axial_delta_deg(a: float, b: float) -> float:
    """Angular distance between two AXIAL bearings (a lateral has no
    direction): fold to [0°, 180°), take the shorter way around."""
    d = abs(a % 180.0 - b % 180.0)
    return min(d, 180.0 - d)


def _axial_spread_deg(vals: Sequence[float]) -> float:
    """Max pairwise axial distance. Callers guarantee len(vals) >= 2."""
    return max(_axial_delta_deg(a, b) for a, b in itertools.combinations(vals, 2))


def _projected_offset_ft(
    heel_lon: float,
    heel_lat: float,
    toe_lon: float,
    toe_lat: float,
    azimuth_deg: float,
    origin_lon: float,
    origin_lat: float,
) -> float:
    """Reproduce narvi's §6 signed cross-section offset from WGS84 leg
    endpoints: transform to the work CRS, take the leg midpoint THERE
    (narvi projects work-CRS midpoints), project onto the axis 90°
    clockwise of the folded azimuth through the origin. Mirrors narvi
    ``placement.cross_axis`` / ``placement.gunbarrel_offset_ft``."""
    tf = _to_work_crs()
    hx, hy = tf.transform(heel_lon, heel_lat)
    tx, ty = tf.transform(toe_lon, toe_lat)
    ox, oy = tf.transform(origin_lon, origin_lat)
    mx, my = (float(hx) + float(tx)) / 2.0, (float(hy) + float(ty)) / 2.0
    a = math.radians(azimuth_deg % 180.0)
    px, py = math.cos(a), -math.sin(a)
    return ((mx - float(ox)) * px + (my - float(oy)) * py) * FT_PER_M


def _labeled_inventory(data: BlueOxExportData) -> list[tuple[str, InventoryRow]]:
    """(area, row) over every inventory row, unzoned PDP context included."""
    rows = [(z.zone_name, inv) for z in data.zones for inv in z.inventory]
    rows.extend(data.pdp_context_rows)
    return rows


def _inv_legs(inv: InventoryRow) -> list[tuple[str, float, float, float, float]]:
    """("a"|"b", heel_lon, heel_lat, toe_lon, toe_lat) per leg with all
    four endpoints populated (blank ``_b`` cells on singles are
    sanctioned; a partially blank leg is simply not checkable)."""
    legs: list[tuple[str, float, float, float, float]] = []
    if (
        inv.heel_a_lon is not None
        and inv.heel_a_lat is not None
        and inv.toe_a_lon is not None
        and inv.toe_a_lat is not None
    ):
        legs.append(("a", inv.heel_a_lon, inv.heel_a_lat, inv.toe_a_lon, inv.toe_a_lat))
    if (
        inv.heel_b_lon is not None
        and inv.heel_b_lat is not None
        and inv.toe_b_lon is not None
        and inv.toe_b_lat is not None
    ):
        legs.append(("b", inv.heel_b_lon, inv.heel_b_lat, inv.toe_b_lon, inv.toe_b_lat))
    return legs


def _header_index(headers: Sequence[str], name: str) -> int | None:
    for i, h in enumerate(headers):
        if str(h).casefold() == name:
            return i
    return None


def _lateral_col_index(headers: Sequence[str]) -> int | None:
    for i, h in enumerate(headers):
        if "lateral" in str(h).casefold():
            return i
    return None


def _num(v: Any) -> float | None:
    if isinstance(v, bool) or not isinstance(v, (int, float)):
        return None
    f = float(v)
    return f if math.isfinite(f) else None


def _findings(
    check: str, n_checked: int, unit: str, fails: list[str], warns: list[str]
) -> list[SweepFinding]:
    out: list[SweepFinding] = [SweepFinding(check, "fail", d) for d in fails]
    out.extend(SweepFinding(check, "warn", d) for d in warns)
    if not out:
        out.append(SweepFinding(check, "pass", f"{n_checked} {unit} checked"))
    return out


def _check_heel_toe_lateral(data: BlueOxExportData) -> list[SweepFinding]:
    """bro_time guard: wherever heel AND toe are populated, the geodesic
    heel->toe span must be consistent with the stated producing lateral.
    A 3-vertex stick's mid-lateral "heel" halves the span — flagged at
    >40% mismatch. U-turn inventory rows sum both legs (each leg is
    ~half the producing lateral by construction)."""
    fails: list[str] = []
    n = 0
    for area, inv in _labeled_inventory(data):
        legs = _inv_legs(inv)
        lat_ft = float(inv.producing_lateral_ft)
        if not legs or lat_ft <= 0:
            continue
        span_ft = sum(_geodesic_ft(hlon, hlat, tlon, tlat) for _leg, hlon, hlat, tlon, tlat in legs)
        n += 1
        if abs(span_ft - lat_ft) / lat_ft > HEEL_TOE_MISMATCH_FRAC:
            fails.append(
                f"inventory {area}/{inv.well_name or '?'}: heel->toe spans "
                f"{span_ft:,.0f} ft vs producing_lateral_ft {lat_ft:,.0f} "
                f"({span_ft / lat_ft:.0%}) — mid-lateral heel signature"
            )
    for z in data.zones:
        h_lon = _header_index(z.analog_headers, "heel_lon")
        h_lat = _header_index(z.analog_headers, "heel_lat")
        t_lon = _header_index(z.analog_headers, "toe_lon")
        t_lat = _header_index(z.analog_headers, "toe_lat")
        lat_i = _lateral_col_index(z.analog_headers)
        if h_lon is None or h_lat is None or t_lon is None or t_lat is None or lat_i is None:
            continue  # no geo block on this sheet (legacy shape) — nothing to check
        api_i = _api_col_index(z.analog_headers)
        for row in z.analog_rows:
            cells = [_num(row[i]) if i < len(row) else None for i in (h_lon, h_lat, t_lon, t_lat)]
            lat_ft2 = _num(row[lat_i]) if lat_i < len(row) else None
            if any(c is None for c in cells) or lat_ft2 is None or lat_ft2 <= 0:
                continue  # blank heel cells (non-4-vertex stick) are sanctioned
            hlon, hlat, tlon, tlat = (float(c) for c in cells if c is not None)
            span_ft = _geodesic_ft(hlon, hlat, tlon, tlat)
            n += 1
            if abs(span_ft - lat_ft2) / lat_ft2 > HEEL_TOE_MISMATCH_FRAC:
                fails.append(
                    f"analog {z.zone_name}/{row[api_i]}: heel->toe spans "
                    f"{span_ft:,.0f} ft vs lateral {lat_ft2:,.0f} "
                    f"({span_ft / lat_ft2:.0%}) — mid-lateral heel signature"
                )
    return _findings("heel_toe_lateral_consistency", n, "heel/toe rows", fails, [])


def _check_azimuth_spread(data: BlueOxExportData) -> list[SweepFinding]:
    """toucan guard: PDP rows carry each well's OWN as-built bearing —
    multiple PDPs in a unit sharing one lateral_azimuth_deg (within
    0.1° axially) means a single value was stamped across the unit.
    Planned-only units are exempt: generated wells legitimately share
    the scenario frame azimuth. A unit whose single PDP matches every
    planned row is only WARNed — the frame azimuth can legitimately
    derive from a kept existing stick (rule-16 trust order)."""
    fails: list[str] = []
    warns: list[str] = []
    units: dict[str, list[tuple[str, float]]] = {}
    for _area, inv in _labeled_inventory(data):
        az = inv.lateral_azimuth_deg
        if inv.dsu_id is not None and az is not None:
            units.setdefault(inv.dsu_id, []).append((inv.category, float(az)))
    n = 0
    for dsu, rows in sorted(units.items()):
        vals = [az for _cat, az in rows]
        pdp_vals = [az for cat, az in rows if cat == "PDP"]
        if len(vals) < 2:
            continue
        n += 1
        if len(pdp_vals) >= 2 and _axial_spread_deg(pdp_vals) < AZIMUTH_SPREAD_MIN_DEG:
            fails.append(
                f"{dsu}: {len(pdp_vals)} PDP rows share one lateral_azimuth_deg "
                f"(axial spread {_axial_spread_deg(pdp_vals):.3f}° < "
                f"{AZIMUTH_SPREAD_MIN_DEG}°) — as-built bearings stamped with a "
                "single value (toucan signature)"
            )
        elif len(pdp_vals) == 1 and _axial_spread_deg(vals) < AZIMUTH_SPREAD_MIN_DEG:
            warns.append(
                f"{dsu}: the unit's only PDP row carries the same "
                "lateral_azimuth_deg as every planned row (axial spread < "
                f"{AZIMUTH_SPREAD_MIN_DEG}°) — verify the PDP bearing is its "
                "own as-built value, not the frame azimuth"
            )
    return _findings("azimuth_spread", n, "multi-row units", fails, warns)


def _check_coordinate_order(data: BlueOxExportData) -> list[SweepFinding]:
    """§8 guard: every emitted pair is lon-first. Permian lon/lat live
    in disjoint ranges, so a swapped pair fails both bounds at once.
    Covers inventory heel/toe, dsu_meta origins, and the analog sheets'
    surface/heel/toe columns."""
    fails: list[str] = []
    n = 0

    def probe(label: str, lon: float | None, lat: float | None) -> None:
        nonlocal n
        for axis, v, (lo, hi) in (
            ("lon", lon, PERMIAN_LON_RANGE),
            ("lat", lat, PERMIAN_LAT_RANGE),
        ):
            if v is None:
                continue
            n += 1
            if not lo <= float(v) <= hi:
                fails.append(
                    f"{label}: {axis}={float(v):.4f} outside Permian {axis} "
                    f"range [{lo}, {hi}] — check lon-first ordering (§8)"
                )

    for area, inv in _labeled_inventory(data):
        who = f"inventory {area}/{inv.well_name or '?'}"
        probe(f"{who} heel_a", inv.heel_a_lon, inv.heel_a_lat)
        probe(f"{who} toe_a", inv.toe_a_lon, inv.toe_a_lat)
        probe(f"{who} heel_b", inv.heel_b_lon, inv.heel_b_lat)
        probe(f"{who} toe_b", inv.toe_b_lon, inv.toe_b_lat)
    for frame in data.dsu_meta:
        probe(f"dsu_meta {frame.dsu_id} origin", frame.origin_lon, frame.origin_lat)
    for z in data.zones:
        api_i = _api_col_index(z.analog_headers)
        for lon_name, lat_name in (
            ("surface_lon", "surface_lat"),
            ("heel_lon", "heel_lat"),
            ("toe_lon", "toe_lat"),
        ):
            lon_i = _header_index(z.analog_headers, lon_name)
            lat_i = _header_index(z.analog_headers, lat_name)
            if lon_i is None or lat_i is None:
                continue
            for row in z.analog_rows:
                probe(
                    f"analog {z.zone_name}/{row[api_i]} {lon_name.removesuffix('_lon')}",
                    _num(row[lon_i]) if lon_i < len(row) else None,
                    _num(row[lat_i]) if lat_i < len(row) else None,
                )
    return _findings("coordinate_order", n, "coordinates", fails, [])


def _check_offset_reproducibility(data: BlueOxExportData) -> list[SweepFinding]:
    """§6 guard: every gunbarrel_offset_ft (and _b_ft) must reproduce
    from dsu_meta.azimuth_deg + origin — the signed projection of the
    leg midpoint onto the axis 90° clockwise of the folded azimuth —
    to within ~1 ft. This holds for ALL rows including adopted wells
    whose own bearing differs from the frame (§10: one axis per unit).
    Offsets that CAN'T be verified (incomplete frame, legacy rows with
    no leg endpoints) warn rather than fail — absence of evidence, not
    a detected defect."""
    fails: list[str] = []
    warns: list[str] = []
    frames = {f.dsu_id: f for f in data.dsu_meta}
    incomplete_frames: set[str] = set()
    no_endpoints: dict[str, int] = {}
    n = 0
    for area, inv in _labeled_inventory(data):
        if inv.dsu_id is None:
            continue
        legs: dict[str, tuple[str, float, float, float, float] | None] = {"a": None, "b": None}
        for parsed_leg in _inv_legs(inv):
            legs[parsed_leg[0]] = parsed_leg
        for col, offset, leg_key in (
            ("gunbarrel_offset_ft", inv.gunbarrel_offset_ft, "a"),
            ("gunbarrel_offset_b_ft", inv.gunbarrel_offset_b_ft, "b"),
        ):
            if offset is None:
                continue
            leg = legs[leg_key]
            if leg is None:
                no_endpoints[inv.dsu_id] = no_endpoints.get(inv.dsu_id, 0) + 1
                continue
            frame = frames.get(inv.dsu_id)
            if (
                frame is None
                or frame.azimuth_deg is None
                or frame.origin_lon is None
                or frame.origin_lat is None
            ):
                incomplete_frames.add(inv.dsu_id)
                continue
            expected = _projected_offset_ft(
                leg[1],
                leg[2],
                leg[3],
                leg[4],
                frame.azimuth_deg,
                frame.origin_lon,
                frame.origin_lat,
            )
            n += 1
            if abs(expected - float(offset)) > OFFSET_REPRO_TOL_FT:
                fails.append(
                    f"inventory {area}/{inv.well_name or '?'}: {col}="
                    f"{float(offset):,.1f} ft but the §6 projection from "
                    f"dsu_meta[{inv.dsu_id}] gives {expected:,.1f} ft "
                    f"(Δ {abs(expected - float(offset)):,.1f} ft > "
                    f"{OFFSET_REPRO_TOL_FT} ft)"
                )
    for dsu in sorted(incomplete_frames):
        warns.append(
            f"{dsu}: dsu_meta frame is incomplete (blank azimuth/origin) — "
            "its §6 offsets cannot be verified or reproduced downstream"
        )
    for dsu, count in sorted(no_endpoints.items()):
        warns.append(
            f"{dsu}: {count} offset(s) carried without leg heel/toe endpoints "
            "(legacy narvi save?) — not independently verifiable"
        )
    return _findings("offset_reproducibility", n, "offsets", fails, warns)


def _check_vector_sanity(data: BlueOxExportData) -> list[SweepFinding]:
    """Zone-sheet vector sanity. Negative/non-finite values are already
    hard build-refusals in ``_validate``; here: (a) an all-zero
    delivered oil/gas stream is a FAIL (a delivered level must carry
    volume — ngl's deliberate all-zero is exempt by construction),
    all-zero water WARNs (optional column — drop it instead); (b) peak
    monthly magnitudes above the loose normalization ceilings WARN —
    deliberately far above physical rates so they flag basis mistakes
    (e.g. pre-multiplied laterals), never real curves."""
    fails: list[str] = []
    warns: list[str] = []
    levels = _delivered_levels(data)
    n = 0
    for z in data.zones:
        scale = 1.0 if z.normalization_basis == "per_1000_lateral_ft" else LATERAL_MAX_FT / 1000.0
        checks: list[tuple[str, str]] = [(lv, s) for lv in levels for s in ("oil", "gas")]
        if "water" in z.volumes.get("P50", {}):
            checks.append(("P50", "water"))
        for lv, stream in checks:
            vec = z.volumes.get(lv, {}).get(stream)
            if vec is None or not vec:
                continue  # missing levels are _validate's hard failure
            vals = [float(v) for v in vec]
            n += 1
            peak = max(vals)
            if peak <= 0.0:
                msg = f"{z.zone_name}: {stream} {lv} is all-zero"
                if stream == "water":
                    warns.append(msg + " — water is optional; drop the column instead")
                else:
                    fails.append(msg + " — a delivered stream must carry volume")
                continue
            ceiling = _STREAM_MONTH_CEILING_PER_1000FT[stream] * scale
            if peak > ceiling:
                warns.append(
                    f"{z.zone_name}: {stream} {lv} peak month {peak:,.0f} exceeds "
                    f"the loose {z.normalization_basis} ceiling {ceiling:,.0f} — "
                    "check the normalization basis (pre-multiplied laterals?)"
                )
    return _findings("vector_sanity", n, "stream vectors", fails, warns)


def presend_sweep(data: BlueOxExportData) -> tuple[SweepFinding, ...]:
    """Value-level pre-send sweep — the automated half of the
    blueox-curve-drop skill's §7 checklist.

    Returns the full structured report (every check contributes at
    least one finding; "pass" findings carry the count of rows
    actually inspected, so an accidentally-empty check is visible).
    :func:`build_blueox_workbook` runs this after ``_validate`` and
    refuses the build on any "fail"; "warn" findings are surfaced to
    callers via this function for the human sweep."""
    out: list[SweepFinding] = []
    out.extend(_check_heel_toe_lateral(data))
    out.extend(_check_azimuth_spread(data))
    out.extend(_check_coordinate_order(data))
    out.extend(_check_offset_reproducibility(data))
    out.extend(_check_vector_sanity(data))
    return tuple(out)


# ============================ sheet writers ============================


def _bold_row(ws: Any, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).font = _BOLD


def _zone_sheet_columns(z: ZoneData, levels: Sequence[str]) -> list[tuple[str, str, str]]:
    """(header, level, stream) triples in contract column order."""
    cols: list[tuple[str, str, str]] = [
        ("oil_bbl", "P50", "oil"),
        ("gas_mcf", "P50", "gas"),
        ("ngl_bbl", "P50", "ngl"),
    ]
    if "water" in z.volumes.get("P50", {}):
        cols.append(("water_bbl", "P50", "water"))
    for lv in levels:
        if lv == "P50":
            continue
        suffix = lv.lower()  # "P10" -> "p10"
        cols.append((f"oil_bbl_{suffix}", lv, "oil"))
        cols.append((f"gas_mcf_{suffix}", lv, "gas"))
        cols.append((f"ngl_bbl_{suffix}", lv, "ngl"))
    return cols


def _write_zone_sheet(ws: Any, z: ZoneData, data: BlueOxExportData) -> None:
    levels = _delivered_levels(data)
    cols = _zone_sheet_columns(z, levels)
    ws.append([c[0] for c in cols])
    _bold_row(ws, 1, len(cols))

    vectors: list[Sequence[float]] = []
    zero = [0.0] * data.curve_months
    for _header, lv, stream in cols:
        if stream == "ngl":
            vectors.append(zero)  # amendment: Blue Ox derives NGL via yield
        else:
            vectors.append(z.volumes[lv][stream])
    for i in range(data.curve_months):
        ws.append([float(vec[i]) for vec in vectors])

    ws.freeze_panes = "A2"
    for col_idx in range(1, len(cols) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].number_format = "#,##0.0"
        ws.column_dimensions[letter].width = 14


def _write_meta(ws: Any, data: BlueOxExportData) -> None:
    ws.append(["area", "normalization_basis", "reserve_category"])
    _bold_row(ws, 1, 3)
    for z in data.zones:
        ws.append([z.zone_name, z.normalization_basis, z.reserve_category])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18


def _all_inventory_rows(data: BlueOxExportData) -> list[InventoryRow]:
    return [inv for z in data.zones for inv in z.inventory] + [
        inv for _area, inv in data.pdp_context_rows
    ]


def _write_inventory(ws: Any, data: BlueOxExportData) -> None:
    all_rows = _all_inventory_rows(data)
    # well_name is optional; include the column only when every row has
    # one (contract formatting rule: no blank cells mid-column).
    all_named = all(inv.well_name for inv in all_rows)
    # Geometry columns (2026-07-27 amendment) ride along whenever any
    # row carries them — strictly appended after the v1 columns so a
    # lenient loader keeps working unchanged. Blank cells inside the
    # block are sanctioned (e.g. `_b` legs on single-lateral wells).
    with_geo = any(inv.dsu_id is not None for inv in all_rows)
    header = ["area", "category", "producing_lateral_ft", "drilled_lateral_ft"]
    if all_named:
        header.append("well_name")
    if with_geo:
        header.extend(_INVENTORY_GEO_COLS)
    ws.append(header)
    _bold_row(ws, 1, len(header))

    def _row(area: str, inv: InventoryRow) -> None:
        row: list[Any] = [
            area,
            inv.category,
            float(inv.producing_lateral_ft),
            float(inv.drilled_lateral_ft),
        ]
        if all_named:
            row.append(inv.well_name)
        if with_geo:
            row.extend(getattr(inv, col) for col in _INVENTORY_GEO_COLS)
        ws.append(row)

    for z in data.zones:
        for inv in z.inventory:
            _row(z.zone_name, inv)
    # Existing producers with no zone here: area = the bench code (the
    # one sanctioned mismatch with the zone-sheet names; PDP rows only).
    for area, inv in data.pdp_context_rows:
        _row(area, inv)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
    for letter in ("C", "D"):
        ws.column_dimensions[letter].number_format = "#,##0"
        ws.column_dimensions[letter].width = 20
    ws.column_dimensions["E"].width = 26


def _write_novi_comparison(ws: Any, data: BlueOxExportData) -> None:
    """Long format: one row per zone per 30-day period — the median
    Novi ML per-1,000-ft period volumes, IP-aligned. Zones with no
    eligible sticks contribute no rows (declared on the meta sheet)."""
    ws.append(["area", "month", "oil_bbl", "gas_mcf", "water_bbl"])
    _bold_row(ws, 1, 5)
    for c in data.novi_comparison:
        for i in range(len(c.oil_bbl)):
            ws.append(
                [
                    c.zone_name,
                    i + 1,
                    float(c.oil_bbl[i]),
                    float(c.gas_mcf[i]),
                    float(c.water_bbl[i]),
                ]
            )
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 28
    for letter in ("C", "D", "E"):
        ws.column_dimensions[letter].number_format = "#,##0.0"
        ws.column_dimensions[letter].width = 14


_NOVI_META_COLS: tuple[str, ...] = (
    "area",
    "n_sticks",
    "n_self",
    "n_neighborhood",
    "n_pud",
    "n_res",
    "n_wells_no_set",
    "radius_m",
    "lateral_tol",
    "intel_vintage",
    "low_n_flag",
    "stale_vintage_flag",
    "tc_risked",
)


def _write_novi_comparison_meta(ws: Any, data: BlueOxExportData) -> None:
    ws.append(list(_NOVI_META_COLS))
    _bold_row(ws, 1, len(_NOVI_META_COLS))
    for c in data.novi_comparison:
        ws.append(
            [
                c.zone_name,
                c.n_sticks,
                c.n_self,
                c.n_neighborhood,
                c.n_pud,
                c.n_res,
                c.n_wells_no_set,
                c.radius_m,
                c.lateral_tol,
                c.intel_vintage,
                c.low_n,
                c.stale_vintage,
                c.tc_risked,
            ]
        )
    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, len(_NOVI_META_COLS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


def _write_dsu_meta(ws: Any, data: BlueOxExportData) -> None:
    ws.append(["dsu_id", "azimuth_deg", "origin_lon", "origin_lat"])
    _bold_row(ws, 1, 4)
    for frame in data.dsu_meta:
        ws.append(
            [
                frame.dsu_id,
                frame.azimuth_deg,
                frame.origin_lon,
                frame.origin_lat,
            ]
        )
    ws.column_dimensions["A"].width = 34
    for letter in ("B", "C", "D"):
        ws.column_dimensions[letter].width = 14


def _write_analog_sheet(ws: Any, z: ZoneData) -> None:
    """Contract §1.4 shape: a ``per_well_summary`` marker in column A,
    headers on the next row, then one row per analog well."""
    ws.append(["per_well_summary"])
    ws.cell(row=1, column=1).font = _BOLD
    ws.append(list(z.analog_headers))
    _bold_row(ws, 2, len(z.analog_headers))
    for row in z.analog_rows:
        ws.append(list(row))
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 26
    for col_idx in range(3, len(z.analog_headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14


def _write_analog_production(ws: Any, data: BlueOxExportData) -> None:
    ws.append(list(data.production_headers))
    _bold_row(ws, 1, len(data.production_headers))
    for row in data.production_rows:
        ws.append(list(row))
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    for col_idx in range(3, len(data.production_headers) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].number_format = "#,##0.0"
        ws.column_dimensions[letter].width = 12


_CURVE_PARAM_COLS: tuple[str, ...] = (
    "area",
    "stream",
    "level",
    "qi",
    "qi_units",
    "qi_basis",
    "b_factor",
    "di",
    "di_convention",
    "dmin",
    "risk_mult",
    "notes",
)
QI_BASIS = "fitted_qi"
DI_CONVENTION = "nominal_annual"


def _write_curve_params(ws: Any, data: BlueOxExportData) -> None:
    ws.append(list(_CURVE_PARAM_COLS))
    _bold_row(ws, 1, len(_CURVE_PARAM_COLS))
    level_order = {lv: i for i, lv in enumerate(BLUEOX_LEVEL_ORDER)}
    for z in data.zones:
        rows = sorted(
            z.curve_params,
            key=lambda r: (str(r.get("stream")), level_order.get(str(r.get("level")), 99)),
        )
        for r in rows:
            ws.append(
                [
                    z.zone_name,
                    r.get("stream"),
                    r.get("level"),
                    r.get("qi"),
                    r.get("qi_units"),
                    # Per-row basis (2026-07-24 risking amendment): a risked
                    # row declares fitted_qi_risked + its multiplier; rows
                    # from callers predating the amendment fall back to the
                    # unrisked constants.
                    r.get("qi_basis", QI_BASIS),
                    r.get("b_factor"),
                    r.get("di"),
                    DI_CONVENTION,
                    r.get("dmin"),
                    r.get("risk_mult", 1.0),
                    r.get("notes"),
                ]
            )
    ws.column_dimensions["A"].width = 28
    for letter in ("D", "G", "H", "J", "K"):
        ws.column_dimensions[letter].number_format = "0.000"
    for col_idx in range(2, len(_CURVE_PARAM_COLS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    ws.column_dimensions[get_column_letter(len(_CURVE_PARAM_COLS))].width = 48


def _write_manifest(ws: Any, data: BlueOxExportData) -> None:
    block_a: list[tuple[str, Any]] = [
        ("deal_codename", data.codename),
        ("export_date", data.export_date.isoformat()),
        ("source_system", data.source_system),
        ("governing_export", data.governing_export),
        ("percentile_orientation", PERCENTILE_ORIENTATION),
        ("gas_basis", GAS_BASIS),
        ("ngl_basis", NGL_BASIS),
        ("risking", data.risking),
        ("curve_months", data.curve_months),
        ("production_history_through", data.production_history_through),
        ("curve_params_source", data.curve_params_source),
        ("prepared_by", data.prepared_by),
    ]
    if data.history_exceptions:
        block_a.append(
            (
                "analog_history_exceptions",
                ", ".join(data.history_exceptions) + " (no monthly history in source)",
            )
        )
    if data.inventory_exclusions:
        block_a.append(
            (
                "inventory_benches_excluded",
                "; ".join(data.inventory_exclusions),
            )
        )
    # Scenario-scoped zones declare their DSU subset (2026-07-24
    # amendment, optional key — absent on unscoped zones so legacy
    # drops are byte-identical).
    for z in data.zones:
        if z.scenario_scope:
            block_a.append(
                (
                    f"zone_scenario_scope[{z.zone_name}]",
                    "; ".join(z.scenario_scope),
                )
            )
    block_a.append(
        (
            "inventory_category_basis",
            "PDP = existing producer (context only, not counted); "
            "PUD = pdp_count_3mi >= 3; UPSIDE = pdp_count_3mi <= 2 or unscored; "
            "narvi user overrides win",
        )
    )
    # TC-vs-Novi benchmark declarations (2026-07-27 amendment) — only
    # when the comparison ships, so legacy drops stay byte-identical.
    if data.novi_comparison:
        radius = data.novi_comparison[0].radius_m
        # ll tolerance is per-basin (2026-07-30 amendment). Single-basin
        # deals — the practical case — declare the one value; a
        # mixed-basin deal defers to the per-zone lateral_tol column in
        # novi_comparison_meta rather than declaring one zone's value
        # for all.
        tols = {z.lateral_tol for z in data.novi_comparison}
        tol: float | str = tols.pop() if len(tols) == 1 else "per_zone_see_novi_comparison_meta"
        block_a.extend(
            [
                ("novi_intel_vintage", data.novi_intel_vintage),
                ("novi_selection_radius_m", radius),
                ("novi_selection_lateral_tol", tol),
                ("novi_alignment", NOVI_ALIGNMENT),
                ("novi_rate_to_volume_days", NOVI_RATE_TO_VOLUME_DAYS),
            ]
        )
    for key, value in block_a:
        ws.append([key, value])
        ws.cell(row=ws.max_row, column=1).font = _BOLD

    ws.append([])
    ws.append([])

    # Block B — reconciliation targets, computed from the SAME vectors /
    # inventory rows the sheets were written from (contract: "compute
    # Block B from the final sheets, not from the source system").
    header = [
        "area",
        "eur_oil_bbl",
        "eur_gas_mcf",
        "eur_ngl_bbl",
        "gross_locations",
        "avg_producing_lateral_ft",
        "avg_drilled_lateral_ft",
    ]
    ws.append(header)
    _bold_row(ws, ws.max_row, len(header))
    for z in data.zones:
        # PDP rows are display-only context — the valued location count
        # and lateral means cover PUD/UPSIDE rows exclusively.
        counted = [i for i in z.inventory if i.category != "PDP"]
        n_inv = len(counted)
        avg_prod = sum(float(i.producing_lateral_ft) for i in counted) / n_inv if n_inv else 0.0
        avg_drill = sum(float(i.drilled_lateral_ft) for i in counted) / n_inv if n_inv else 0.0
        ws.append(
            [
                z.zone_name,
                sum(float(v) for v in z.volumes["P50"]["oil"]),
                sum(float(v) for v in z.volumes["P50"]["gas"]),
                0.0,  # ngl_bbl delivered all-zero (amendment)
                n_inv,
                avg_prod,
                avg_drill,
            ]
        )
    ws.column_dimensions["A"].width = 34
    for col_idx in range(2, len(header) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].number_format = "#,##0.0"
        ws.column_dimensions[letter].width = 20
    ws.column_dimensions["B"].width = 44  # Block A values share column B


# ============================ entry point ============================


def build_blueox_workbook(data: BlueOxExportData) -> bytes:
    """Validate against the contract and emit the workbook bytes.

    Raises :class:`BlueOxContractError` (with every violation listed)
    rather than emitting a workbook that would bounce off the Blue Ox
    acceptance gate. After the structural contract checks, the value-
    level pre-send sweep runs (:func:`presend_sweep`); any "fail"
    finding also blocks the build — the sweep's "warn" findings do
    not, but callers should surface them for the human checklist.
    """
    _validate(data)
    sweep_failures = [f for f in presend_sweep(data) if f.status == "fail"]
    if sweep_failures:
        raise BlueOxContractError(
            "Blue Ox pre-send sweep failures:\n- "
            + "\n- ".join(f"[{f.check}] {f.detail}" for f in sweep_failures)
        )

    wb = Workbook()
    wb.remove(wb.active)

    for z in data.zones:
        _write_zone_sheet(wb.create_sheet(z.zone_name), z, data)
    _write_meta(wb.create_sheet("meta"), data)
    _write_inventory(wb.create_sheet("inventory"), data)
    if data.dsu_meta:
        _write_dsu_meta(wb.create_sheet("dsu_meta"), data)
    for z in data.zones:
        _write_analog_sheet(wb.create_sheet(f"{z.zone_name} meta"), z)
    _write_analog_production(wb.create_sheet("analog_production"), data)
    _write_curve_params(wb.create_sheet("curve_params"), data)
    if data.novi_comparison:
        _write_novi_comparison(wb.create_sheet("novi_comparison"), data)
        _write_novi_comparison_meta(wb.create_sheet("novi_comparison_meta"), data)
    _write_manifest(wb.create_sheet("manifest"), data)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
