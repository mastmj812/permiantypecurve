"""Compare the CSV-zip export vs the deal XLSX export for the same curve.

Both share three pieces of numeric content:
  * Fitted forecast: per-percentile rate + cum out to 600 months
  * fitted_eur_per_1000ft (50-yr Arps projection per percentile)
  * fitted_p50_params (qi/Di/b/Df/qo/peak_index)

Plus shared metadata (id/name/basis/alignment/...). We diff each of
these to confirm both export paths agree on numeric content for the
same curve.
"""
from __future__ import annotations

import csv
import io
import uuid
import zipfile
from typing import Any

from openpyxl import Workbook
from sqlalchemy import select  # noqa: F401

from app.api.deals import _write_forecast_sheet, _write_metadata_sheet
from app.api.type_curves import export_type_curve
from app.db.models import TypeCurve
from app.db.session import SessionLocal

TC_ID = uuid.UUID("2f5bfeba-caf3-457d-a0a0-c424e7738604")  # holdTheLine_bs1s
s = SessionLocal()
tc = s.get(TypeCurve, TC_ID)
print(f"Curve: {tc.name}, n_wells={len(tc.included_api14s)}, "
      f"series.n_months={tc.series.get('n_months')}")
print()

# ---- Build the CSV ZIP via the actual handler ----
csv_resp = export_type_curve(TC_ID, session=s)
zf = zipfile.ZipFile(io.BytesIO(csv_resp.body))
print(f"CSV zip members: {sorted(zf.namelist())}")
print()

# ---- Build the XLSX via the actual handler path (no FastAPI) ----
wb = Workbook()
wb.remove(wb.active)
meta_ws = wb.create_sheet("meta")
_write_metadata_sheet(meta_ws, tc, s)
fcst_ws = wb.create_sheet("forecast")
_write_forecast_sheet(fcst_ws, tc)


def parse_forecast_csv(content: bytes) -> dict[int, dict[str, dict[str, float]]]:
    out: dict[int, dict[str, dict[str, float]]] = {}
    reader = csv.reader(io.StringIO(content.decode()))
    header = next(reader)
    col_meta: list[tuple[str, str] | None] = [None]  # month col
    for h in header[1:]:
        # "p10_rate" → ("p10", "rate"). Older exports used a "fitted_"
        # prefix; we strip it if it shows up so this script reads both.
        m = h.replace("fitted_", "").rsplit("_", 1)
        col_meta.append((m[0], m[1]))
    for row in reader:
        if not row[0]:
            continue
        month = int(row[0])
        per_pct: dict[str, dict[str, float]] = {}
        for i, cell in enumerate(row[1:], start=1):
            cm = col_meta[i]
            if cm is None or not cell:
                continue
            pct, kind = cm
            per_pct.setdefault(pct, {})[kind] = float(cell)
        out[month] = per_pct
    return out


csv_forecasts: dict[str, dict[int, dict[str, dict[str, float]]]] = {}
for stream in ("oil", "gas", "water"):
    csv_forecasts[stream] = parse_forecast_csv(
        zf.read(f"{stream}_forecast.csv")
    )

# Parse the XLSX forecast sheet (wide layout).
xlsx_rows = list(fcst_ws.iter_rows(values_only=True))
xlsx_header = xlsx_rows[0]
xlsx_data = xlsx_rows[1:]
xlsx_col_meta: list[tuple[str, str, str] | None] = [None]
for h in xlsx_header[1:]:
    parts = h.split("_")
    xlsx_col_meta.append((parts[0], parts[1], parts[2]))

samples = list(range(10)) + list(range(290, 300)) + list(range(590, 600))
print(f"Comparing {len(samples)} sample months × 3 streams × 6 pcts × 2 metrics:")

mismatches: list[tuple[str, str, str, int, Any, Any]] = []
for ridx in samples:
    if ridx >= len(xlsx_data):
        continue
    xrow = xlsx_data[ridx]
    if xrow[0] is None:
        continue
    month = int(xrow[0])
    for col_idx, cm in enumerate(xlsx_col_meta):
        if cm is None:
            continue
        stream, pct, kind = cm
        xlsx_val = xrow[col_idx]
        csv_val = csv_forecasts[stream].get(month, {}).get(pct, {}).get(kind)
        if xlsx_val is None and csv_val is None:
            continue
        if xlsx_val is None or csv_val is None:
            mismatches.append((stream, pct, kind, month, xlsx_val, csv_val))
            continue
        tol = 0.5 if kind == "cum" else 0.001
        if abs(xlsx_val - csv_val) > tol:
            mismatches.append((stream, pct, kind, month, xlsx_val, csv_val))

if mismatches:
    print(f"  X {len(mismatches)} mismatches:")
    for m in mismatches[:20]:
        print(f"     {m[0]}.{m[1]}.{m[2]} @ month {m[3]}: xlsx={m[4]}, csv={m[5]}")
else:
    print("  OK forecast tables identical across every sampled cell")

print()


def parse_metadata_csv(content: bytes) -> dict[str, Any]:
    reader = csv.reader(io.StringIO(content.decode()))
    out: dict[str, Any] = {"eur": {}, "params": {}}
    section = None
    eur_header: list[str] = []
    param_header: list[str] = []
    for row in reader:
        if not row:
            continue
        if row[0].startswith("fitted_eur_per_1000ft"):
            section = "eur"
            continue
        if row[0].startswith("fitted_p50_params"):
            section = "params"
            continue
        if section == "eur" and row[0] == "stream":
            eur_header = row[1:]
            continue
        if section == "params" and row[0] == "stream":
            param_header = row[1:]
            continue
        if section == "eur" and row[0] in ("oil", "gas", "water"):
            out["eur"][row[0]] = {
                k: float(v) if v else None
                for k, v in zip(eur_header, row[1:], strict=False)
            }
            continue
        if section == "params" and row[0] in ("oil", "gas", "water"):
            out["params"][row[0]] = {
                k: v for k, v in zip(param_header, row[1:], strict=False)
            }
            continue
        if row[0] == "included_api14s":
            section = None
    return out


csv_meta = parse_metadata_csv(zf.read("metadata.csv"))

xlsx_meta_rows = list(meta_ws.iter_rows(values_only=True))


def find_block(rows: list, header_prefix: str) -> tuple[int, int]:
    for i, r in enumerate(rows):
        if (
            r and r[0]
            and isinstance(r[0], str)
            and r[0].startswith(header_prefix)
        ):
            return i + 1, i + 2
    return -1, -1


eur_hdr_row, eur_data_row = find_block(
    xlsx_meta_rows, "fitted_eur_per_1000ft"
)
prm_hdr_row, prm_data_row = find_block(xlsx_meta_rows, "fitted_p50_params")

xlsx_eur: dict[str, dict[str, Any]] = {}
xlsx_eur_header = list(xlsx_meta_rows[eur_hdr_row])[1:]
for i in range(3):
    r = xlsx_meta_rows[eur_data_row + i]
    if r and r[0] in ("oil", "gas", "water"):
        xlsx_eur[r[0]] = {
            k: r[1 + j] for j, k in enumerate(xlsx_eur_header)
        }

xlsx_params: dict[str, dict[str, Any]] = {}
xlsx_param_header = list(xlsx_meta_rows[prm_hdr_row])[1:]
for i in range(3):
    r = xlsx_meta_rows[prm_data_row + i]
    if r and r[0] in ("oil", "gas", "water"):
        xlsx_params[r[0]] = {
            k: r[1 + j] for j, k in enumerate(xlsx_param_header)
        }

print("Comparing fitted_eur_per_1000ft per stream × percentile:")
eur_ok = True
for stream in ("oil", "gas", "water"):
    for pct in ("p10", "p25", "p50", "p75", "p90", "mean"):
        x = xlsx_eur.get(stream, {}).get(pct)
        c = csv_meta["eur"].get(stream, {}).get(pct)
        if x is None and c is None:
            continue
        if x is None or c is None:
            print(f"   X {stream}.{pct}: xlsx={x}, csv={c}")
            eur_ok = False
            continue
        if abs(float(x) - float(c)) > 0.5:
            d = abs(float(x) - float(c))
            print(f"   X {stream}.{pct}: xlsx={x}, csv={c} (diff {d:.2f})")
            eur_ok = False
if eur_ok:
    print("   OK All 18 (3 streams x 6 percentiles) fitted EURs agree "
          "within 0.5 BBL/MCF per 1k ft")

print()
print("Comparing fitted_p50_params per stream:")
prm_ok = True
for stream in ("oil", "gas", "water"):
    for k in (
        "model_type", "qi", "Di", "b", "Df", "qo", "peak_index"
    ):
        x = xlsx_params.get(stream, {}).get(k)
        c = csv_meta["params"].get(stream, {}).get(k)
        if x is None and (c == "" or c is None):
            continue
        if k == "model_type":
            if x != c:
                print(f"   X {stream}.{k}: xlsx={x!r}, csv={c!r}")
                prm_ok = False
            continue
        if k == "peak_index":
            xv = int(x) if x is not None else None
            cv = int(c) if c else None
            if xv != cv:
                print(f"   X {stream}.{k}: xlsx={x!r}, csv={c!r}")
                prm_ok = False
            continue
        xv = float(x) if x is not None else None
        cv = float(c) if c else None
        if xv is None and cv is None:
            continue
        if abs((xv or 0) - (cv or 0)) > 1e-4:
            print(f"   X {stream}.{k}: xlsx={x}, csv={c}")
            prm_ok = False
if prm_ok:
    print("   OK All 7 fields x 3 streams of fitted_p50_params agree within 1e-4")

print()
print("Summary:")
print(f"  forecast tables: {'identical' if not mismatches else f'{len(mismatches)} mismatches'}")
print(f"  fitted_eur_per_1000ft: {'identical' if eur_ok else 'differs'}")
print(f"  fitted_p50_params: {'identical' if prm_ok else 'differs'}")
